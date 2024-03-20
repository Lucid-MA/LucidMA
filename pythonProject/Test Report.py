import os
import win32com.client
from openpyxl import load_workbook

def generate_report():
    # Variable Declarations
    template_path = r"S:\Lucid\Investment Committee & Risk\Nightly Risk Packets\Report Generator\Risk Packet Template - streamlined.xlsx"
    out_path_curr = r"S:\Lucid\Investment Committee & Risk\Nightly Risk Packets\test_report.xlsx"
    save_path = r"S:\Lucid\Investment Committee & Risk\Nightly Risk Packets\test_report_2.xlsx"
    consolidated_prime_path = r"S:\Lucid\Investment Committee & Risk\Nightly Risk Packets\Report Generator\pofo_reports\ConsolidatedPrime_20240109.xlsx"
    cash_ladder_report_path = r"S:\Lucid\Trading & Markets\Trading and Settlement Tools\Cash Ladder.xlsx"
    cash_summary_path = r"S:\Mandates\Operations\Daily Reconciliation\CashSummary.xlsx"
    var_overview_path = r"S:\Lucid\Investment Committee & Risk\VAR Workspace\Dynamic Overview.xlsx"
    report_date = "2024-03-17"  # Example date

    # Outlook Object Creation
    olApp = win32com.client.Dispatch("Outlook.Application")
    olMail = olApp.CreateItem(0)
    olMail.Recipients.Add("tony.hoang@lucidma.com")

    # Data Retrieval and Report Generation
    wbto = load_workbook(template_path)
    menupage = wbto["Main"]
    menupage["B6"] = report_date

    # VAR
    wbfrom = load_workbook(var_overview_path, read_only=True)
    wsto = wbto["VAR Overview"]
    # Copy data and charts from wbfrom to wsto
    # Close wbfrom

    # Consolidated Prime Fund Report
    wbfrom = load_workbook(consolidated_prime_path, read_only=True)
    wsto = wbto["Consolidated Prime"]
    # Copy data from wbfrom to wsto
    # Close wbfrom

    # Cash Ladder
    wbfrom = load_workbook(cash_ladder_report_path, read_only=True)
    wsto = wbto["Cash Ladder"]
    # Copy data from wbfrom to wsto
    # Close wbfrom

    # Cash Summary for Failing Trade Tab
    wbfrom = load_workbook(cash_summary_path, read_only=True)
    wsto = wbto["Cash Summary"]
    # Copy data from wbfrom to wsto
    # Close wbfrom

    # Email Notification
    all_checks = ""
    if all_checks:
        olMail.Subject = f"SOME CHECKS FAILED - Nightly Risk packet - {report_date}"
        olMail.HTMLBody = all_checks
    else:
        olMail.Subject = f"Nightly Risk packet - {report_date}"

    wbto.save(save_path)
    olMail.Attachments.Add(save_path)
    olMail.Send()

    # Cleanup
    os.remove(out_path_curr)
    os.remove(save_path)

if __name__ == "__main__":
    generate_report()