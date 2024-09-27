import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from sqlalchemy import create_engine


def Output_Engine(Results: dict, N, DATE):

    ledger_list_full = {
        "USG": ["Monthly"],
        "Prime": [
            "Master",
            "Monthly",
            "Custom1",
            "Quarterly1",
            "MonthlyIG",
            "QuarterlyX",
            "Q364",
            "2YIG",
        ],
    }

    folder_path = "VAR Results"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    for Fund in ledger_list_full:
        for ledger in ledger_list_full.get(Fund):
            if isinstance(Results.get(("N", Fund, ledger)), pd.DataFrame):
                wb = Workbook()
                wb = load_workbook(filename="Ledger Template.xlsx")
                ws1 = wb["Results"]

                ws1.cell(row=3, column=2).value = DATE
                ws1.cell(row=6, column=2).value = N

                if Fund == "USG":
                    ws1.cell(row=5, column=2).value = "Monthly"
                    ws1.cell(row=4, column=2).value = "USG Fund"
                else:
                    ws1.cell(row=5, column=2).value = ledger
                    ws1.cell(row=4, column=2).value = "Prime Fund"

                today_out = DATE.strftime("%m_%d_%Y")
                if Fund == "USG":
                    wb.save(os.path.join(folder_path, f"USG_Monthly_{today_out}.xlsx"))
                else:
                    wb.save(
                        os.path.join(
                            folder_path, f"Prime Fund_{ledger}_{today_out}.xlsx"
                        )
                    )
                wb.close()
                NormalStress = Results.get(("N", Fund, ledger))
                StressedStress = Results.get(("S", Fund, ledger))
                CriticStress = Results.get(("C", Fund, ledger))

                NormalStress = NormalStress.reindex(
                    np.roll(NormalStress.index, shift=1)
                )
                StressedStress = StressedStress.reindex(
                    np.roll(StressedStress.index, shift=1)
                )
                CriticStress = CriticStress.reindex(
                    np.roll(CriticStress.index, shift=1)
                )

                if Fund == "USG":
                    with pd.ExcelWriter(
                        os.path.join(folder_path, f"USG_Monthly_{today_out}.xlsx"),
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                    ) as writer:
                        writing(NormalStress, writer, StressedStress, CriticStress)
                else:
                    with pd.ExcelWriter(
                        os.path.join(
                            folder_path, f"Prime Fund_{ledger}_{today_out}.xlsx"
                        ),
                        mode="a",
                        engine="openpyxl",
                        if_sheet_exists="overlay",
                    ) as writer:
                        writing(NormalStress, writer, StressedStress, CriticStress)


def writing(NormalStress, writer, StressedStress, CriticStress):
    NormalStress.reset_index().to_excel(
        writer, sheet_name="Results", header=None, startrow=10, startcol=2, index=False
    )
    StressedStress.reset_index().to_excel(
        writer, sheet_name="Results", header=None, startrow=10, startcol=17, index=False
    )
    CriticStress.reset_index().to_excel(
        writer, sheet_name="Results", header=None, startrow=10, startcol=32, index=False
    )


# Create the database engine once, to be reused across function calls
conn_string = "postgresql://dbmasteruser:lnRz*(N_7aOf~7Hx6oRo8;,<vYp|~#PC@luciddb1.czojmxqfrx7k.us-east-1.rds.amazonaws.com/spiral_prod"
db = create_engine(conn_string)


#### UPLOAD TO MS SQL ####
import platform

# Configuration
DB_CONFIG = {
    "postgres": {
        "db_endpoint": "luciddb1.czojmxqfrx7k.us-east-1.rds.amazonaws.com",
        "db_port": "5432",
        "db_user": "dbmasteruser",
        "db_password": "lnRz*(N_7aOf~7Hx6oRo8;,<vYp|~#PC",
        "db_name": "spiral_prod",
    },
    "sql_server_1": {
        "driver": "ODBC+Driver+17+for+SQL+Server",
        "server_mac": "172.31.0.10",
        "server_windows": "LUCIDSQL1",
        "database": "HELIXREPO_PROD_02",
        "trusted_connection": "yes",
        "user_mac": "Lucid\\tony.hoang",
        "user_windows": "tony.hoang",
        "password": os.getenv("MY_PASSWORD"),
    },
    "sql_server_2": {
        "driver": "ODBC+Driver+17+for+SQL+Server",
        "server_mac": "172.31.32.100",
        "server_windows": "LUCIDSQL2",
        "database": "Prod1",
        "trusted_connection": "yes",
        "user_mac": "Lucid\\tony.hoang",
        "user_windows": "tony.hoang",
        "password": "Ar0undthe$un",
    },
}


def get_database_engine(db_type):
    if db_type == "postgres":
        database_url = f"postgresql://{DB_CONFIG['postgres']['db_user']}:{DB_CONFIG['postgres']['db_password']}@{DB_CONFIG['postgres']['db_endpoint']}:{DB_CONFIG['postgres']['db_port']}/{DB_CONFIG['postgres']['db_name']}"
        return create_engine(database_url)

    elif db_type.startswith("sql_server"):
        if platform.system() == "Darwin":  # macOS
            conn_str = f"mssql+pymssql://{DB_CONFIG[db_type]['user_mac']}:{DB_CONFIG[db_type]['password']}@{DB_CONFIG[db_type]['server_mac']}/{DB_CONFIG[db_type]['database']}"
            return create_engine(conn_str)

        elif platform.system() == "Windows":
            conn_str = (
                f"mssql+pyodbc://{DB_CONFIG[db_type]['user_windows']}:{DB_CONFIG[db_type]['password']}@"
                f"{DB_CONFIG[db_type]['server_windows']}/{DB_CONFIG[db_type]['database']}?driver={DB_CONFIG[db_type]['driver']}&Trusted_Connection={DB_CONFIG[db_type]['trusted_connection']}"
            )
            return create_engine(conn_str)

        else:
            raise Exception("Unsupported platform")


engine = get_database_engine("sql_server_2")


###################################################
def Output_SQL(TempDF, N, StressRun, Fund, ledger, TableName, DATE):
    conn = db.connect()
    try:
        TempDF["Fund"] = Fund
        TempDF["ledger"] = ledger
        TempDF["StressRun"] = StressRun
        TempDF["N"] = N
        TempDF["Date"] = DATE
        TempDF.loc["Total FUND", "Spd Diff"] = None
        # TempDF.to_sql(TableName, db, if_exists='append', schema='VAR_Model', index=True)
        output_excel_path = (
            "S:\Lucid\Investment Committee & Risk\VAR Workspace\Tony\Output"
        )
        file_name = f"{Fund}_{ledger}_{StressRun}.xlsx"
        # Combine the output path and file name
        full_output_path = os.path.join(output_excel_path, file_name)
        # Write TempDF to Excel
        TempDF.to_excel(full_output_path, index=False)
        # Upsert to table
        # To PostGres
        TempDF.to_sql(TableName, db, if_exists="append", schema="VAR_Model", index=True)
        # To SQL
        TempDF.to_sql("VAR_Results", con=engine, if_exists="append", index=False)
    except Exception as e:
        conn.close()
        return e
    finally:
        conn.close()


def SingleCounterpartyVAR(Results, counterparty, Fund, ledger):
    df_N = Results.get(("N", Fund, ledger)).loc[[counterparty]]
    df_S = Results.get(("S", Fund, ledger)).loc[[counterparty]]
    df_C = Results.get(("C", Fund, ledger)).loc[[counterparty]]
    return pd.concat([df_N, df_S, df_C])


def Overview(Results, DATE):
    today_out = DATE.strftime("%m_%d_%Y")

    # Define directory path for 'VAR Results' folder
    directory_path = "VAR Results"

    # Check if 'VAR Results' folder exists, create it if necessary
    if not os.path.exists(directory_path):
        os.makedirs(directory_path)

    # Create and save the output workbook
    wb = load_workbook(filename="Overview Template.xlsx")
    ws_summary = wb["Summary"]
    ws_summary.cell(row=2, column=1).value = DATE
    wb.save(os.path.join(directory_path, f"VaR_{today_out}_Overview.xlsx"))
    wb.close()

    excel_index = {
        "USG": {"Monthly": 3},
        "Prime": {
            "Master": 8,
            "Monthly": 28,
            "Custom1": 18,
            "Quarterly1": 38,
            "MonthlyIG": 23,
            "QuarterlyX": 43,
            "Q364": 33,
            "2YIG": 13,
        },
    }

    with pd.ExcelWriter(
        os.path.join(directory_path, f"VaR_{today_out}_Overview.xlsx"),
        mode="a",
        engine="openpyxl",
        if_sheet_exists="overlay",
    ) as writer:
        List_Max = {}
        VaR_list = [0.8, 0.9, 0.95, 0.98, 0.99, 0.995, 0.999, 0.9999, 0.99995, 0.99999]
        for StressRun, Fund, ledger in Results.keys():
            # Check if the starting row exists
            start_row = excel_index.get(Fund, {}).get(ledger)
            if start_row is None:
                continue

            # Create a dataframe from the results
            df = Results.get((StressRun, Fund, ledger)).loc[["Total FUND"]]

            # Write the dataframe to the output workbook
            i = {"N": 0, "S": 1, "C": 2}[StressRun]
            df.loc[:, df.columns != "Spd Diff"].to_excel(
                writer,
                sheet_name="Summary",
                header=None,
                startrow=start_row + i,
                startcol=2,
                index=False,
            )

            if StressRun == "C":
                df_max = Results.get((StressRun, Fund, ledger))
                df_max = df_max.loc[:, df_max.columns != "Total FUND"]
                param = 1
                while len(df_max.loc[df_max[VaR_list[param]] != 0]) == 0:
                    param += 1
                List_Max[(Fund, ledger)] = (
                    df_max.loc[df_max[VaR_list[6]] != 0].head(1).index[0]
                )

        for (Fund, ledger), counterparty in List_Max.items():
            SingleCounterparty_VaR = SingleCounterpartyVAR(
                Results, counterparty, Fund, ledger
            )
            start_row = excel_index.get(Fund, {}).get(ledger)
            if start_row is None:
                continue
            SingleCounterparty_VaR.loc[:, df.columns != "Spd Diff"].to_excel(writer, sheet_name="Summary", header=None, startrow=start_row, startcol=16, index=False)  # type: ignore

    wb = load_workbook(os.path.join(directory_path, f"VaR_{today_out}_Overview.xlsx"))
    ws_summary = wb["Summary"]
    for (Fund, ledger), counterparty in List_Max.items():
        ws_summary.cell(row=excel_index.get(Fund).get(ledger), column=16).value = counterparty  # type: ignore
    wb.save(os.path.join(directory_path, f"VaR_{today_out}_Overview.xlsx"))
    wb.close()
