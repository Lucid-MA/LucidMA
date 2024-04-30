# pull prices from JPPD and IDC into source file. if AM, gets cusips from helix. if PM, re-prices cusips already there
# (new cusips added throughout the day aren't to be included in PM price)
import time

# JJ Vulopas

import openpyxl as op
import pandas as pd
import requests, base64, http.client, os, subprocess
from datetime import datetime, timedelta
from pathlib import Path, PureWindowsPath
from pandas.tseries.offsets import BDay
import pymssql
from pymssql import Error
import xlwings as xw
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
import ast
import datetime as dt
import pandas as pd
import time
from Utils.Common import get_file_path

# currdest = PureWindowsPath(Path("S:/Lucid/Data/Bond Data/Price Source/Price_Source.xlsx"))
# savedest = PureWindowsPath(Path("S:/Lucid/Data/Bond Data/Price Source/Archives/Price_Source_" + datetime.now().strftime("%m_%d_%Y") + ".xlsx"))
currdest = get_file_path(r"S:/Lucid/Data/Bond Data/Price Source/Price_Source.xlsx")
savedest = get_file_path(r"S:/Lucid/Data/Bond Data/Price Source/Archives/Price_Source_" + datetime.now().strftime("%m_%d_%Y") + ".xlsx")
# returns tuple of 1) map of fund to all cusips in that fund and 2) pulltime. reads from helix backup file
def old_helix_match():
    print("Fetching CUSIPs from Helix...")
    cusips_all = dict()
    cusips_all["Prime"] = set()
    cusips_all["USG"] = set()
    cusips_all["MMT"] = set()
    cusips_all["LMCP Inv"] = set()

    helix_backup_date = (datetime.now() - timedelta(1)).strftime("%m_%d_%Y")
    helix_pull_time = datetime.now()

    for f in cusips_all.keys():
        helix_backup_file = PureWindowsPath(
            Path("S:/Mandates/Operations/Helix Trade Files/" + f + (" Fund" if f != "LMCP Inv" else "") + ".txt"))
        fund_df = pd.read_csv(helix_backup_file, sep="\t")

        for c in fund_df["BondID"]: # changed 10/3/2019 from "Cusip" to "BondID"
            if isinstance(c, str):
                if c.strip().isalnum(): # added strip 11/26 in case errant spaces
                    cusips_all[f].add(c)

    for extra_mmt in ['XS1951177309','SOSPRUCE1','SOSPRUCE2','XS2091648928','EASTCYPR1']:
        cusips_all["MMT"].add(extra_mmt)

    for f in cusips_all.keys():
        to_add = set()
        for x in cusips_all[f]:
            if len(x) >= 3:
                if x[:3] == "PNI":
                    to_add.add(x[3:len(x)])
        for extra_f in to_add:
            cusips_all[f].add(extra_f)

    return cusips_all, helix_pull_time



# returns tuple of 1) map of fund to all cusips in that fund and 2) pulltime, thru direct SQL query to Helix
# Yating Liu on Feb 27: remove all dummy cusips
def helix_match():
    funds = ["Prime", "USG", "MMT", "LMCP Inv"]
    cusips_all = dict()
    for f in funds:
        cusips_all[f] = set()
        helix_pull_time = datetime.now()
    try:
        print("Fetching CUSIPs from Helix...")
        connection = pymssql.connect(server='LUCIDSQL1',
                                     user='mattiasalmers',
                                     password='12345',
                                     database='HELIXREPO_PROD_02')
        cursor = connection.cursor()
        query =  """
            select distinct
            case when tradepieces.company = 44 then 'USG Fund' when tradepieces.company = 45 then 'Prime Fund' when tradepieces.COMPANY = 46 then 'MMT IM Fund' when tradepieces.COMPANY = 48 then 'LMCP Inv Fund'  when tradepieces.COMPANY = 49 then 'LucidRepo' end Fund,
            ltrim(rtrim(Tradepieces.ISIN)) BondID
            from tradepieces 
            where (tradepieces.isvisible = 1 or tradepieces.company = 49)
            and tradepieces.company in (44,45,46,48,49)
	    and ltrim(rtrim(Tradepieces.ISIN)) not in ('HEXZETA01','HEXZT----','HZLNT----','MCHY-----','MNTNCHRY1','OLIVEEUR-','OLIVEUSD-','OPPOR----','OPPORTUN1','PAAPLEUR-','PAAPLUSD-','PFIR-----','SSPRUCE--','STAPL----','STHAPPLE1','TREATY---','TREATYUS1','ALM2EUR--','ALM2USD--','ALMNDUSD1','ALMONDEUR','ALMONDUSD','ECYP-----','EELM-----','EWILLEUR-','EWILLUSD-')
            order by Fund ASC
        """
        cursor.execute(query)
        records = cursor.fetchall()
        for row in records:
            if row[0] in ['USG Fund', 'Prime Fund', 'MMT IM Fund','LMCP Inv Fund','LucidRepo']:
                k = row[0][:-5] if row[0] in ['USG Fund', 'Prime Fund'] else 'MMT'
                cusips_all[k].add(row[1])
    except Error as e:
        # if this happens must flag and notify else might wipe cusips from the pricer (unless only clean in the morning)
        print("Error reading data from SQL, reverting to last night backup", e)
        return old_helix_match()
    finally:
        connection.close()
        cursor.close()

    for f in cusips_all.keys():
        to_add = set()
        for x in cusips_all[f]:
            if len(x) >= 3:
                if x[:3] == "PNI":
                    to_add.add(x[3:len(x)])
        for extra_f in to_add:
            cusips_all[f].add(extra_f)
    return cusips_all, helix_pull_time


# File to track processed files
base_path = get_file_path(r"S:/Lucid/Data/Bond Data/Historical Price/")

def fetch_jppd(cusips, vdate):
    start_time = time.time()
    print(f"Fetching prices from Pricing Direct for {vdate}...")
    vdate = datetime.strptime(vdate, '%Y%m%d')
    url = 'https://www.pricing-direct.com/pricingdirect/request/priceWsFICusips'
    data = {
        "Cusip": cusips,
        "Date": [vdate.strftime("%m/%d/%Y")],
        "CloseType": ["BOND"],
        "PriceType": ["BID"]
    }
    namepass = "yating:19960601Lyt"
    headers_pd = {"Authorization": "Basic " + base64.b64encode(namepass.encode("utf-8")).decode("utf-8")}
    response = requests.post(url, json=data, headers=headers_pd)

    if response.status_code == 200:
        src = ast.literal_eval(response.text)
        data = []
        for i in range(0, len(src) - 1):  # skip disclaimer, the last
            x = src[i]
            cusip = x["SecurityID"]
            price = x['Bid Evaluation']
            try:
                float_price = float(price)  # Attempt to convert price to float
                data.append([cusip, float_price])  # Append if successful
            except ValueError:
                continue  # Skip appending if conversion fails

        # Create a DataFrame with CUSIPs and prices
        df = pd.DataFrame(data, columns=["CUSIP", "price"])

        # Export to Excel
        vdate = vdate.strftime('%Y%m%d')
        file_name = f"PD_{vdate}"
        output_path = f"{base_path}{file_name}.xlsx"
        df.to_excel(output_path, engine="openpyxl", index=False)
        print(f"Data exported to {output_path}")
        end_time = time.time()
        print(f"Time taken: {end_time - start_time:.2f} seconds")
        return df
    else:
        print(f"Error fetching data from Pricing Direct: {response.status_code}")
        print(f"Response content: {response.content}")
        return None

def fetch_idc(cusips, price_date):
    start_time = time.time()
    print("Fetching prices from IDC...")
    url = "https://rplus.intdata.com/cgi/nph-rplus"
    user = "d4lucid"
    password = "Spring17!"

    # Prepare the request
    auth = base64.b64encode(f"{user}:{password}".encode()).decode()
    headers = {"Authorization": f"Basic {auth}"}
    idc_req = f'GET,({" ,".join(cusips)}),(PRC),{price_date},,D,TITLES=SHORT,DATEFORM=YMD'

    # Send the request
    response = requests.post(url, headers=headers, data={"Request": idc_req, "Done": "flag"})

    if response.status_code == 200:
        lines = response.text.strip().split("\n")
        data = []

        for line in lines[1:-1]:  # Skip the first line (header) and the last line (CRC)
            parts = line.split(",")
            cusip = parts[0].strip('"')  # Remove the quotes around the CUSIP
            price = parts[1]
            try:
                float_price = float(price)  # Attempt to convert price to float
                data.append([cusip, float_price])  # Append if successful
            except ValueError:
                continue  # Skip appending if conversion fails

        # Create a DataFrame with dates as columns and CUSIPs as rows
        df = pd.DataFrame(data, columns=["CUSIP", "price"])
        file_name = f"IDC_{price_date}"
        output_path = get_file_path(base_path + file_name + ".xlsx")

        # Export to Excel
        df.to_excel(output_path, engine="openpyxl")
        print(f"Data exported to {output_path}")
        end_time = time.time()
        print(f"Time taken: {end_time - start_time:.2f} seconds")
    else:
        print(f"Error fetching data from IDC: {response.status_code}")
        return None


# fetch from idc using remoteplus, returns tuple of 1) map of cusips to prices and 2) time request sent
def idc_fetch(cusips):
    print("Fetching prices from IDC...")
    idc_test_req = "GET,(3140H8QR9,13080BAE1, 3136A66B5),(PRC),,, ,Titles=SHORT,DATEFORM=YMD" # confirmed that IDC response format same if doesn't have cusip
    idc_req = "GET,(" + (",".join(cusips)) + "),(PRC),,, ,Titles=SHORT,DATEFORM=YMD"
    #java_path = r"C:\Users\jvulopas\python_projects\IDCPricing"
    java_path = "IDCPricing"
    # HTTP post to remoteplus through java program
    idc_response = subprocess.check_output("java " + java_path + " \"" + idc_req + "\"", shell=True).decode("utf-8")
    idc_response_time = datetime.now()
    print(idc_response)
    idc_prices = dict()

    for p in idc_response.split(";"):
        if len(p.split(",")) != 2:
            continue
        cusip, px = p.split(",")
        idc_prices[cusip] = float(px)
    return idc_prices, idc_response_time

# fetch all bond data from bloomberg and write to bond data csv
def bb_fetch(cusips):
    print("Fetching data from Bloomberg...")
    # write cusips to file (Sept 2020 replaces passing through command line)
    list_path = "curr_fetch_batch.txt"
    outp = open(list_path,"w") # should overwrite
    for x in cusips:
        outp.write(x+"\n")
    outp.close()
    bbpy_path = r"bb_fetch_processor.py"

    try:
        subprocess.check_output("py " + bbpy_path + " " + list_path + " \"Helix\"") # jjv march 2
    except:
        print("Error fetching bond data from Bloomberg. Check bb_fetch_processor.py.")

# JP Morgan pricing direct request (deprecated past 2.7, so I wrote pricing_direct_pricer with different interpreter)
# returns tuple of 1) map of cusips to prices and 2) time request sent
def pd_fetch(cusips):
    print("Fetching prices from JPPD...")
    py27_pd_path = r"pricing_direct_pricer.py"
    pd_cusip_req = ""

    for c in cusips:
        pd_cusip_req = pd_cusip_req + " " + c

    pd_cusip_req = "\n".join(cusips)  # Join the CUSIPs with newline separator

    pddate = datetime.now()
    pddate = pddate if pddate.hour >= 16 and pddate.minute >= 30 else (pddate - BDay(1))

    # command = f"py -2.7 {py27_pd_path} {pddate.strftime('%Y%m%d')}"
    command = f"python {py27_pd_path} {pddate.strftime('%Y%m%d')}"
    process = subprocess.Popen(command, stdin=subprocess.PIPE, stdout=subprocess.PIPE, shell=True, text=True)
    pd_response, _ = process.communicate(pd_cusip_req)

    pd_response_time = datetime.now()
    print(pd_response)
    pd_prices = dict()

    for p in pd_response.splitlines():
        cusip, px = p.split(",")
        try:
            pd_prices[cusip] = float(px)
        except ValueError:
            continue

    return pd_prices, pd_response_time

# write price data to price file
def write_to_price_file_AM(cusips, helix_outp, idc_outp, pd_outp):
    print("Writing to price source...")
    idc_prices = idc_outp[0]
    pd_prices = pd_outp[0]

    wb = op.Workbook()
    
    px_sheet = wb.active
    px_sheet.title = "AM Prices"
    cusip_sheet = wb.create_sheet(title="Helix Cusips")
    info_sheet = wb.create_sheet(title="Info")

    # populate price sheet

    px_sheet["A1"] = "Cusip/ISIN"
    px_sheet["C1"] = "IDC"
    px_sheet["D1"] = "Pricing Direct"

    row = 2
    for c in cusips:
        px_sheet.cell(row=row, column=1).value = c
        if c in idc_prices:
            px_sheet.cell(row=row, column=3).value = idc_prices[c]
        else:
            px_sheet.cell(row=row, column=3).value = "N/A"
        if c in pd_prices:
            px_sheet.cell(row=row, column=4).value = pd_prices[c]
        else:
            px_sheet.cell(row=row, column=4).value = "N/A"
        row = row + 1

    # populate cusip sheet
    col = 1
    for f in helix_outp[0].keys():
        row = 1
        cusip_sheet.cell(row=1, column=col).value = f
        for c in helix_outp[0][f]:
            row = row + 1
            cusip_sheet.cell(row=row, column=col).value = c
        col = col + 1

    # populate log sheet
    info_sheet["A1"] = "Price fetch info"
    info_sheet["A2"] = "Pulled cusips from Helix"
    info_sheet["B2"] = helix_outp[1].strftime("%m/%d/%Y")
    info_sheet["C2"] = helix_outp[1].strftime("%H:%M:%S")
    info_sheet["A3"] = "Helix file used"
    info_sheet["B3"] = (helix_outp[1] - timedelta(1)).strftime("%m_%d_%Y")
    info_sheet["A4"] = "Pulled prices from IDC"
    info_sheet["B4"] = idc_outp[1].strftime("%m/%d/%Y")
    info_sheet["C4"] = idc_outp[1].strftime("%H:%M:%S")
    info_sheet["A5"] = "Pulled prices from PD"
    info_sheet["B5"] = pd_outp[1].strftime("%m/%d/%Y")
    info_sheet["C5"] = pd_outp[1].strftime("%H:%M:%S")
    # save prices
    wb.save(currdest) # overwrite the current pricer
    wb.save(savedest) # save backup
    wb.close()

# kickoff AM process. match against helix and price
def old_AM_process():
    print("AM process.")
    helix_proc = helix_match()
    pxable_cusips = [c for f in helix_proc[0].keys() for c in helix_proc[0][f]]
    pxable_cusips = list(dict.fromkeys(pxable_cusips)) # remove duplicates
    idc_proc = idc_fetch(pxable_cusips)
    pd_proc = pd_fetch(pxable_cusips)
    write_to_price_file_AM(pxable_cusips, helix_proc, idc_proc, pd_proc)
    #bb_fetch(pxable_cusips)
    print("All done.")# kickoff AM process. match against helix and price
    
# Yating Liu
def AM_process():
    print("AM process.")
    helix_proc = helix_match()
    pxable_cusips = [c for f in helix_proc[0].keys() for c in helix_proc[0][f]]
    pxable_cusips = list(dict.fromkeys(pxable_cusips)) # remove duplicates
    
    # open last day's price: 
    # read from PM prices tab if exists, otherwise read from AM prices tab 
    currdest = "S:/Lucid/Data/Bond Data/Price Source/Price_Source.xlsx"
    savedest = "S:/Lucid/Data/Bond Data/Price Source/Archives/Price_Source_" + datetime.now().strftime("%m_%d_%Y") + ".xlsx"
    xl = pd.ExcelFile(currdest)
    if 'PM Prices' in xl.sheet_names:  
        priceT_1 = pd.read_excel(currdest, sheet_name = "PM Prices")  
    else:
        priceT_1 = pd.read_excel(currdest, sheet_name = "AM Prices")  
    # get the date and time for IDC and PD pulls: 
    infoT_1 = pd.read_excel(currdest, sheet_name = "Info")       
   
    lastDayCusip = list(priceT_1['Cusip/ISIN'])
    columns = ['Cusip/ISIN', '', 'IDC', 'Pricing Direct']
    output = pd.DataFrame(columns = columns)
    for i in range(len(pxable_cusips)): 
        cusip = pxable_cusips[i]
        # need IDC_price and PD_price 
        if cusip not in lastDayCusip: 
            IDC_price = "N/A"
            PD_price = "N/A"
        else:
            ind = priceT_1.index[priceT_1['Cusip/ISIN']==cusip]
            price = priceT_1.iloc[ind, [2,3]].reset_index()
            IDC_price = "N/A" if pd.isna(price['IDC'][0]) else price['IDC'][0]
            PD_price = "N/A" if pd.isna(price['Pricing Direct'][0]) else price['Pricing Direct'][0]
        output = output.append({'Cusip/ISIN': cusip, '': '', 'IDC': IDC_price, 'Pricing Direct': PD_price}, ignore_index=True)

    wb = op.Workbook()
    px_sheet = wb.active
    px_sheet.title = "AM Prices"
    # populate AM Prices tab: 
    rows = dataframe_to_rows(output, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            if c_idx != 2: 
                if pd.isnull(value): 
                    px_sheet.cell(row=r_idx, column=c_idx, value="N/A")
                else:
                    px_sheet.cell(row=r_idx, column=c_idx, value=value)
    px_sheet["B1"] = ""    # remove the column name for empty column
    
    cusip_sheet = wb.create_sheet(title="Helix Cusips")
    info_sheet = wb.create_sheet(title="Info")
    # populate cusips tab 
    col = 1
    for f in helix_proc[0].keys():
        row = 1
        cusip_sheet.cell(row=1, column=col).value = f
        for c in helix_proc[0][f]:
            row = row + 1
            cusip_sheet.cell(row=row, column=col).value = c
        col = col + 1
    
    # populate info tab: 
    info_sheet["A1"] = "Price fetch info"
    info_sheet["A2"] = "Pulled cusips from Helix"
    info_sheet["B2"] = helix_proc[1].strftime("%m/%d/%Y")
    info_sheet["C2"] = helix_proc[1].strftime("%H:%M:%S")
    info_sheet["A3"] = "Helix file used"
    info_sheet["B3"] = (helix_proc[1] - timedelta(1)).strftime("%m_%d_%Y")
    info_sheet["A4"] = "Pulled prices from IDC"
    info_sheet["B4"] = infoT_1.iloc[2, 1]
    info_sheet["C4"] = infoT_1.iloc[2, 2]
    info_sheet["A5"] = "Pulled prices from PD"
    info_sheet["B5"] = infoT_1.iloc[3, 1]
    info_sheet["C5"] = infoT_1.iloc[3, 2]
    
    wb.save(currdest)
    wb.save(savedest)
    wb.close()
    print("All done.")

def just_bond_data():
    helix_proc = helix_match()
    pxable_cusips = [c for f in helix_proc[0].keys() for c in helix_proc[0][f]]
    pxable_cusips = list(dict.fromkeys(pxable_cusips)) # remove duplicates
    bb_fetch(pxable_cusips)
    print("All done.")

# kickoff PM process. retrieve fresher prices for securities already in helix that morning
def PM_process():
    print("PM process.")

    try:
        wb = op.load_workbook(currdest)
    except:
        print("file not found")
        exit()

    if 'PM Prices' in wb.sheetnames:
        px_sheet = wb["PM Prices"] # overwrite current PM page if exists
    else:
        px_sheet = wb.copy_worksheet(wb["AM Prices"]) # make new (so copy AM) if doesn't

    px_sheet.title = "PM Prices"
    pxable_cusips = []

    row = 2
    curr = px_sheet.cell(row=row, column=1)
    while (curr.value):
        pxable_cusips.append(curr.value)
        row = row + 1
        curr = px_sheet.cell(row=row, column=1)

    # NEW PROCESS HERE
    # Get the current time in seconds since the epoch
    current_time = time.time()
    # Convert to a datetime object
    date_time = datetime.fromtimestamp(current_time)
    # Format the date as YYYYMMDD
    query_date = date_time.strftime('%Y%m%d')
    fetch_idc(pxable_cusips, query_date)
    fetch_jppd(pxable_cusips, query_date)

    idc_outp = idc_fetch(pxable_cusips)
    pd_outp = pd_fetch(pxable_cusips)
    print("Saving to price source...")
    row = 2
    curr = px_sheet.cell(row=row, column=1)
    while (curr.value):
        if curr.value in idc_outp[0]:
            px_sheet.cell(row=row, column=3).value = idc_outp[0][curr.value]
        else:
            px_sheet.cell(row=row, column=3).value = "N/A"
        if curr.value in pd_outp[0]:
            px_sheet.cell(row=row, column=4).value = pd_outp[0][curr.value]
        else:
            px_sheet.cell(row=row, column=4).value = "N/A"
        row = row + 1
        curr = px_sheet.cell(row=row, column=1)

    info_sheet = wb["Info"]
    # populate log sheet
    info_sheet["A4"] = "Pulled prices from IDC"
    info_sheet["B4"] = idc_outp[1].strftime("%m/%d/%Y")
    info_sheet["C4"] = idc_outp[1].strftime("%H:%M:%S")
    info_sheet["A5"] = "Pulled prices from PD"
    info_sheet["B5"] = pd_outp[1].strftime("%m/%d/%Y")
    info_sheet["C5"] = pd_outp[1].strftime("%H:%M:%S")

    wb.save(currdest) # overwrite current
    wb.save(savedest) # overwrite backup
    wb.close()

    batch_file = open('BatchLogs.txt', 'a')
    batch_file.write("\n(" + datetime.now().strftime("%a %Y-%m-%d;%H:%M:%S") + ";\"price_fetcher\")") 
    batch_file.close()

    print("All done.")

# main process
if __name__ == "__main__":
    if (datetime.now().hour < 12):
        AM_process()
    else:
        PM_process()
