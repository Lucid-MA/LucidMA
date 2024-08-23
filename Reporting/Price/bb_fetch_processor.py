# calls to Java program to fetch bond data from BB using their Server API (iffily documented) from their SDK, then
# filters it to be relevant for lucid purposes
import logging

# to be run during the am process. BB server api jar and BloombergFetch class ought to be in same directory

# kicked off given set of bond ids from price_sourcer.py
# 11/5/2019 replaced yas_ispread with flt_spread

# JJ Vulopas

import openpyxl as op
import pandas as pd
import numpy as np
import subprocess
from datetime import datetime
import sys

from Bronze_tables.Price.bloomberg_utils import BloombergDataFetcher, bb_fields
from Utils.Common import get_file_path

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# fetch recent data from BB given cusips/isins and fields
# commented out prev awb coupon date
legacydate = datetime.strptime("2008-01-01", "%Y-%m-%d")
valdate = datetime.today()
# added some col
fields = (
    "SECURITY_TYP,ISSUER,Collat Typ,Name,Industry Sector,Issue DT,Maturity,Amt Outstanding,Coupon,Floater,"
    "MTG Factor,PX Bid,PX Mid,Int Acc,Mtg WAL,MTG ORIG_WAL,DUR ADJ OAS BID,YAS_MOD_DUR,Days Acc,YLD_ytm_BID,I_SPRD_BID,"
    "FLT_SPREAD,OAS_SPREAD_ASK,MTG TRANCHE TYP LONG,MTG PL CPR 1M,MTG PL CPR 6M,MTG_WHLN_GEO1,MTG_WHLN_GEO2,"
    "MTG_WHLN_GEO3,RTG_SP,RTG_MOODY,RTG_FITCH,RTG_KBRA,RTG_DBRS,RTG_EGAN_JONES,DELIVERY_TYP,DTC_REGISTERED,DTC_ELIGIBLE,MTG_DTC_TYP,"
    "TRADE_DT_ACC_INT,PRINCIPAL_FACTOR,MTG_PREV_FACTOR,MTG_RECORD_DT,MTG_FACTOR_PAY_DT,MTG_NXT_PAY_DT_SET_DT,IDX_RATIO"
)  # 6/24/2020 added principal factor, 9/18/2020 added more factor data

cols = (
    "CUSIP,SECURITY_TYP,ISSUER,Collat Typ,Name,Industry Sector,Issue DT,Maturity,Amt Outstanding,Coupon,Floater,"
    "MTG Factor,PX Bid,PX Mid,Int Acc,Mtg WAL,DUR ADJ OAS BID,YAS_MOD_DUR,USED DURATION,Days Acc,YLD_ytm_BID,"
    "I_SPRD_BID,FLT_SPREAD,OAS_SPREAD_ASK,MTG TRANCHE TYP LONG,MTG PL CPR 1M,MTG PL CPR 6M,MTG_WHLN_GEO1,"
    "MTG_WHLN_GEO2,MTG_WHLN_GEO3,RATINGS BUCKET,RTG_SP,RTG_MOODY,RTG_FITCH,RTG_KBRA,RTG_DBRS,RTG_EGAN_JONES,"
    "DELIVERY_TYP,Est'd Asset Class,CUSIP or ISIN,MTG_PREV_FACTOR,MTG_RECORD_DT,MTG_FACTOR_PAY_DT,MTG_NXT_PAY_DT_SET_DT,IDX_RATIO"
)

# natwest_bonds = ['195117730', 'SOSPRUCE1', 'EASTCYPR1']
# extra_sospruce = ['SOSPRUCE2']
# #extra_sospruce = ['SOSPRUCE2', 'SOSPRUCE3', 'SOSPRUCE4', 'SOSPRUCE5']
# special_cusips = ['CASHUSD01'] + natwest_bonds + extra_sospruce


def depr_wal_since(wal, t1, t0):
    try:
        return wal - ((t1 - t0).days / 365.0)
    except:
        return wal


# cusip, startdate, wal
custom_wals = {"61033WAN2": (datetime.strptime("2021-01-04", "%Y-%m-%d"), 6.25)}


# funds hardwired here
# returns tuple of 1) map of fund to all cusips in that fund and 2) pulltime. reads from helix backup file, 11/25 to read in used market symbol
def fetch_helix_symbols():
    print("Fetching CUSIPs from Helix backup...")
    cusips_all = dict()
    for f in ["Prime", "USG", "MMT", "LMCP Inv"]:
        helix_backup_file = get_file_path(
            "S:/Mandates/Operations/Helix Trade Files/"
            + f
            + (" Fund" if f != "LMCP Inv" else "")
            + ".txt"
        )
        fund_df = pd.read_csv(helix_backup_file, sep="\t")
        for c in fund_df["BondID"]:
            if isinstance(c, str) and c.strip().isalnum():
                cusips_all[c] = (
                    fund_df[fund_df["BondID"] == c]["Description"].iloc[0]
                ).strip()
    return cusips_all


# to translate from helix symbol to the associated code
symbol_classifiers = {
    "Treasury Bill": "TBILL",
    "Treasury Note": "TNOTE",
    "Treasury Bond": "TBOND",
    "Agency NOTE": "AGNCYNT",
    "Agency MBS FIX PT": "MBSFIX",
    "Agency MBS ARM PT": "MBSARM",
    "Agency MBS TRUST": "MBSTRUST",
    "Agency MBS CMO": "MBSCMO",
    "Agency MBS IO FIX": "MBSIOFIX",
    "Agency MBS IO INV": "MBSIOINV",
    "Agency MBS PO": "MBSPO",
    "Agency COMMERCIAL PT": "CMRCLPT",
    "Agency COMMERCIAL IO": "CMRCLIO",
    "Agency PROJECT LOAN PT": "PROJECTPT",
    "Agency PROJECT LOAN IO": "PROJECTIO",
    "GNMA MBS FIX PT": "MBSFIXGMNA",
    "GNMA MBS ARM PT": "MBSARMGMNA",
    "GNMA MBS TRUST": "MBSTRUSTGMNA",
    "GNMA MBS CMO": "MBSCMOGMNA",
    "GNMA MBS IO FIX": "MBSIOFIXGMNA",
    "GNMA MBS IO INV": "MBSIOINVGMNA",
    "GNMA MBS PO": "MBSPOGMNA",
    "GNMA COMMERCIAL PT": "CMRCLPTGMNA",
    "GNMA COMMERCIAL IO": "CMRCLIOGMNA",
    "GNMA PROJECT LOAN PT": "PROJECTPTGMNA",
    "GNMA PROJECT LOAN IO": "PROJECTIOGMNA",
    "SBA SBA REAL ESTATE": "SBARE",
    "SBA SBA EQUIP": "SBAEQUIP",
    "CLO BSL": "CLOBSL",
    "CLO MM": "CLOMM",
    "CLO CRE": "CLOCRE",
    "CLO OTHER": "CLOOTHER",
    "RMBS CRT": "RMBSCRT",
    "RMBS LEGACY": "RMBSLGCY",
    "RMBS RECENT": "RMBSRECNT",
    "RMBS RPL": "RMBSRPL",
    "RMBS QM": "RMBSQM",
    "RMBS NON QM": "RMBSNONQM",
    "RMBS OTHER": "RMBSOTHER",
    "CMBS CONDUIT": "CMBSCOND",
    "CMBS NON-AGNC": "CMBSNA",
    "CMBS SINGLE PROPERTY": "CMBSSINGLE",
    "CMBS OTHER": "CMBSOTHER",
    "ABS AUTO": "ABSAUTOS",
    "ABS CREDIT CARD": "ABSCC",
    "ABS EQUIPMENT": "ABSEQUIP",
    "ABS AIRCRAFT": "ABSAIRCFT",
    "ABS OTHER": "ABSOTHER",
    "CDO OTHER": "CDOOTHER",
    "CDO TRUPS": "CDOTRUPS",
    "CDO BONDS": "CDOBONDS",
    "EURO ABS OTHER": "EUROABSOTHER",
    "FINANCIAL SENIOR FLOAT": "FINSENFLT",
    "FINANCIAL SENIOR FIX": "FINSENFIX",
    "FINANCIAL SUB FLOAT": "FINSUBFLT",
    "FINANCIAL SUB FIX": "FINSUBFIX",
    "CORPORATE 1ST LIEN FLOAT": "CORP1LFLT",
    "CORPORATE 1ST LIEN FIX": "CORP1LFIX",
    "CORPORATE 2ND LIEN FLOAT": "CORP2LFLT",
    "CORPORATE 2ND LIEN FIX": "CORP2LFIX",
    "CORPORATE SENIOR FLOAT": "CORPSENFLT",
    "CORPORATE SENIOR FIX": "CORPSENFIX",
    "CORPORATE SUB FLOAT": "CORPSUBFLT",
    "CORPORATE SUB FIX": "CORPSUBFIX",
    "CD": "MMCD",
    "CP": "MMFCP",
    "ABS CP": "MMABSCP",
    "NON FIN CP": "MMNFCP",
    "VRDN": "MMMUNI",
    "MUNI GO FIX": "MUNIGOFIX",
    "MUNI GO FLT": "MUNIGOFLT",
    "MUNI REV FIX": "MUNIREVFIX",
    "MUNI REV FLT": "MUNIREVFLT",
    "MUNI ABS": "MUNIABS",
    "CMBS CONDUIT IO FIX": "CMBSCONDIO",
    "CMBS SMALL BALANCE": "CMBSSBC",
    "ABS STUDENT LOAN": "ABSSTLN",
    "ABS WHOLE BUSINESS": "ABSWBS",
    "ABS CONSUMERS": "ABSCON",
    "ABS TIMESHARE": "ABSTSHR",
    "ABS SMALL BUS LOAN": "ABSSBLN",
    "TRI PARTY IG CORP": "TRIPTYIGCORP",
    "RMBS SFR": "RMBSSFR",
    "RMBS PRIVATE CRT": "RMBSCRTPVT",
    "RMBS IO FIX": "RMBSIO",
    "RMBS IO INV": "RMBSIOINV",
    "RMBS MTG SVC RIGHTS": "RMBSMSR",
    "RMBS MTG INSURANCE": "RMBSMI",
    "RMBS MANUF HOUSING": "RMBSMNF",
    "GERMAN BUND": "GEBUND",
    "FRENCH OAT": "FROAT",
    "USD Cash": "CASH",
    "French Govt Bond": "FROAT",
}

# these are cusips in BBerg but different ticker to access
diff_cusip_map = {
    "EURFX": "EUR Curncy",
    "XS2606220999": "PPG91FR41",
    "XS2592024009": "PPGA0OKN5",
    "XS2592025071": "PPG80PFP8",
    "ALMNDUSD7": "PPFM1GA98",
    "ALMNDUSD8": "PPFR5TTD6",
    "STHAPPLE5": "PPFT6V9U0",
    "OPPORTUN1": "PPFQKMXT6",
    "STHAPPLE4": "PPF54YY06",
    "MNTNCHRY1": "PPEBEGFI4",
    "MNTNCHRY2": "PPFX3C1P5",
    "52953BBJ1": "PPEG3JY56",
    "STHAPPLE2": "PPED2BZX9",
    "XS2373029664": "PPEZDK875",
    "52468JX82": "PPE43E6K2",
    "XS2373029748": "PPE0FKE65",
    "STHAPPLE1": "PPE32P4N6",
    "HEXZETA01": "PPE9DMNR8",
    "HEXZETA02": "PPE4DGC45",
    "STHAPPLE3": "PPE939O30",
    "ALMNDUSD4": "PPE0F22A9",
    "ALMNDUSD5": "PPEA34F46",
    "ALMNDUSD6": "PPEBFPHO8",
    "ALMNDEUR4": "PPE5GG8O0",
    "ALMNDEUR3": "PPEA34F53",
    "ALMNDUSD3": "PPE139546",
    "ALMNDUSD2": "PPEE2UU10",
    "ALMNDUSD1": "PPE32P4O4",
    "ALMNDEUR2": "PPE42XYF1",
    "ALMNDEUR1": "PPEA2SM53",
    "XS2225938831": "PPEF1MMY3",
    "XS2091648928": "PP9FCKDZ8",
    "XS1951177309": "PP30JD700",
    "XS2004377136": "PP075HWJ5",
    "XS2643730695": "PPG64I278",
    "XS2644211281": "PPG24HYG4",
    "XS2644210986": "PPG64I468",
    "XS2644211109": "PPG64I4G6",
    "TREATYUS1": "PPG1JQ3D1",
    "XS2643730695": "PPG1K4CZ9",
    "XS2644210986": "PPG5K61F1",
    "XS2644211109": "PPG1K4CY2",
    "XS2644211281": "PPG5K61D6",
}

diff_cusip_inv = {v: k for k, v in diff_cusip_map.items()}  # assumes bijection


# this is data fetchable from bberg
def bb_fetch(cusips):
    print("Fetching data from Bloomberg...")
    df = pd.DataFrame(columns=fields.split(","))
    df["CUSIP"] = cusips
    cusip_pass = [
        (diff_cusip_map[x] if x in diff_cusip_map.keys() else x) for x in cusips
    ]
    cusip_pass = [
        (
            "/cusip/"
            if len(x) == 9
            else "/mtge/" if x in ("3137F8RH8", "3137F8ZC0") else "/isin/"
        )
        + x
        for x in cusip_pass
    ]
    cusip_pass.append("EUR Curncy")
    cypress_mapped = False
    # Sept 2020 write to file instead of potentially overloading command line
    list_path = "curr_fetch_batch.txt"
    outp = open(list_path, "w")  # should overwrite
    outp.write(",".join(cusip_pass))
    outp.close()

    df.set_index("CUSIP", inplace=True)
    cmd = 'java -cp "blpapi3.jar;" BloombergFetch "' + list_path + '" "' + fields + '"'
    bb_response = subprocess.check_output(cmd, shell=True).decode("utf-8")
    if "LIMIT (Workflow review needed" in bb_response:
        raise Exception("Bloomberg data limit reached. Contact help desk.")
    if "xxfailurexx" in bb_response:
        raise Exception("Error fetching from Bloomberg.")
    bb_response_time = datetime.now()
    print(bb_response)
    data_by_cusip = bb_response.splitlines()
    print(data_by_cusip)
    for line in data_by_cusip:
        if line and ("BAD_SEC" not in line):
            spl = line.split(":")
            cusip = spl[0]
            cusip = diff_cusip_inv[cusip] if cusip in diff_cusip_inv.keys() else cusip
            flds = spl[1].split(";")
            for entry in flds:
                if entry:
                    pair = entry.split("~")
                    df.loc[cusip, pair[0]] = pair[1]
                    # cypress mapped twice
                    if cusip == "EASTCYPR1" and ("XS2004377136" in df.index):
                        df.loc["XS2004377136", pair[0]] = pair[1]
    return df


def bb_fetch_v2(cusips):
    cusip_pass = [
        (diff_cusip_map[x] if x in diff_cusip_map.keys() else x) for x in cusips
    ]
    cusip_pass = [
        (
            "/cusip/"
            if len(x) == 9
            else "/mtge/" if x in ("3137F8RH8", "3137F8ZC0") else "/isin/"
        )
        + x
        for x in cusip_pass
    ]
    cusip_pass.append("EUR Curncy")
    cypress_mapped = False
    # Sept 2020 write to file instead of potentially overloading command line
    list_path = "curr_fetch_batch_v2.txt"
    outp = open(list_path, "w")  # should overwrite
    outp.write(",".join(cusip_pass))
    outp.close()

    fetcher = BloombergDataFetcher()

    logging.info("Fetching security attributes...")
    security_attributes_df = fetcher.get_security_attributes(cusip_pass, bb_fields)

    # Invert the diff_cusip_map dictionary so that values become keys and keys become values
    inverted_diff_cusip_map = {v: k for k, v in diff_cusip_map.items()}

    # Replace CUSIP values using the inverted_diff_cusip_map dictionary
    security_attributes_df["CUSIP"] = security_attributes_df["CUSIP"].map(
        lambda x: inverted_diff_cusip_map.get(x, x)
    )
    security_attributes_df.set_index("CUSIP", inplace=True)
    logging.info(security_attributes_df)
    return security_attributes_df


# jan 2021
def bb_fetch_with_overrides(mktsymbol_map):
    print("Fetching overrides by product type")
    param_file = get_file_path("S:/Lucid/Data/Bond Data/Bloomberg_Calc_Parameters.xlsx")

    param_df = pd.read_excel(param_file, header=2, usecols="B:E").dropna()
    cols = []

    for col in param_df:
        cols.append(col)

    try:
        cols.remove("Asset Class")
    except:
        pass
    param_df.set_index("Asset Class", inplace=True)
    df = pd.DataFrame(columns=cols + ["CUSIP", "Asset Class"])
    df.set_index("CUSIP", inplace=True)

    for cusip in mktsymbol_map:
        asset_class = mktsymbol_map[cusip]
        defined = True
        for col in param_df:
            try:
                df.loc[cusip, col] = param_df.loc[asset_class, col]
            except:
                defined = False
                continue
        if defined:
            df.loc[cusip, "Asset Class"] = asset_class
        else:
            try:
                df.drop([cusip])
            except:
                continue

    df_no_index = df.reset_index()
    distincts = df_no_index[cols].drop_duplicates()

    list_path = "curr_fetch_batch.txt"
    fields = "MTG_WAL"
    overrides_init = " MTG_PREPAY_TYP CPR DEFAULT_TYPE CDR"

    override_name_map = dict()
    override_name_map["CPR"] = "PREPAY_SPEED_VECTOR"
    override_name_map["CDR"] = "DEFAULT_SPEED_VECTOR"

    for i, row in distincts.iterrows():
        iterdf = df_no_index
        overrides = overrides_init
        for col in cols:
            iterdf = iterdf.loc[df_no_index[col] == row[col]]
            overrides = (
                overrides
                + " "
                + (override_name_map[col] if col in override_name_map else col)
                + " "
                + str(row[col])
            )
        cusip_pass = iterdf["CUSIP"].values
        cusip_pass = [
            (
                "/cusip/"
                if len(x) == 9
                else "/mtge/" if x in ("3137F8RH8", "3137F8ZC0") else "/isin/"
            )
            + x
            for x in cusip_pass
        ]

        outp = open(list_path, "w")  # should overwrite
        outp.write(",".join(cusip_pass))
        outp.close()

        cmd = (
            'java -cp "blpapi3.jar;" BloombergFetch "'
            + list_path
            + '" "'
            + fields
            + '"'
            + overrides
        )
        print("Fetching for " + overrides + "...")
        bb_response = subprocess.check_output(cmd, shell=True).decode("utf-8")
        if "xxfailurexx" in bb_response:
            print("problem")  # TODO error handle
        data_by_cusip = bb_response.splitlines()
        for line in data_by_cusip:
            if line and ("BAD_SEC" not in line):
                spl = line.split(":")
                cusip = spl[0]
                flds = spl[1].split(";")
                for entry in flds:
                    if entry:
                        pair = entry.split("~")
                        try:
                            df.loc[cusip, pair[0]] = float(pair[1])
                        except:
                            df.loc[cusip, pair[0]] = pair[1]

    return df.rename(columns={"MTG_WAL": "Mtg WAL"}).fillna("")


def bb_fetch_with_overrides_v2(mktsymbol_map):
    print("Fetching overrides by product type")
    param_file = get_file_path("S:/Lucid/Data/Bond Data/Bloomberg_Calc_Parameters.xlsx")

    param_df = pd.read_excel(param_file, header=2, usecols="B:E").dropna()
    cols = []

    for col in param_df:
        cols.append(col)

    try:
        cols.remove("Asset Class")
    except:
        pass
    param_df.set_index("Asset Class", inplace=True)
    df = pd.DataFrame(columns=cols + ["CUSIP", "Asset Class"])
    df.set_index("CUSIP", inplace=True)

    for cusip in mktsymbol_map:
        asset_class = mktsymbol_map[cusip]
        defined = True
        for col in param_df:
            try:
                df.loc[cusip, col] = param_df.loc[asset_class, col]
            except:
                defined = False
                continue
        if defined:
            df.loc[cusip, "Asset Class"] = asset_class
        else:
            try:
                df.drop([cusip])
            except:
                continue

    ### Here is fine ###
    df = df.reset_index()

    cusip_pass = df["CUSIP"].values
    cusip_pass = [
        (
            "/cusip/"
            if len(x) == 9
            else "/mtge/" if x in ("3137F8RH8", "3137F8ZC0") else "/isin/"
        )
        + x
        for x in cusip_pass
    ]

    list_path = "curr_fetch_batch.txt"
    fields = ["MTG_WAL"]

    outp = open(list_path, "w")  # should overwrite
    outp.write(",".join(cusip_pass))
    outp.close()

    fetcher = BloombergDataFetcher()

    logging.info("Fetching security attributes...")
    security_attributes_overrides_df = fetcher.get_security_attributes(
        cusip_pass, fields
    )
    logging.info(security_attributes_overrides_df)

    result_df = df.merge(
        security_attributes_overrides_df[["CUSIP", "MTG_WAL"]], on="CUSIP", how="left"
    )

    result_df.set_index("CUSIP", inplace=True)

    return result_df.rename(columns={"MTG_WAL": "Mtg WAL"}).fillna("")


# (checked) msp's lucid rating function, made pythonic
def lucid_rating(sp, moodys, fitch, kroll, dbrs, ej, issuer, sectype):
    # check edge cases
    if sectype[:9] == "Agncy CMO" or sectype[:10] == "Agncy CMBS":
        return "USGCMO"
    elif (
        sectype[:10] == "US GOVERNM"
        or issuer[:10] == "Fannie Mae"
        or issuer[:11] == "Freddie Mac"
        or issuer[:28] == "Government National Mortgage"
        or sectype[:3] == "SBA"
    ):
        if not (
            issuer[-5:] == "STACR"
            or issuer[-3:] == "CRT"
            or issuer == "Fannie Mae - CAS"
        ):  # check if CRT
            return "USG"

    # check other ratings not moody's
    # for x in {sp, fitch, kroll, dbrs, ej}:
    #    if str(x)[-1:] in {'1', '2', '3'}:
    #        return "Error: Moody's rating where it shouldn't be"

    rmap = lambda x: (
        0
        if x[:3] == "AAA"
        else (
            1
            if x[:2] == "AA"
            else (
                2
                if x[:1] == "A"
                else (
                    3
                    if (x[:3] == "BBB" or x[:3] == "BAA")
                    else 4 if (x[:2] == "BB" or x[:2] == "BA") else 5
                )
            )
        )
    )

    ratingsindex = [
        rmap(x.upper())
        for x in [sp[3:] if sp[:3] == "(P)" else sp, moodys, fitch, kroll, dbrs, ej]
    ]
    return ["AAA", "AA", "A", "BBB", "BB", "NR"][min(ratingsindex)]


# (checked) msp's helix market symbol, made *slightly* more pythonic. will give current rating
def helix_mkt_symbol(
    sectype,
    issuer,
    colltype,
    issuedt,
    bondname,
    indsector,
    tranchetype,
    floater,
    legacydt,
):
    if issuer[-5:] == "STACR" or issuer == "Fannie Mae - CAS":
        return "RMBSCRT"  # hardcoded in jan 2020
    if not isinstance(issuedt, datetime):
        issuedt = datetime.strptime(issuedt, "%m/%d/%Y")
    tobeset = True
    isFixIO = False
    isPO = False
    isInvIo = False
    isCommercial = False
    isFix = False

    outp = ""

    gnmasuffix = "GMNA" if issuer[:28].lower() == "government national mortgage" else ""
    fixsuffix = "FIX" if floater == "N" else "FLT"
    if sectype[:3] == "MBS":
        outp = ("MBSARM" if floater == "Y" else "MBSFIX") + gnmasuffix
        tobeset = False

    # if agency cmo
    if tobeset and sectype[:5] == "Agncy":
        tobeset = False
        if "CMBS" in sectype:
            isCommercial = True
        if sectype[-2:] == "IO" or "IO" in tranchetype:
            if "INV" in tranchetype:
                isInvIo = True
            else:
                isFixIO = True
        if sectype[-2:] == "PO":
            isPO = True
        outp = (
            (
                "CMRCLIO"
                if isFixIO
                else "CMRCLIOINVFIX" if isInvIo else "CMRCLPO" if isPO else "CMRCLPT"
            )
            if isCommercial
            else (
                "MBSIOFIX"
                if isFixIO
                else "MBSIOINV" if isInvIo else "MBSPO" if isPO else "MBSCMO"
            )
        ) + gnmasuffix

    # if sba
    if tobeset and sectype[:3] == "SBA":
        tobeset = False
        outp = "SBAEQUIP"
    # if cmbs
    if tobeset and sectype[:4] == "CMBS":
        tobeset = False
        outp = "CMBSNA"
    # if treasury
    if tobeset and colltype == "US GOVT":
        tobeset = False
        outp = (
            "TBill"
            if bondname == "TREASURY BILL"
            else "TNote" if bondname == "US TREASURY N/B" else "Unknown!"
        )
    # if clo
    if tobeset and colltype[:6] == "CF-CLO":
        tobeset = False
        outp = (
            "CLOBSL"
            if colltype[-2:] == "LL"
            else "CLOMM" if colltype[-3:] == "MML" else "Unknown!"
        )
        # if muni
    if tobeset and indsector == "Government":
        tobeset = False
        outp = "MUNI"
    # if financial
    if tobeset and indsector == "Financial":
        tobeset = False
        outp = (
            "FINSUB" + fixsuffix
            if colltype == "SUBORDINATED"
            else "FINSEN" + fixsuffix if colltype == "SR UNSECURED" else "Unknown!"
        )
    # if rmbs
    if tobeset and sectype[:3] == "ABS":
        # here not setting tobeset to false
        if colltype[:3] == "RES" or colltype[:4] == "HOME":
            tobeset = False
            outp = "RMBSLGCY" if issuedt < legacydt else "RMBSRECNT"
    if tobeset and sectype[:8] == "Prvt CMO":
        tobeset = False
        outp = "RMBSLGCY" if issuedt < legacydt else "RMBSRECNT"
    # if other abs
    if tobeset and sectype[:9] == "ABS Other":
        tobeset = False
        outp = "ABSOTHER"
    # if auto abs
    if tobeset and sectype[:8] == "ABS Auto":
        tobeset = False
        outp = "ABSAutos"

    # default to corps
    if tobeset:
        outp = "CORPFLT" if floater == "Y" else "CORPFIX"
    return outp


# (checked) fetch data for special cusips
def fetch_spec_df():
    outp = pd.DataFrame(
        [
            {
                "CUSIP": "CASHUSD01",
                "SECURITY_TYP": "CASH",
                "ISSUER": "US TREASURY",
                "Collat Typ": "Cash",
                "Name": "Cash",
                "Industry Sector": "USD Cash",
                "Issue DT": "7/4/1776",
                "Maturity": "12/31/2099",
                "Amt Outstanding": "1",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "0",
                "YAS_MOD_DUR": "0",
                "USED DURATION": "0.00273972",
                "Days Acc": "0",
                "YLD_ytm_BID": "0",
                "I_SPRD_BID": "0",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "0",
                "MTG TRANCHE TYP LONG": "CASH",
                "MTG PL CPR 1M": "0",
                "MTG PL CPR 6M": "0",
                "MTG_WHLN_GEO1": "0",
                "MTG_WHLN_GEO2": "0",
                "MTG_WHLN_GEO3": "0",
                "RATINGS BUCKET": "USG",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_EGAN_JONES": "",
                "DELIVERY_TYP": "",
                "Est'd Asset Class": "CASH",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "1",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPCASHUSD",
                "SECURITY_TYP": "CASH",
                "ISSUER": "US TREASURY",
                "Collat Typ": "Cash",
                "Name": "Cash",
                "Industry Sector": "USD Cash",
                "Issue DT": "7/4/1776",
                "Maturity": "12/31/2099",
                "Amt Outstanding": "1",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "0",
                "YAS_MOD_DUR": "0",
                "USED DURATION": "0.00273972",
                "Days Acc": "0",
                "YLD_ytm_BID": "0",
                "I_SPRD_BID": "0",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "0",
                "MTG TRANCHE TYP LONG": "CASH",
                "MTG PL CPR 1M": "0",
                "MTG PL CPR 6M": "0",
                "MTG_WHLN_GEO1": "0",
                "MTG_WHLN_GEO2": "0",
                "MTG_WHLN_GEO3": "0",
                "RATINGS BUCKET": "USG",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_EGAN_JONES": "",
                "DELIVERY_TYP": "",
                "Est'd Asset Class": "CASH",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "1",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "CASHEUR01",
                "SECURITY_TYP": "CASH",
                "ISSUER": "ECB",
                "Collat Typ": "Cash",
                "Name": "Cash",
                "Industry Sector": "EUR Cash",
                "Issue DT": "1/1/1999",
                "Maturity": "12/31/2099",
                "Amt Outstanding": "1",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "0",
                "YAS_MOD_DUR": "0",
                "USED DURATION": "0.00273972",
                "Days Acc": "0",
                "YLD_ytm_BID": "0",
                "I_SPRD_BID": "0",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "0",
                "MTG TRANCHE TYP LONG": "CASH",
                "MTG PL CPR 1M": "0",
                "MTG PL CPR 6M": "0",
                "MTG_WHLN_GEO1": "0",
                "MTG_WHLN_GEO2": "0",
                "MTG_WHLN_GEO3": "0",
                "RATINGS BUCKET": "USG",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_EGAN_JONES": "",
                "DELIVERY_TYP": "",
                "Est'd Asset Class": "CASH",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "1",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "74039RAA8",
                "SECURITY_TYP": "",
                "ISSUER": "Preeti Trust SWMC 2021",
                "Collat Typ": "",
                "Name": "",
                "Industry Sector": "",
                "Issue DT": "6/1/2021",
                "Maturity": "12/31/2021",
                "Amt Outstanding": "50000000",
                "Coupon": "2.96",
                "Floater": "",
                "MTG Factor": "1",
                "PX Bid": "101.5",
                "PX Mid": "101.5",
                "Int Acc": "0",
                "Mtg WAL": "5.2",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "USG",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "MBSTRUST",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "371494AK1",
                "SECURITY_TYP": "",
                "ISSUER": "Tri Party IGCorp",
                "Collat Typ": "",
                "Name": "Tri Party IGCorp",
                "Industry Sector": "",
                "Issue DT": "8/10/2021",
                "Maturity": "8/10/2031",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "5",
                "DUR ADJ OAS BID": "5",
                "YAS_MOD_DUR": "5",
                "USED DURATION": "5",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "TRIPTYIGCORP",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "2063C0VL3",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "Concord",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "A1/P1",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "MMFCP",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-TH2O",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-TH2O",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-SCHF1",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-SCHF1",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-DYM1",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-DYM1",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-4631",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-4631",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-PBTA1",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-PBTA1",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-SVI1",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-SVI1",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            {
                "CUSIP": "JPM-MANT1",
                "SECURITY_TYP": "",
                "ISSUER": "Concord",
                "Collat Typ": "",
                "Name": "JPM-MANT1",
                "Industry Sector": "",
                "Issue DT": "8/13/2021",
                "Maturity": "8/20/2021",
                "Amt Outstanding": "10000000",
                "Coupon": "0",
                "Floater": "N",
                "MTG Factor": "1",
                "PX Bid": "100",
                "PX Mid": "100",
                "Int Acc": "0",
                "Mtg WAL": "0",
                "DUR ADJ OAS BID": "",
                "YAS_MOD_DUR": "",
                "USED DURATION": "",
                "Days Acc": "",
                "YLD_ytm_BID": "",
                "I_SPRD_BID": "",
                "FLT_SPREAD": "",
                "OAS_SPREAD_ASK": "",
                "MTG TRANCHE TYP LONG": "",
                "MTG PL CPR 1M": "",
                "MTG PL CPR 6M": "",
                "MTG_WHLN_GEO1": "",
                "MTG_WHLN_GEO2": "",
                "MTG_WHLN_GEO3": "",
                "RATINGS BUCKET": "BBB",
                "RTG_SP": "",
                "RTG_MOODY": "",
                "RTG_FITCH": "",
                "RTG_KBRA": "",
                "RTG_DBRS": "",
                "RTG_Egan_Jones": "",
                "DELIVERY_TYP": "PHYS",
                "Est'd Asset Class": "FUNDINT",
                "CUSIP or ISIN": " CUSIP",
                "MTG_PREV_FACTOR": "",
                "MTG_RECORD_DT": "",
                "MTG_FACTOR_PAY_DT": "",
                "MTG_NXT_PAY_DT_SET_DT": "",
            },
            # {'CUSIP': 'JPM-TH2O', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''},
            # {'CUSIP': 'JPM-SCHF1', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''},
            # {'CUSIP': 'JPM-TH2O', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''},
            # {'CUSIP': 'JPM-DYM1', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''},
            # {'CUSIP': 'JPM-4631', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''},
            # {'CUSIP': 'JPM-PBTA1', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''},
            # {'CUSIP': 'JPM-MANT1', 'SECURITY_TYP': '', 'ISSUER': 'JTOP', 'Collat Typ': '', 'Name': 'JPM-TH2O',
            # 'Industry Sector': '', 'Issue DT': '8/13/2021', 'Maturity': '8/20/2021', 'Amt Outstanding': '10000000', 'Coupon': '0',
            # 'Floater': 'N', 'MTG Factor': '1', 'PX Bid': '100', 'PX Mid': '100', 'Int Acc': '0', 'Mtg WAL': '0',
            # 'DUR ADJ OAS BID': '', 'YAS_MOD_DUR': '', 'USED DURATION': '', 'Days Acc': '', 'YLD_ytm_BID': '', 'I_SPRD_BID': '',
            # 'FLT_SPREAD': '', 'OAS_SPREAD_ASK': '', 'MTG TRANCHE TYP LONG': '', 'MTG PL CPR 1M': '', 'MTG PL CPR 6M': '',
            # 'MTG_WHLN_GEO1': '', 'MTG_WHLN_GEO2': '', 'MTG_WHLN_GEO3': '', 'RATINGS BUCKET': 'BBB', 'RTG_SP': '', 'RTG_MOODY': '',
            # 'RTG_FITCH': '', 'RTG_KBRA': '', 'RTG_DBRS': '', 'RTG_Egan_Jones': '', 'DELIVERY_TYP': 'PHYS', "Est'd Asset Class": 'FUNDINT',
            # 'CUSIP or ISIN': ' CUSIP', 'MTG_PREV_FACTOR': '', 'MTG_RECORD_DT': '', 'MTG_FACTOR_PAY_DT': '', 'MTG_NXT_PAY_DT_SET_DT': ''}
            # {"CUSIP": "195117730", "SECURITY_TYP": "Secured PE Loan", "ISSUER": "Eastern Elm Co", "Collat Typ": "PE",
            #  "Name": "East Elm 2019-1", "Industry Sector": "Asset Backed Securities", "Issue DT": "2/12/2019",
            #  "Maturity": "3/15/2029", "Amt Outstanding": "187500000", "Coupon": "4.3685", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "1.09838888888889", "Mtg WAL": "9.51506849315068",
            #  "DUR ADJ OAS BID": "9.51506849315068", "YAS_MOD_DUR": "9.51506849315068", "USED DURATION": "9.51506849315068",
            #  "Days Acc": "85", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "-", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
            # {"CUSIP": "SOSPRUCE1", "SECURITY_TYP": "Secured Loan", "ISSUER": "Southern Spruce Co", "Collat Typ": "Fund",
            #  "Name": "Southern Spruce 2019-1", "Industry Sector": "Asset Backed Securities", "Issue DT": "4/30/2019",
            #  "Maturity": "6/15/2029", "Amt Outstanding": "35000000", "Coupon": "4.7935", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "1.19873611111111", "Mtg WAL": "9.76712328767123",
            #  "DUR ADJ OAS BID": "9.76712328767123", "YAS_MOD_DUR": "9.76712328767123", "USED DURATION": "9.76712328767123",
            #  "Days Acc": "85", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "SOSPRUCE", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
            # {"CUSIP": "EASTCYPR1", "SECURITY_TYP": "Secured Loan", "ISSUER": "Eastern Cypress Co", "Collat Typ": "Fund",
            #  "Name": "Eastern Cypress 2019-1", "Industry Sector": "Asset Backed Securities", "Issue DT": "6/4/2019",
            #  "Maturity": "5/29/2029", "Amt Outstanding": "56000000", "Coupon": "1.667", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "0.42075", "Mtg WAL": "9.72054794520548",
            #  "DUR ADJ OAS BID": "9.72054794520548", "YAS_MOD_DUR": "9.72054794520548", "USED DURATION": "9.72054794520548",
            #  "Days Acc": "85", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "SOSPRUCE", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
            # {"CUSIP": "SOSPRUCE2", "SECURITY_TYP": "Secured Loan", "ISSUER": "Southern Spruce Co", "Collat Typ": "Fund",
            #  "Name": "Southern Spruce 2019-2", "Industry Sector": "Asset Backed Securities", "Issue DT": "4/30/2019",
            #  "Maturity": "6/15/2029", "Amt Outstanding": "35000000", "Coupon": "4.7935", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "1.19873611111111", "Mtg WAL": "9.76712328767123",
            #  "DUR ADJ OAS BID": "9.76712328767123", "YAS_MOD_DUR": "9.76712328767123", "USED DURATION": "9.76712328767123",
            #  "Days Acc": "85", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "SOSPRUCE", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
            # {"CUSIP": "SOSPRUCE3", "SECURITY_TYP": "Secured Loan", "ISSUER": "Southern Spruce Co", "Collat Typ": "Fund",
            #  "Name": "Southern Spruce 2019-3", "Industry Sector": "Asset Backed Securities", "Issue DT": "4/30/2019",
            #  "Maturity": "6/15/2029", "Amt Outstanding": "35000000", "Coupon": "5.077", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "0.423083333333333", "Mtg WAL": "9.76712328767123",
            #  "DUR ADJ OAS BID": "9.76712328767123", "YAS_MOD_DUR": "9.76712328767123", "USED DURATION": "9.76712328767123",
            #  "Days Acc": "30", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "SOSPRUCE", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
            # {"CUSIP": "SOSPRUCE4", "SECURITY_TYP": "Secured Loan", "ISSUER": "Southern Spruce Co", "Collat Typ": "Fund",
            #  "Name": "Southern Spruce 2019-4", "Industry Sector": "Asset Backed Securities", "Issue DT": "4/30/2019",
            #  "Maturity": "6/15/2029", "Amt Outstanding": "35000000", "Coupon": "5.077", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "0.310261111111111", "Mtg WAL": "9.76712328767123",
            #  "DUR ADJ OAS BID": "9.76712328767123", "YAS_MOD_DUR": "9.76712328767123", "USED DURATION": "9.76712328767123",
            #  "Days Acc": "22", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "SOSPRUCE", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
            # {"CUSIP": "SOSPRUCE5", "SECURITY_TYP": "Secured Loan", "ISSUER": "Southern Spruce Co", "Collat Typ": "Fund",
            #  "Name": "Southern Spruce 2019-5", "Industry Sector": "Asset Backed Securities", "Issue DT": "4/30/2019",
            #  "Maturity": "6/15/2029", "Amt Outstanding": "35000000", "Coupon": "5.077", "Floater": "Y", "MTG Factor": "1",
            #  "PX Bid": "100", "PX Mid": "100", "Int Acc": "0.310261111111111", "Mtg WAL": "9.76712328767123",
            #  "DUR ADJ OAS BID": "9.76712328767123", "YAS_MOD_DUR": "9.76712328767123", "USED DURATION": "9.76712328767123",
            #  "Days Acc": "22", "YLD_ytm_BID": "", "I_SPRD_BID": "", "FLT_SPREAD": "", "OAS_SPREAD_ASK": "0",
            #  "MTG TRANCHE TYP LONG": "SOSPRUCE", "MTG PL CPR 1M": "0", "MTG PL CPR 6M": "0", "MTG_WHLN_GEO1": "0",
            #  "MTG_WHLN_GEO2": "0", "MTG_WHLN_GEO3": "0", "RATINGS BUCKET": "A", "RTG_SP": "", "RTG_MOODY": "A1",
            #  "RTG_FITCH": "", "RTG_KBRA": "", "RTG_DBRS": "", "RTG_EGAN_JONES": "", "DELIVERY_TYP": "",
            #  "Est'd Asset Class": "ABSOTHER", "CUSIP or ISIN": " CUSIP"},
        ]
    )
    outp.set_index("CUSIP", inplace=True)
    # ss_isdate = {
    #     #'SOSPRUCE2': datetime.strptime("2019-09-25","%Y-%m-%d"),
    #     # 'SOSPRUCE3': datetime.strptime("2019-08-12", "%Y-%m-%d"),
    #     # 'SOSPRUCE4': datetime.strptime("2019-08-20", "%Y-%m-%d"),
    #     # 'SOSPRUCE5': datetime.strptime("2019-09-10", "%Y-%m-%d"),
    # }
    # for c in natwest_bonds + extra_sospruce:
    #     days_acc = (max(0, (valdate-(prev_awbury_coup if c in natwest_bonds else ss_isdate[c])).days))
    #     outp.at[c, 'Int Acc'] = days_acc * float(outp['Coupon'][c]) / 360
    #     wal = (datetime.strptime(outp['Maturity'][c],"%m/%d/%Y") - valdate).days / 365
    #     outp.at[c, 'Mtg WAL'] = wal
    #     outp.at[c, 'DUR ADJ OAS BID'] = wal
    #     outp.at[c, 'YAS_MOD_DUR'] = wal
    #     outp.at[c, 'USED DURATION']= wal
    #     outp.at[c, 'Days Acc'] = days_acc
    return outp


# handle any exceptional cases, called before and after processing to ensure that there in both states.
def hardwired_adjustments(df):
    # df.loc[df['CUSIP'] == '87299AVF8', 'YAS_MOD_DUR'] = 0# loc will only update if cusip exists
    # df.loc[df['CUSIP'] == '87299AVF8', 'USED DURATION'] = df.loc[df['CUSIP'] == '87299AVF8', 'Mtg WAL'].item()
    for c in diff_cusip_map.keys():
        if c in df.index:
            df.at[c, "RATINGS BUCKET"] = "A"
            df.at[c, "Est'd Asset Class"] = "ABSOTHER"
            # df.at['87299AVF8', 'YAS_MOD_DUR'] = 0# loc will only update if cusip exists
            # df.at['87299AVF8', 'USED DURATION'] = df.at['87299AVF8', 'Mtg WAL']
            # df.at['87299AVF8', 'Issue DT'] = "2018-9-11"
    if "XS2004377136" in df.index:
        df.at["XS2004377136", "Est'd Asset Class"] = "FROAT"
        df.at["XS2004377136", "RATINGS BUCKET"] = "A"
    return df


# highly dependent on specific structure of df of raw fetch from bloomberg
# go column by column and filter
def process_bb_data(raw_df, mktsymbol_map, df_custom_overrides):
    print("Processing data from Bloomberg...")
    df = raw_df.copy()
    # df = hardwired_adjustments(df)
    # add custom columns (not among queried bb fields)
    for newcol in np.setdiff1d(cols.split(","), fields.split(",")):
        if newcol != "CUSIP":
            df[newcol] = ""

    df["correct_days"] = ""
    # outp = pd.DataFrame(columns=df.columns)
    newrows = []
    hardcoded_usg_rating = [
        "31424WCR1",
        "3136G4X32",
        "3133ENUV0",
        "3130AGFP5",
        "3130AJP60",
        "31422XLK6",
        "3130AJZ36",
    ]
    hardcoded_usgcmo_rating = [
        "3136ASL95",
        "3137BHE45",
        "3137F8RH8",
        "3137F97B1",
        "3137FLYE8",
        "3137FPZ73",
        "38379Q5M3",
        "38380TXD3",
        "38382CXN6",
    ]
    hardcoded_principal_factor_field = [
        "XS1951177309",
        "XS2606220999",
        "XS2373029664",
        "3134GW5R3",
        "XS2225938831",
        "XS2091648928",
        "XS2373029748",
        "XS2004377136",
        "00908PAB3",
        "00909DAA1",
        "XS2592024009",
        "XS2644211109",
        "XS2644211281",
        "XS2643730695",
        "XS2644210986",
        "XS2592025071",
    ]
    hardcoded_inflation_indexed = [
        "9128287D6",
        "912828H45",
        "912810SM1",
        "912810RF7",
        "912828YL8",
    ]
    for cusip, row in df.iterrows():  # no concurrent modification problem
        for col in cols.split(",") + ["correct_days"]:
            try:
                if col == "Collat Typ":
                    row[col] = (row[col], "US GOVT")[
                        row["ISSUER"][:5] == "US TR" or row["ISSUER"][:5] == "TREAS"
                    ]
                elif col in {"Coupon", "PX Bid", "PX Mid", "Int Acc", "Days Acc"}:
                    try:
                        row[col] = 0 if not row[col] else round(float(row[col]), 6)
                    except:
                        row[col] = 0
                elif col == "Issue DT" or col == "Maturity":
                    try:
                        row[col] = (
                            datetime.strptime("1900-01-01", "%Y-%m-%d")
                            if not row[col]
                            else datetime.strptime(row[col], "%Y-%m-%d").strftime(
                                "%m/%d/%Y"
                            )
                        )
                    except:
                        row[col] = datetime.strptime("1900-01-01", "%Y-%m-%d")
                elif col in {"MTG Factor"}:
                    try:
                        row[col] = 1 if not row[col] else float(row[col])
                    except:
                        row[col] = 1
                    if cusip in hardcoded_principal_factor_field:
                        try:
                            row[col] = float(row["PRINCIPAL_FACTOR"])
                        except:
                            row[col] = 1
                    if cusip in hardcoded_inflation_indexed:
                        try:
                            row[col] = float(row["IDX_RATIO"])
                        except:
                            row[col] = 1
                elif col == "Mtg WAL":
                    if cusip in custom_wals:
                        row[col] = depr_wal_since(
                            custom_wals[cusip][1],
                            datetime.today(),
                            custom_wals[cusip][0],
                        )
                    else:
                        row[col] = row["MTG ORIG_WAL"] if not row[col] else row[col]
                elif col == "USED DURATION":
                    if not row["DUR ADJ OAS BID"]:
                        if not row["Mtg WAL"]:
                            if not row["YAS_MOD_DUR"]:
                                try:
                                    row[col] = (
                                        (
                                            datetime.strptime(
                                                row["Maturity"], "%m/%d/%Y"
                                            )
                                            - datetime.today()
                                        ).days
                                    ) / 365.0
                                except:
                                    row[col] = ""
                            else:
                                row[col] = row["YAS_MOD_DUR"]
                        else:
                            row[col] = row["Mtg WAL"]
                    else:
                        row[col] = row["DUR ADJ OAS BID"]
                elif col in {
                    "Name",
                    "Industry Sector",
                    "Floater",
                    "Amt Outstanding",
                    "YLD_ytm_BID",
                    "I_SPRD_BID",
                    "FLT_SPREAD",
                    "OAS_SPREAD_ASK",
                    "MTG TRANCHE TYP LONG",
                    "MTG PL CPR 1M",
                    "MTG PL CPR 6M",
                    "MTG_WHLN_GEO1",
                    "MTG_WHLN_GEO2",
                    "MTG_WHLN_GEO3",
                }:
                    row[col] = "" if not row[col] else row[col]
                elif col == "RATINGS BUCKET":
                    row[col] = lucid_rating(
                        row["RTG_SP"],
                        row["RTG_MOODY"],
                        row["RTG_FITCH"],
                        row["RTG_KBRA"],
                        row["RTG_DBRS"],
                        row["RTG_EGAN_JONES"],
                        row["ISSUER"],
                        row["SECURITY_TYP"],
                    )
                    if cusip in hardcoded_usg_rating:
                        row[col] = "USG"
                    if cusip in hardcoded_usgcmo_rating:
                        row[col] = "USGCMO"
                    if row[col] == "US GCMO":  # hardwired ratings here
                        if cusip == "30297HAW1":
                            row[col] = "BBB"
                        if cusip == "055631HH1":
                            row[col] = "BB"
                    if row[col] == "NR":  # hardwired ratings here
                        print(cusip)
                        if cusip == "58960CAC3":
                            row[col] = "BBB"
                        elif cusip == "23802WAJ0":
                            row[col] = "BBB"
                        elif cusip == "64831HAN3":
                            row[col] = "AAA"
                        elif cusip == "31397VXD4":
                            row[col] = "USGCMO"
                        elif cusip == "29460XAC3":
                            row[col] = "BBB"
                        elif cusip == "03465BAA5":
                            row[col] = "AAA"
                        elif cusip == "50202KAA4":
                            row[col] = "AAA"
                        elif cusip == "822866AA7":
                            row[col] = "AAA"
                        elif cusip == "55284AAA6":
                            row[col] = "AAA"
                        elif cusip == "STHAPPLE1":
                            row[col] = "A"
                        elif cusip == "XS2373029748":
                            row[col] = "A"
                        elif cusip == "13080TAU6":
                            row[col] = "A"
                        elif cusip == "STHAPPLE2":
                            row[col] = "A"
                        elif cusip == "STHAPPLE3":
                            row[col] = "A"
                        elif cusip == "ALMNDUSD1":
                            row[col] = "A"
                        elif cusip == "ALMNDEUR1":
                            row[col] = "A"
                        elif cusip == "ALMNDEUR2":
                            row[col] = "A"
                        elif cusip == "ALMNDUSD4":
                            row[col] = "A"
                        elif cusip == "ALMNDUSD6":
                            row[col] = "A"
                        elif cusip == "ALMNDEUR3":
                            row[col] = "A"
                        elif cusip == "ALMNDEUR4":
                            row[col] = "A"
                        elif cusip == "ALMNDUSD5":
                            row[col] = "A"
                        elif cusip == "EASTCYPR1":
                            row[col] = "A"
                        elif cusip == "SOSPRUCE1":
                            row[col] = "A"
                        elif cusip == "SOSPRUCE2":
                            row[col] = "A"
                        elif cusip == "52953BBJ1":
                            row[col] = "A1/P1"
                        elif cusip == "20632C3S4":
                            row[col] = "A1/P1"
                        elif cusip == "20632C4C8":
                            row[col] = "A1/P1"
                        elif cusip == "20632C4G9":
                            row[col] = "A1/P1"
                        elif cusip == "20632C4W4":
                            row[col] = "A1/P1"
                        elif cusip == "2063C0AK8":
                            row[col] = "A1/P1"
                        elif cusip == "52468JDK7":
                            row[col] = "A1/P1"
                        elif cusip == "52953AAK1":
                            row[col] = "A1/P1"
                        elif cusip == "52953BZE6":
                            row[col] = "A1/P1"

                        elif cusip == "04942PAN7":
                            row[col] = "A"
                        elif cusip == "03465BAC1":
                            row[col] = "A"
                        elif cusip == "31422XSV5":
                            row[col] = "USG"
                        elif cusip == "3130ALYU2":
                            row[col] = "USG"
                        elif cusip == "3133ENXW5":
                            row[col] = "USG"
                elif col == "DELIVERY_TYP":
                    if len(cusip) == 12:
                        row[col] = "ECL"
                    else:
                        if (
                            row["DTC_REGISTERED"] == "Y"
                            or row["DTC_ELIGIBLE"] == "Y"
                            or row[col][:3] == "DTC"
                        ):
                            row[col] = "DTC"
                        elif row["RATINGS BUCKET"][:3] == "USG":
                            row[col] = "FED"
                        elif row["DTC_ELIGIBLE"] == "N":
                            row[col] = "CHECK SETTLEMENT"
                        else:
                            row[col] = "DTC"
                elif col == "Est'd Asset Class":
                    if cusip in mktsymbol_map.keys():
                        row[col] = mktsymbol_map[cusip]
                    else:
                        row[col] = helix_mkt_symbol(
                            row["SECURITY_TYP"],
                            row["ISSUER"],
                            row["Collat Typ"],
                            row["Issue DT"],
                            row["Name"],
                            row["Industry Sector"],
                            row["MTG TRANCHE TYP LONG"],
                            row["Floater"],
                            legacydate,
                        )
                elif col == "CUSIP or ISIN":
                    row[col] = " CUSIP" if len(cusip) == 9 else " ISIN"
                elif col == "correct_days":
                    try:
                        row["TRADE_DT_ACC_INT"] = (
                            row["Int Acc"]
                            if not row["TRADE_DT_ACC_INT"]
                            else round(float(row["TRADE_DT_ACC_INT"]), 6)
                        )
                    except:
                        row["TRADE_DT_ACC_INT"] = row["Int Acc"]
                    try:
                        row[col] = int(
                            row["Days Acc"] * row["TRADE_DT_ACC_INT"] / row["Int Acc"]
                        )  # for finding days accd using today as settle
                    except:
                        row[col] = row["Days Acc"]
                # finally, if column in the custom override df, default to that
                try:
                    if cusip not in custom_wals:
                        row[col] = df_custom_overrides.loc[cusip, col]
                        print("Using custom override " + col + " for " + cusip)
                except:
                    pass

            except Exception as e:
                print("Check " + cusip + ", " + col + ", " + str(e))

        row["Days Acc"] = row[
            "correct_days"
        ]  # after traversing all cols update the accruals cols
        row["Int Acc"] = row["TRADE_DT_ACC_INT"]
        newrows.append(row)
        # outp.loc[cusip] = row.tolist() # populate the output dataframe with the processed row
    # OAS spread ask seems closer to YB than with all of the complicated model params, defaulting to that.

    outp = pd.concat(newrows, axis=1).T  # transpose
    return outp[(cols.split(","))[1:]]


# construct the pni cusip rows
def fetch_pni_df(pni_cusips, other_df):
    outp = pd.DataFrame()
    for pnic in pni_cusips:
        undr = pnic[3:]
        if undr in other_df.index:  # "underlying" cusip should be in helix already
            outp = outp.append(
                [
                    {
                        "CUSIP": pnic,
                        "SECURITY_TYP": "PNI",
                        "ISSUER": other_df.at[undr, "ISSUER"],
                        "Collat Typ": "PNI",
                        "Name": "PNI",
                        "Industry Sector": "Mortgage Securities",
                        "Issue DT": other_df.at[undr, "Issue DT"],
                        "Maturity": other_df.at[undr, "Maturity"],
                        "Amt Outstanding": "1",
                        "Coupon": "0",
                        "Floater": "N",
                        "MTG Factor": "1",
                        "PX Bid": "100",
                        "PX Mid": "100",
                        "Int Acc": "0",
                        "Mtg WAL": "0",
                        "DUR ADJ OAS BID": "0",
                        "YAS_MOD_DUR": "0",
                        "USED DURATION": "0",
                        "Days Acc": "0",
                        "YLD_ytm_BID": "0",
                        "I_SPRD_BID": "0",
                        "FLT_SPREAD": "",
                        "OAS_SPREAD_ASK": "0",
                        "MTG TRANCHE TYP LONG": "PNI",
                        "MTG PL CPR 1M": "0",
                        "MTG PL CPR 6M": "0",
                        "MTG_WHLN_GEO1": "0",
                        "MTG_WHLN_GEO2": "0",
                        "MTG_WHLN_GEO3": "0",
                        "RATINGS BUCKET": "USG",
                        "RTG_SP": "",
                        "RTG_MOODY": "",
                        "RTG_FITCH": "",
                        "RTG_KBRA": "",
                        "RTG_DBRS": "",
                        "RTG_EGAN_JONES": "",
                        "DELIVERY_TYP": "FED",
                        "Est'd Asset Class": other_df.at[undr, "Est'd Asset Class"],
                        "CUSIP or ISIN": " ISIN",
                        "MTG_PREV_FACTOR": "1",
                        "MTG_RECORD_DT": "MTG_PREV_FACTOR",
                        "MTG_FACTOR_PAY_DT": "",
                        "MTG_NXT_PAY_DT_SET_DT": "",
                    }
                ]
            )
        else:
            print(pnic + " not here")
            outp = outp.append(
                [
                    {
                        "CUSIP": pnic,
                        "SECURITY_TYP": "PNI",
                        "ISSUER": "",
                        "Collat Typ": "PNI",
                        "Name": "PNI",
                        "Industry Sector": "Mortgage Securities",
                        "Issue DT": "",
                        "Maturity": "",
                        "Amt Outstanding": "1",
                        "Coupon": "0",
                        "Floater": "N",
                        "MTG Factor": "1",
                        "PX Bid": "100",
                        "PX Mid": "100",
                        "Int Acc": "0",
                        "Mtg WAL": "0",
                        "DUR ADJ OAS BID": "0",
                        "YAS_MOD_DUR": "0",
                        "USED DURATION": "0",
                        "Days Acc": "0",
                        "YLD_ytm_BID": "0",
                        "I_SPRD_BID": "0",
                        "FLT_SPREAD": "",
                        "OAS_SPREAD_ASK": "0",
                        "MTG TRANCHE TYP LONG": "PNI",
                        "MTG PL CPR 1M": "0",
                        "MTG PL CPR 6M": "0",
                        "MTG_WHLN_GEO1": "0",
                        "MTG_WHLN_GEO2": "0",
                        "MTG_WHLN_GEO3": "0",
                        "RATINGS BUCKET": "USG",
                        "RTG_SP": "",
                        "RTG_MOODY": "",
                        "RTG_FITCH": "",
                        "RTG_KBRA": "",
                        "RTG_DBRS": "",
                        "RTG_EGAN_JONES": "",
                        "DELIVERY_TYP": "FED",
                        "Est'd Asset Class": "",
                        "CUSIP or ISIN": " ISIN",
                        "MTG_PREV_FACTOR": "1",
                        "MTG_RECORD_DT": "MTG_PREV_FACTOR",
                        "MTG_FACTOR_PAY_DT": "",
                        "MTG_NXT_PAY_DT_SET_DT": "",
                    }
                ]
            )
    outp.set_index("CUSIP", inplace=True)
    return outp


if __name__ == "__main__":

    cusip_set = sys.argv[
        1
    ]  # changed from cusips to filepath to set of cusips so no need to overload cmd line Sept 2020
    bond_data_type = sys.argv[2]

    cusips_filepath = open(cusip_set, "r")
    input_cusips = []
    for c in cusips_filepath:
        input_cusips.append(c.strip())
    cusips_filepath.close()

    # these are the used market symbols in helix. not all cusips will be in helix; for those that aren't, we use the mkt_symbol function
    mktsymbol_map = fetch_helix_symbols()
    # some hardwired
    mktsymbol_map["05548WAJ6"] = "CMBS SINGLE PROPERTY"
    mktsymbol_map["30227FAL4"] = "CMBS SINGLE PROPERTY"
    mktsymbol_map["30227FAN0"] = "CMBS SINGLE PROPERTY"

    # reverse
    mktsymbol_map = {
        k: ("ABSOTHER" if v not in symbol_classifiers.keys() else symbol_classifiers[v])
        for k, v in mktsymbol_map.items()
    }

    if bond_data_type == "Helix":
        # include additional cusips to data-fetch
        addl_df = pd.read_excel(
            get_file_path("S:/Lucid/Data/Bond Data/Non-Collateral Cusips.xlsx"),
            skiprows=3,
        )
        input_cusips = (
            input_cusips
            + [x for x in addl_df["Vantage Proxies"] if str(x) != "nan"]
            + [x for x in addl_df["Other"] if str(x) != "nan"]
        )
        input_cusips = input_cusips + [x for x in diff_cusip_map]
        input_cusips = list(set(input_cusips))

        # fetch data for hardwired cusips
        special_bond_data = fetch_spec_df()

        # isolate cusips, changed 12/9/2019 to read from fetch_spec_df() bc should be consistent
        special_cusips = [x for x in special_bond_data.index]

        pni_cusips = [x for x in input_cusips if x[:3] == "PNI"]

        # don't call bloomberg for data on them
        bb_cusips = np.setdiff1d(
            input_cusips, special_cusips
        )  # remove special cusips from bb fetch proc
        bb_cusips = np.setdiff1d(
            bb_cusips, pni_cusips
        )  # remove pni dummy cusips from bb fetch proc
        # raw_df = bb_fetch(bb_cusips)
        raw_df_v2 = bb_fetch_v2(bb_cusips)
        df_custom_overrides = bb_fetch_with_overrides(mktsymbol_map)
        df_custom_overrides_v2 = bb_fetch_with_overrides_v2(mktsymbol_map)
        procd_df = process_bb_data(
            raw_df.fillna(""), mktsymbol_map, df_custom_overrides
        )  # crucial to replace the nan's with "" because changed isnull to falsy throughout the processing. did to avoid errors trying to stringify nan

        # append special cusips
        df_to_write = special_bond_data.append(procd_df, sort=False)
        # df_to_write = pd.concat([special_bond_data, procd_df], ignore_index=True, sort=False)

        # append pni cusips
        if len(pni_cusips) > 0:
            pni_df = fetch_pni_df(pni_cusips, df_to_write)
            df_to_write = pni_df.append(df_to_write, sort=False)
            # df_to_write = pd.concat([pni_df, df_to_write], ignore_index=True, sort=False)

        df_to_write = df_to_write[(cols.split(","))[1:]]
        df_to_write = hardwired_adjustments(df_to_write)
    elif bond_data_type == "Proxies":
        raw_df = bb_fetch(input_cusips)
        df_to_write = process_bb_data(
            raw_df.fillna(""), mktsymbol_map, pd.DataFrame()
        )  # crucial to replace the nan's with "" because changed isnull to falsy throughout the processing. did to avoid errors trying to stringify nan

    ## write to sheet

    print("Writing to bond data sheet...")
    vdate = datetime.now()

    if bond_data_type == "Helix":
        # then do helix and non-collateral cusips (prices fetched daily so bond data is too)
        currdest = get_file_path("S:/Lucid/Data/Bond Data/Bond Data.xlsx")
        savedeststr = get_file_path(
            "S:/Lucid/Data/Bond Data/Historical/Bond_Data_"
            + vdate.strftime("%m_%d_%Y")
            + "_AM.xlsx"
        )
    elif bond_data_type == "Proxies":
        # since prices fetched less often here needs to save down separately
        currdest = get_file_path(
            "S:/Lucid/Data/Bond Data/ProxySecurities Bond Data.xlsx"
        )

        savedeststr = get_file_path(
            "S:/Lucid/Data/Bond Data/Historical/ProxySecurities_Bond_Data_"
            + vdate.strftime("%m_%d_%Y")
            + ".xlsx"
        )

    savedest = savedeststr

    try:
        wb = op.load_workbook(currdest)
    except:
        print("file not found")
        exit()
    sht = wb["Bloomberg Data"]

    rn = 8
    acell = sht["A" + str(rn)]
    while acell.value != None:
        colnum = 1
        colcell = sht.cell(row=5, column=colnum)
        while colcell.value != None:
            sht.cell(rn, colnum).value = None
            colnum = colnum + 1
            colcell = sht.cell(row=5, column=colnum)
        rn = rn + 1
        acell = sht["A" + str(rn)]

    rn = 8
    for cusip in list(df_to_write.index.values):
        sht["A" + str(rn)].value = str(cusip)
        rn = rn + 1
    lastrow = rn
    colnum = 2
    colnamecell = sht.cell(row=5, column=colnum)
    while colnamecell.value != None:
        rn = 8
        cusipcell = sht["A" + str(rn)]
        while cusipcell.value != None:
            val = df_to_write.at[cusipcell.value, colnamecell.value]
            wrt = 0
            try:
                try:
                    wrt = float(val)
                except:  # susc to type or value error most likely
                    wrt = val
            except:
                wrt = 0
            sht.cell(row=rn, column=colnum).value = wrt
            rn = rn + 1
            cusipcell = sht["A" + str(rn)]
        colnum = colnum + 1
        colnamecell = sht.cell(row=5, column=colnum)

    sht["B3"].value = vdate.strftime("%m/%d/%Y")
    sht["B4"].value = vdate.strftime("%Y%m%d")
    sht["D2"].value = vdate.strftime("%m/%d/%Y %H:%M")
    sht["D3"].value = savedeststr

    wb.save(currdest)  # overwrite current
    wb.save(savedest)  # overwrite backup
    if bond_data_type == "Helix":
        savedeststrPM = get_file_path(
            "S:/Lucid/Data/Bond Data/Historical/Bond_Data_"
            + vdate.strftime("%m_%d_%Y")
            + "_PM.xlsx"
        )
        savedestPM = savedeststrPM
        wb.save(savedestPM)  # save PM file
    wb.close()
