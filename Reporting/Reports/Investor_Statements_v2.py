import openpyxl as op
import yaml
import logging
from datetime import datetime
from report_generator import generate_fund_report, generate_note_report, generate_pdf

def main():
    logging.basicConfig(level=logging.INFO)

    with open("config.yaml", "r") as f:
        config = yaml.safe_load(f)

    try:
        wb = op.load_workbook("Master Data.xlsx")
    except FileNotFoundError:
        logging.error("Master Data.xlsx not found.")
        return

    overview_ws = None
    for ws in wb.worksheets:
        if ws.title == "Platform Data":
            overview_ws = ws
            break

    if overview_ws is None:
        logging.warning("Platform Data worksheet not found.")

    reports_generated = []
    bad_reports = []

    for ws in wb.worksheets:
        try:
            if ws.title != "Mandate Template" and ws["B2"].value == "Mandate Data":
                report_name = ws.title
                logging.info(f"Processing {report_name}")

                if ws["C5"].value.upper() == "FUND":
                    pdf_file = generate_fund_report(ws, overview_ws, config)
                elif ws["C5"].value.upper() == "NOTE":
                    pdf_file = generate_note_report(ws, overview_ws, config)
                else:
                    logging.warning(f"Unsupported report type for {report_name}")
                    continue

                if pdf_file:
                    reports_generated.append(report_name)
                else:
                    bad_reports.append(report_name)

        except Exception as e:
            logging.exception(f"Error generating {report_name}")
            bad_reports.append(report_name)

    wb.close()

    logging.info("Generated reports:")
    for report in reports_generated:
        logging.info(report)

    if bad_reports:
        logging.warning("Failed to generate reports:")
        for report in bad_reports:
            logging.warning(report)

if __name__ == "__main__":
    main()