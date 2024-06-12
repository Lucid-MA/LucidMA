import platform
import subprocess
from datetime import datetime
from pathlib import Path, PureWindowsPath

import openpyxl as op
import pandas as pd

from Reporting.Utils.Common import get_file_path
from Reporting.Utils.Constants import (
    lucid_series,
    benchmark_shortern,
    reverse_cusip_mapping,
)
from Reporting.Utils.database_utils import get_database_engine, read_table_from_db
from Reports.Constants import fund_report_template, note_report_template
from Reports.Utils import (
    diff_period_rate,
    heightmap,
    stretches,
    hspacemap,
    xmap,
    barwidthmap,
    tablevstretch,
    form_as_percent,
    accs_since_start,
    month_wordify,
    benchmark_shorten,
    bps_spread,
)

# CONSTANT
reporting_series = [
    "PRIME-C10",
    "PRIME-M00",
    "PRIME-MIG",
    "PRIME-Q10",
    "PRIME-Q10",
    "PRIME-Q36",
    "PRIME-QX0",
    "USGFD-M00",
]

# Variable
# current_date = datetime.now()
current_date = datetime.strptime("2024-05-16", "%Y-%m-%d")
report_date_formal = current_date.strftime("%B %d, %Y")
report_date = current_date.strftime("%Y-%m-%d")
reporting_fund = reporting_series[-1]
reporting_fund_name = lucid_series[reporting_fund]

# Table names
db_type = "postgres"
attributes_table_name = "bronze_series_attributes"
historical_returns_table_name = "historical_returns"
target_return_table_name = "target_returns"
benchmark_table_name = "bronze_benchmark"
oc_rate_table_name = "oc_rates"
daily_nav_table_name = "bronze_daily_nav"
roll_schedule_table_name = "roll_schedule"
cash_balance_table_name = "bronze_cash_balance"
benchmark_comparison_table_name = "silver_return_by_series"

# Connect to the PostgreSQL database
engine = get_database_engine("postgres")

# Read the table into a pandas DataFrame
df_attributes = read_table_from_db(attributes_table_name, db_type)

df_historical_returns = read_table_from_db(historical_returns_table_name, db_type)

df_target_return = read_table_from_db(target_return_table_name, db_type)

df_benchmark = read_table_from_db(benchmark_table_name, db_type)

df_oc_rates = read_table_from_db(oc_rate_table_name, db_type)

df_daily_nav = read_table_from_db(daily_nav_table_name, db_type)

df_roll_schedule = read_table_from_db(roll_schedule_table_name, db_type)

df_cash_balance = read_table_from_db(cash_balance_table_name, db_type)

df_benchmark_comparison = read_table_from_db(benchmark_comparison_table_name, db_type)

## REPORTING VARIABLE ##
# DATES
roll_schedule_condition = df_roll_schedule["series_id"] == reporting_fund
df_roll_schedule = df_roll_schedule[roll_schedule_condition]


def get_current_reporting_dates(reporting_date):
    reporting_date = datetime.strptime(reporting_date, "%Y-%m-%d")
    current = df_roll_schedule[
        (df_roll_schedule["start_date"] <= reporting_date.strftime("%Y-%m-%d"))
        & (df_roll_schedule["end_date"] >= reporting_date.strftime("%Y-%m-%d"))
    ]
    if not current.empty:
        current_row = current.iloc[0]
        return (
            current_row["start_date"].strftime("%Y-%m-%d"),
            current_row["end_date"].strftime("%Y-%m-%d"),
            current_row["withdrawal_date"].strftime("%Y-%m-%d"),
            current_row["notice_date"].strftime("%Y-%m-%d"),
        )
    return (None, None, None, None)


def get_previous_reporting_dates(reporting_date):
    reporting_date = datetime.strptime(reporting_date, "%Y-%m-%d")
    previous = df_roll_schedule[
        df_roll_schedule["end_date"] < reporting_date.strftime("%Y-%m-%d")
    ]
    if not previous.empty:
        previous = previous.sort_values(by="end_date", ascending=False)
        previous_row = previous.iloc[0]
        return (
            previous_row["start_date"].strftime("%Y-%m-%d"),
            previous_row["end_date"].strftime("%Y-%m-%d"),
        )
    return (None, None)


def get_next_reporting_dates(reporting_date):
    reporting_date = datetime.strptime(reporting_date, "%Y-%m-%d")
    next_dates = df_roll_schedule[
        df_roll_schedule["start_date"] > reporting_date.strftime("%Y-%m-%d")
    ]
    if not next_dates.empty:
        next_dates = next_dates.sort_values(by="start_date")
        next_row = next_dates.iloc[0]
        return (
            next_row["start_date"].strftime("%Y-%m-%d"),
            next_row["end_date"].strftime("%Y-%m-%d"),
            next_row["withdrawal_date"].strftime("%Y-%m-%d"),
            next_row["notice_date"].strftime("%Y-%m-%d"),
        )
    return (None, None, None, None)


curr_start, curr_end, curr_withdrawal, curr_notice = get_current_reporting_dates(
    report_date
)
prev_start, prev_end = get_previous_reporting_dates(report_date)
next_start, next_end, next_withdrawal, next_notice = get_next_reporting_dates(
    report_date
)


# TARGET RETURN
curr_target_return_condition = (df_target_return["security_id"] == reporting_fund) & (
    df_target_return["date"] == curr_start
)
benchmark_name = df_target_return[curr_target_return_condition]["benchmark_name"].iloc[
    0
]
benchmark_short = benchmark_shortern[benchmark_name]
target_outperform_range = df_target_return[curr_target_return_condition][
    "target_range"
].iloc[0]
target_outperform_net = df_target_return[curr_target_return_condition][
    "net_spread"
].iloc[0]
benchmark = df_target_return[curr_target_return_condition]["benchmark"].iloc[0]

current_target_return = form_as_percent(benchmark, 2)


prev_target_return_condition = (df_target_return["security_id"] == reporting_fund) & (
    df_target_return["date"] == prev_start
)
prev_target_outperform = (
    str(df_target_return[prev_target_return_condition]["net_spread"].iloc[0]) + " bps"
)

# HISTORICAL RETURN
pool_name_encoded = reverse_cusip_mapping[reporting_fund]
historical_return_condition = (df_historical_returns["end_date"] == prev_end) & (
    df_historical_returns["pool_name"] == pool_name_encoded
)
df_historical_returns = df_historical_returns[historical_return_condition]
prev_return = form_as_percent(
    df_historical_returns["annualized_returns_360"].iloc[0], 2
)


# FUND ATTRIBUTES
fund_attribute_condition = df_attributes["security_id"] == reporting_fund
df_attributes = df_attributes[fund_attribute_condition]
fund_name = df_attributes["fund_name"].iloc[0]
series_name = df_attributes["series_name"].iloc[0]
expense_ratio_footnote_text = f"Fund Series expense ratio currently capped at an all-in ratio of {df_attributes['expense_ratio_cap'].iloc[0]} bps and can vary over time."
series_abbrev = df_attributes["series_abbreviation"].iloc[0]


# OC RATES
oc_rate_condition = (
    (df_oc_rates["fund"] == fund_name.upper())
    & (df_oc_rates["series"] == series_name.upper())
    & (df_oc_rates["report_date"] == report_date)
)
df_oc_rates = df_oc_rates[oc_rate_condition]

# CASH BALANCE
cash_balance_condition = (
    (df_cash_balance["Fund"] == fund_name.upper())
    & (df_cash_balance["Series"] == series_name.upper())
    & (df_cash_balance["Balance_date"] == report_date)
    & (df_cash_balance["Account"] == "MAIN")
)

df_cash_balance = df_cash_balance[cash_balance_condition]
cash_balance = df_cash_balance["Sweep_Balance"].iloc[0]


# RETURN COMPARISON
benchmark_comparison_condition = (
    df_benchmark_comparison["series_id"] == reporting_fund
) & (df_benchmark_comparison["start_date"] == curr_start)
df_benchmark_comparison = df_benchmark_comparison[benchmark_comparison_condition]

benchmark_dictionary = {
    "PRIME-C10":[],
    "PRIME-M00":["1m SOFR", "1m A1/P1 CP", "1m T-Bill"],
    "PRIME-MIG":[],
    "PRIME-Q10":[],
    "PRIME-Q36":[],
    "PRIME-QX0":[],
    "USGFD-M00":[],
}
r_a =
r_b =
r_c =


def calculate_oc_metrics(data):
    global cash_balance

    total_investment = data["investment_amount"].sum() + cash_balance

    def get_values(rating):
        if rating in data["rating_buckets"].values:
            row = data[data["rating_buckets"] == rating].iloc[0]
            return row["collateral_mv_allocated"], row["investment_amount"]
        return 0, 0

    col_mv_allocated_aaa, inv_aaa = get_values("AAA")
    col_mv_allocated_aa, inv_aa = get_values("AA")
    col_mv_allocated_a, inv_a = get_values("A")
    col_mv_allocated_bbb, inv_bbb = get_values("BBB")
    col_mv_allocated_usg, inv_usg = get_values("USG")
    col_mv_allocated_usgcmo, inv_usgcmo = get_values("USGCMO")

    oc_total = data["collateral_mv_allocated"].sum() / data["investment_amount"].sum()

    oc_usg_aaa = (
        (col_mv_allocated_aaa + col_mv_allocated_usg + col_mv_allocated_usgcmo)
        / (inv_aaa + inv_usg + inv_usgcmo)
        if (inv_aaa + inv_usg + inv_usgcmo) != 0
        else 0
    )
    oc_aa_a = (
        (col_mv_allocated_aa + col_mv_allocated_a) / (inv_aa + inv_a)
        if (inv_aa + inv_a) != 0
        else 0
    )
    oc_bbb = col_mv_allocated_bbb / inv_bbb if inv_bbb != 0 else 0
    oc_tbills = 0

    aloc_usg_aaa = (inv_aaa + inv_usg + inv_usgcmo) / total_investment
    aloc_aa_a = (inv_aa + inv_a) / total_investment
    aloc_bbb = inv_bbb / total_investment
    aloc_tbills = cash_balance / total_investment

    return (
        oc_total,
        oc_usg_aaa,
        oc_aa_a,
        oc_bbb,
        oc_tbills,
        aloc_usg_aaa,
        aloc_aa_a,
        aloc_bbb,
        aloc_tbills,
    )


(
    oc_total,
    oc_usg_aaa,
    oc_aa_a,
    oc_bbb,
    oc_tbills,
    aloc_usg_aaa,
    aloc_aa_a,
    aloc_bbb,
    aloc_tbills,
) = calculate_oc_metrics(df_oc_rates)


def get_fund_size(fund_name, report_date):
    # use df_daily_nav
    return


def get_series_size(fund_name, report_date):
    # use df_daily_nav
    return


def get_aum(report_date):
    return


# intialize script templates

# only thing hardwired: descriptions for various entities. currently most match but might change in future
fund_descriptions = dict()
series_descriptions = dict()
fund_descriptions[
    "USG"
] = r"""The fund series seeks income generation with 100\% capital preservation and invests solely in US government backed securities (USG) and repurchase agreements secured by USG, subject to the proprietary Lucid Investment Process.
"""

fund_descriptions[
    "PRIME"
] = r"""
The fund series seeks income generation with 100\% capital preservation and primarily invests in repurchase agreements secured by US Government Backed securities (USG) and Eligible Collateral Securities as well as other Eligible Investments (T-Bills, A1/P1 Commercial Paper and money market funds).
"""

custom_qx_desc = r"""
The fund series seeks income generation with 100\% capital preservation and primarily invests in repurchase agreements limited to Investment Grade and BB rated Collateral Securities (subject to a 50\% limit on BB collateral), Single-A Bank Guaranteed Collateral or other high quality money market securities. 
"""

series_descriptions[
    "USGFUND M"
] = r"""
\textbf{{Series M}} portfolio assets are limited to 1 month maximum maturities.
"""

series_descriptions[
    "PRIMEFUND M"
] = r"""
\textbf{{Series M}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Highly Rated Investment Grade securities (at least 75\% rated between AAA and A- or USG securities).
"""

series_descriptions[
    "PRIMEFUND M1"
] = r"""
\textbf{{Series M1}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Highly Rated Investment Grade securities (at least 75\% rated between AAA and A- or USG securities).
"""

series_descriptions[
    "PRIMEFUND C1"
] = r"""
\textbf{{Series C1}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Highly Rated Investment Grade securities (at least 75\% rated between AAA and A- or USG securities).
"""

series_descriptions[
    "PRIMEFUND Q1"
] = r"""
\textbf{{Series Q1}} interests have (i) maximum maturities of 3 months on all series assets, (ii) quarterly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Investment Grade securities only or USG securities.
"""

# series_descriptions["PRIMEFUND QX"] = r"""
# \textbf{{Series QX}} interests have (i) maximum maturities of 3 months on all series assets and (ii) quarterly withdrawal dates.
# """

# Terri Change 02/01/2024

series_descriptions[
    "PRIMEFUND QX"
] = r"""
\textbf{{Series QX}} # interests have (i) maximum maturities of 3 months on all series assets, (ii) monthly withdrawal dates.                                                                           
"""

series_descriptions[
    "PRIMEFUND MIG"
] = r"""
\textbf{{Series MIG}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Investment Grade securities only or USG securities.
"""

# intervals for various entities, default to 3 months/1 year unless specified
intervals = dict()
intervals["PRIMEFUND Q1"] = (6, 12)
intervals["PRIMEFUND QX"] = (6, 12)

# disclaimer for reduced 12/30-1/28 return
m1_1230_128_rets_disclaimer = r"""\\
	
	\noindent\textit{{* Returns for the period from 12/30/20 to 1/28/21 reflects the expected impact of sizable allocations of investor cash into the M1 Series after the notice date, with manager and series investor consent. For amounts invested within the series guidelines and notice dates, returns are expected to remain in line with overall M1 series targets.}}
"""


# graph dimension adjustments

# fund & series details , portfolio comp table array stretches


# space between bars in bar chart (mm)
# Terri 02/05/2024 change spacing on M graph from 1.7 to
# return [1.7,1.7,1.7,1.36,0.942,0.674,0.558,0.471,0.395,0.35,0.314,0.103,0.255,0.234,0.215,0.198,0.188][nbars]
# return 0.156 # default here, if 16

# width of each bar in bar chart (mm) Terri was 2.50 default


# for notes


def return_table_plot(
    fund_name,
    prev_pd_return,
    series_abbrev,
    r_this_1,
    r_this_2,
    comp_a,
    comp_b,
    comp_c,
    r_a,
    r_b,
    r_c,
    s_a_0,
    s_a_1,
    s_a_2,
    s_b_0,
    s_b_1,
    s_b_2,
    s_c_0,
    s_c_1,
    s_c_2,
):
    if fund_name == "USG":
        out = r"""
	\textbf{{Lucid {fund_name} - Series {series_abbrev}}}                    & \textbf{{{prev_pd_return}}}                              & \textbf{{-}}                                  & \textbf{{{r_this_1}}}                               & \textbf{{-}}                           & \textbf{{{r_this_2}}}                             & \textbf{{-}}                          \\
{comp_a}                       & {r_a_0}                                       & \textbf{{{s_a_0}}}                            & {r_a_1}                               & \textbf{{{s_a_1}}}                     & {r_a_2}                              & \textbf{{{s_a_2}}}                    \\
{comp_b}                       & {r_b_0}                                       & \textbf{{{s_b_0}}}                           & {r_b_1}                               & \textbf{{{s_b_1}}}                     & {r_b_2}                              & \textbf{{{s_b_2}}}                    \\ \arrayrulecolor{{light_grey}}\hline
	"""
        return out.format(
            fund_name=fund_name,
            prev_pd_return=prev_pd_return,
            series_abbrev=series_abbrev,
            r_this_1=r_this_1,
            r_this_2=r_this_2,
            comp_a=comp_a,
            comp_b=comp_b,
            r_a_0=form_as_percent(r_a[0], 2),
            r_a_1=r_a[1],  # already percent
            r_a_2=r_a[2],
            r_b_0=form_as_percent(r_b[0], 2),
            r_b_1=r_b[1],
            r_b_2=r_b[2],
            s_a_0=s_a_0,
            s_a_1=s_a_1,
            s_a_2=s_a_2,
            s_b_0=s_b_0,
            s_b_1=s_b_1,
            s_b_2=s_b_2,
        )
    else:
        out = r"""
	\textbf{{Lucid {fund_name} - Series {series_abbrev}}}                    & \textbf{{{prev_pd_return}}}                              & \textbf{{-}}                                  & \textbf{{{r_this_1}}}                               & \textbf{{-}}                           & \textbf{{{r_this_2}}}                             & \textbf{{-}}                          \\
{comp_a}                       & {r_a_0}                                       & \textbf{{{s_a_0}}}                            & {r_a_1}                               & \textbf{{{s_a_1}}}                     & {r_a_2}                              & \textbf{{{s_a_2}}}                    \\
{comp_b}                       & {r_b_0}                                       & \textbf{{{s_b_0}}}                           & {r_b_1}                               & \textbf{{{s_b_1}}}                     & {r_b_2}                              & \textbf{{{s_b_2}}}                    \\
{comp_c}                       & {r_c_0}                                       & \textbf{{{s_c_0}}}                            & {r_c_1}                               & \textbf{{{s_c_1}}}                     & {r_c_2}                              & \textbf{{{s_c_2}}}                    \\ \arrayrulecolor{{light_grey}}\hline
	"""
        return out.format(
            fund_name=fund_name,
            prev_pd_return=prev_pd_return,
            series_abbrev=series_abbrev,
            r_this_1=r_this_1,
            r_this_2=r_this_2,
            comp_a=comp_a,
            comp_b=comp_b,
            comp_c=comp_c,
            r_a_0=form_as_percent(r_a[0], 2),
            r_a_1=r_a[1],  # already percent
            r_a_2=r_a[2],
            r_b_0=form_as_percent(r_b[0], 2),
            r_b_1=r_b[1],
            r_b_2=r_b[2],
            r_c_0=form_as_percent(r_c[0], 2),
            r_c_1=r_c[1],
            r_c_2=r_c[2],
            s_a_0=s_a_0,
            s_a_1=s_a_1,
            s_a_2=s_a_2,
            s_b_0=s_b_0,
            s_b_1=s_b_1,
            s_b_2=s_b_2,
            s_c_0=s_c_0,
            s_c_1=s_c_1,
            s_c_2=s_c_2,
        )


def addl_coll_breakdown(alloc_aa_a, oc_aa_a, alloc_bbb, oc_bbb, alloc_bb, oc_bb):
    if alloc_aa_a == "n/a":
        return ""
    else:
        outp = r"""IG Repo: AA to A & {alloc_aa_a} & {oc_aa_a} \\
		IG Repo: BBB & {alloc_bbb} & {oc_bbb} \\
		Repo: BB & {alloc_bb} & {oc_bb} \\"""
        return outp.format(
            alloc_aa_a=alloc_aa_a,
            oc_aa_a=oc_aa_a,
            alloc_bbb=alloc_bbb,
            oc_bbb=oc_bbb,
            alloc_bb=alloc_bb,
            oc_bb=oc_bb,
        )


def colltable(
    inclextra,
    secured_by,
    series_abbrev,
    include_aaa_in_usg_bucket,
    alloc_aaa,
    alloc_aa_a,
    alloc_bbb,
    alloc_bb,
    alloc_tbills,
    alloc_total,
    oc_total,
    oc_tbills,
    oc_bb,
    oc_bbb,
    oc_aa_a,
    oc_aaa,
):
    if inclextra:
        out = r"""
		\renewcommand{{\arraystretch}}{{{vstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
		>{{\columncolor[HTML]{{EFEFEF}}}}p{{3.5cm}} 
		>{{\columncolor[HTML]{{EFEFEF}}}}c
		>{{\columncolor[HTML]{{EFEFEF}}}}c!{{\color{{light_grey}}\vrule}}}}
		\arrayrulecolor{{light_grey}}\hline
		\multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{\color[HTML]{{FFFFFF}}\textbf{{Series Collateral Overview\textsuperscript{{4}}}}}}}} \\
		\multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}p{{8.2cm}}!{{\color{{light_grey}}\vrule}}}}{{\rowcolor[HTML]{{EFEFEF}}{{\textbf{{Series {series_abbrev}}}: Secured by \textbf{{{secured_by}}}, with daily valuations \& margining.}}}} \\
		& & \\
		 & \textbf{{\% Portfolio}} & \textbf{{O/C Rate}}\\
		{usg_aaa_cat} & {alloc_aaa} & {oc_aaa} \\
		IG Repo: AA to A & {alloc_aa_a} & {oc_aa_a} \\
		IG Repo: BBB & {alloc_bbb} & {oc_bbb} \\
		Repo: BB & {alloc_bb} & {oc_bb} \\
		T-Bills; Gov't MMF & {alloc_tbills} & {oc_tbills} \\ \cline{{2-2}} \cline{{3-3}} 
		\textbf{{Total}} & {alloc_total} & \textbf{{{oc_total}}} \\\arrayrulecolor{{light_grey}}\hline
		\end{{tabular}}
		"""
        return out.format(
            vstretch=1.48,
            secured_by=secured_by,
            series_abbrev=series_abbrev,
            usg_aaa_cat=(
                "US Govt/AAA Repo" if include_aaa_in_usg_bucket else "US Govt Repo"
            ),
            alloc_aaa=alloc_aaa,
            alloc_aa_a=alloc_aa_a,
            alloc_bbb=alloc_bbb,
            alloc_bb=alloc_bb,
            alloc_tbills=alloc_tbills,
            alloc_total=alloc_total,
            oc_total=oc_total,
            oc_tbills=oc_tbills,
            oc_bb=oc_bb,
            oc_bbb=oc_bbb,
            oc_aa_a=oc_aa_a,
            oc_aaa=oc_aaa,
        )
    else:
        out = r"""
		\renewcommand{{\arraystretch}}{{{vstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
		>{{\columncolor[HTML]{{EFEFEF}}}}p{{3.5cm}} 
		>{{\columncolor[HTML]{{EFEFEF}}}}c
		>{{\columncolor[HTML]{{EFEFEF}}}}c!{{\color{{light_grey}}\vrule}}}}
		\arrayrulecolor{{light_grey}}\hline
		\multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{\color[HTML]{{FFFFFF}}\textbf{{Series Collateral Overview\textsuperscript{{4}}}}}}}} \\
		\multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}p{{8.2cm}}!{{\color{{light_grey}}\vrule}}}}{{\rowcolor[HTML]{{EFEFEF}}{{\textbf{{Series {series_abbrev}}}: Secured by \textbf{{{secured_by}}}, with daily valuations \& margining.}}}} \\
		& & \\
		 & \textbf{{\% Portfolio}} & \textbf{{O/C Rate}}\\
		{usg_aaa_cat} & {alloc_aaa} & {oc_aaa} \\
		T-Bills; Gov't MMF & {alloc_tbills} & {oc_tbills} \\ \cline{{2-2}} \cline{{3-3}} 
		\textbf{{Total}} & {alloc_total} & \textbf{{{oc_total}}} \\\arrayrulecolor{{light_grey}}\hline
		\end{{tabular}}
		"""
        return out.format(
            vstretch=1.91,
            secured_by=secured_by,
            series_abbrev=series_abbrev,
            usg_aaa_cat=(
                "US Govt/AAA Repo" if include_aaa_in_usg_bucket else "US Govt Repo"
            ),
            alloc_aaa=alloc_aaa,
            # alloc_cp=alloc_cp,
            alloc_tbills=alloc_tbills,
            alloc_total=alloc_total,
            oc_total=oc_total,
            oc_tbills=oc_tbills,
            # oc_cp=oc_cp,
            oc_aaa=oc_aaa,
        )


def plotify(ws, x_col, y_col, start, end):
    outp = ""
    for row in range(start, end + 1):
        if ws[x_col + str(row)].value and ws[y_col + str(row)].value:
            outp = (
                outp
                + "("
                + ws[x_col + str(row)].value.strftime("%Y-%m-%d")
                + ","
                + str(round(100 * ws[y_col + str(row)].value, 3))
                + ") "
            )
    return outp


# for performance vs benchmark charts, for series with less than 3 periods data
def snapshot_graph(
    hspace,
    graphwidth,
    graphheight,
    maxreturn,
    series_abbrev,
    comp_a,
    comp_b,
    comp_c,
    this_r,
    ra,
    rb,
    rc,
):
    if comp_c is not None:  # when we have 3 comparables - for all prime series
        out = r"""
		  \hspace*{{{hspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
	\begin{{axis}}[
		title style = {{font = \small}},
		axis line style = {{light_grey}},
		title={{{{Performance vs Benchmarks}}}},
		ymin=3, ymax={maxreturn}, %MAXRETURN HERE
	   symbolic x coords={{Series {series_abbrev},{comp_a},{comp_b},{comp_c}}},
		xtick={{Series {series_abbrev},{comp_a},{comp_b},{comp_c}}},
		x tick label style={{anchor=north,font=\scriptsize,/pgf/number format/assume math mode}},
		yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
		y tick label style = {{/pgf/number format/.cd,
				fixed,
				fixed zerofill,
				precision=2,
				/pgf/number format/assume math mode
		}},
		ytick distance={tickdist},
		bar width = 8mm, x = 2.46cm,
		xtick pos=bottom,ytick pos=left,
		every node near coord/.append style={{font=\fontsize{{8}}{{8}}\selectfont,/pgf/number format/.cd,
				fixed,
				fixed zerofill,
				precision=2,/pgf/number format/assume math mode}},
		]
	\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			(Series {series_abbrev},{this_r}) 
		}};
	\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			({comp_a},{ra}) 
		}};
	\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			({comp_b},{rb}) 
		}};
	\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			({comp_c},{rc}) 
		}};
	\end{{axis}}
		\end{{tikzpicture}}}}

		"""
        return out.format(
            hspace=hspace,
            tickdist=0.5 if max(this_r, ra, rb, rc) + 0.1 >= 1 else 0.25,
            graphwidth=graphwidth,
            graphheight=graphheight,
            maxreturn=max(this_r, ra, rb, rc) + 1,
            series_abbrev=series_abbrev,
            comp_a=benchmark_shorten(comp_a) if "CRANE" in comp_a.upper() else comp_a,
            comp_b=benchmark_shorten(comp_b) if "CRANE" in comp_b.upper() else comp_b,
            comp_c=benchmark_shorten(comp_c) if "CRANE" in comp_c.upper() else comp_c,
            this_r=this_r,
            ra=ra,
            rb=rb,
            rc=rc,
        )
    else:  # when we only have 2 comparables - for all usg
        out = r"""
		  \hspace*{{{hspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
	\begin{{axis}}[
		title style = {{font = \small}},
		axis line style = {{light_grey}},
			title={{{{Performance vs Benchmarks}}}},
		ymin=2, ymax={maxreturn}, %MAXRETURN HERE
	   symbolic x coords={{Series {series_abbrev},{comp_a},{comp_b}}},
		xtick={{Series {series_abbrev},{comp_a},{comp_b}}},
		x tick label style={{anchor=north,font=\scriptsize,/pgf/number format/assume math mode}},
		yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
		y tick label style = {{/pgf/number format/.cd,
				fixed,
				fixed zerofill,
				precision=2,
				/pgf/number format/assume math mode
		}},
		ytick distance={tickdist},
		bar width = 10mm, x = 3.7cm,
		xtick pos=bottom,ytick pos=left,
		every node near coord/.append style={{font=\fontsize{{8}}{{8}}\selectfont,/pgf/number format/.cd,
				fixed,
				fixed zerofill,
				precision=2,/pgf/number format/assume math mode}},
		]
	\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			(Series {series_abbrev},{this_r}) 
		}};
	\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			({comp_a},{ra}) 
		}};
	\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			({comp_b},{rb}) 
		}};
	\end{{axis}}
		\end{{tikzpicture}}}}
	
		"""
        return out.format(
            hspace=hspace,
            tickdist=0.5 if maxreturn >= 1 else 0.25,
            graphwidth=graphwidth,
            graphheight=graphheight,
            maxreturn=max(this_r, ra, rb, rc) + 1,
            series_abbrev=series_abbrev,
            comp_a=benchmark_shorten(comp_a) if "CRANE" in comp_a.upper() else comp_a,
            comp_b=benchmark_shorten(comp_b) if "CRANE" in comp_b.upper() else comp_b,
            this_r=this_r,
            ra=ra,
            rb=rb,
        )


def performance_graph(
    titleincl,
    graphhspace,
    graphwidth,
    graphheight,
    fund_name,
    zero_date,
    min_return,
    max_return,
    graphx,
    graphbarwidth,
    return_plot,
    comp_a_plot,
    comp_b_plot,
    series_abbrev,
    comp_a,
    comp_b,
):
    if fund_name.upper() == "USG":
        out = r"""
			  \hspace*{{{graphhspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
		\begin{{axis}}[
			title style = {{font = \small}},
			axis line style = {{light_grey}},{title}
			date coordinates in=x, date ZERO={zero_date},
			xticklabel=\month/\day/\year,  
			ymin={min_return}, ymax={max_return}, %MAXRETURN HERE
			legend cell align = {{left}},
			legend style={{at={{(0.3,1)}},
			  anchor=north east, font=\tiny, draw=none,fill=none}},
			  x={graphx}mm, %CHANGE THIS to tighten in graph, eg if quarterly
			bar width={graphbarwidth}mm, ybar=2pt, %bar width is width, ybar is space between
		   % symbolic x coords={{Firm 1, Firm 2, Firm 3, Firm 4, Firm 5}},
			xtick=data,
			x tick label style={{rotate=90,anchor=east,font=\tiny,/pgf/number format/assume math mode}},
				 yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
			y tick label style = {{/pgf/number format/.cd,
					fixed,
					fixed zerofill,
					precision=2,
					/pgf/number format/assume math mode
			}},
			nodes near coords align={{vertical}},
			ytick distance=0.5,
			xtick pos=bottom,ytick pos=left,
			every node near coord/.append style={{font=\fontsize{{6}}{{6}}\selectfont,/pgf/number format/.cd,
					fixed,
					fixed zerofill,
					precision=2,/pgf/number format/assume math mode}},
			]
		%\addplot[ybar, nodes near coords, fill=blue] 
		\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
			coordinates {{
				{return_plot}
			}};
		\addplot[draw=dark_red,ultra thick,smooth] 
			coordinates {{
				{comp_a_plot}
			}};
		\addplot[draw=dark_color,ultra thick,smooth] 
			coordinates {{
				{comp_b_plot}
			}};
		\legend{{\hphantom{{A}}{fund_name} Series {series_abbrev},\hphantom{{A}}{comp_a},\hphantom{{A}}{comp_b}}}
		\end{{axis}}
			\end{{tikzpicture}}}}

			"""
    else:
        out = r"""
		  \hspace*{{{graphhspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
	\begin{{axis}}[
		title style = {{font = \small}},
		axis line style = {{light_grey}},{title}
		date coordinates in=x, date ZERO={zero_date},
		xticklabel=\month/\day/\year,  
		ymin={min_return}, ymax={max_return}, %MAXRETURN HERE
		legend cell align = {{left}},
		legend style={{at={{(0.2,1)}},
		  anchor=north east, font=\tiny, draw=none,fill=none}},
		  x={graphx}mm, %CHANGE THIS to tighten in graph, eg if quarterly
		bar width={graphbarwidth}mm, ybar=2pt, %bar width is width, ybar is space between
	   % symbolic x coords={{Firm 1, Firm 2, Firm 3, Firm 4, Firm 5}},
		xtick=data,
		x tick label style={{rotate=90,anchor=east,font=\tiny,/pgf/number format/assume math mode}},
			 yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
		y tick label style = {{/pgf/number format/.cd,
				fixed,
				fixed zerofill,
				precision=2,
				/pgf/number format/assume math mode
		}},
		nodes near coords align={{vertical}},
		ytick distance=0.5,
		xtick pos=bottom,ytick pos=left,
		every node near coord/.append style={{font=\fontsize{{6}}{{6}}\selectfont,/pgf/number format/.cd,
				fixed,
				fixed zerofill,
				precision=2,/pgf/number format/assume math mode}},
		]
	%\addplot[ybar, nodes near coords, fill=blue] 
	\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
		coordinates {{
			{return_plot}
		}};
	\addplot[draw=dark_red,ultra thick,smooth] 
		coordinates {{
			{comp_a_plot}
		}};
	\addplot[draw=dark_color,ultra thick,smooth] 
		coordinates {{
			{comp_b_plot}
		}};
	\legend{{\hphantom{{A}}{fund_name} Series {series_abbrev},\hphantom{{A}}{comp_a},\hphantom{{A}}{comp_b}}}
	\end{{axis}}
		\end{{tikzpicture}}}}
	
		"""
    # Terri graph
    return out.format(
        graphhspace=graphhspace,
        graphwidth=graphwidth,
        graphheight=graphheight,
        fund_name=fund_name,
        title=(
            r"""
		title={{Performance vs Benchmark}},"""
            if titleincl
            else ""
        ),
        zero_date=zero_date,
        min_return=min_return,
        max_return=max_return,
        graphx=graphx,
        graphbarwidth=graphbarwidth,
        return_plot=return_plot,
        comp_a_plot=comp_a_plot,
        comp_b_plot=comp_b_plot,
        series_abbrev=series_abbrev,
        comp_a=comp_a,
        comp_b=comp_b,
    )


def performance_graph_v2(
    titleincl,
    graphhspace,
    graphwidth,
    graphheight,
    fund_name,
    zero_date,
    max_return,
    graphx,
    graphbarwidth,
    return_plot,
    comp_a_plot,
    comp_b_plot,
    series_abbrev,
    comp_a,
    comp_b,
):
    if fund_name.upper() == "USG":
        legend_x = 0.3
        legend_y = 1.0
    else:
        legend_x = 0.2
        legend_y = 1.0

    out = r"""
	  \hspace*{{{graphhspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
\begin{{axis}}[
	title style = {{font = \small}},
	axis line style = {{light_grey}},{title}
	date coordinates in=x, date ZERO={zero_date},
	xticklabel=\month/\day/\year,  
	ymin=0, ymax={max_return}, %MAXRETURN HERE
	legend cell align = {{left}},
	legend style={{at={{{legend_x},{legend_y}}},
	  anchor=north east, font=\tiny, draw=none,fill=none}},
	  x={graphx}mm, %CHANGE THIS to tighten in graph, eg if quarterly
	bar width={graphbarwidth}mm, ybar=2pt, %bar width is width, ybar is space between
   % symbolic x coords={{Firm 1, Firm 2, Firm 3, Firm 4, Firm 5}},
	xtick=data,
	x tick label style={{rotate=90,anchor=east,font=\tiny,/pgf/number format/assume math mode}},
		 yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
	y tick label style = {{/pgf/number format/.cd,
			fixed,
			fixed zerofill,
			precision=2,
			/pgf/number format/assume math mode
	}},
	nodes near coords align={{vertical}},
	ytick distance=0.5,
	xtick pos=bottom,ytick pos=left,
	every node near coord/.append style={{font=\fontsize{{6}}{{6}}\selectfont,/pgf/number format/.cd,
			fixed,
			fixed zerofill,
			precision=2,/pgf/number format/assume math mode}},
	]
%\addplot[ybar, nodes near coords, fill=blue] 
\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
	coordinates {{
		{return_plot}
	}};
\addplot[draw=dark_red,ultra thick,smooth] 
	coordinates {{
		{comp_a_plot}
	}};
\addplot[draw=dark_color,ultra thick,smooth] 
	coordinates {{
		{comp_b_plot}
	}};
\legend{{\hphantom{{A}}{fund_name} Series {series_abbrev},\hphantom{{A}}{comp_a},\hphantom{{A}}{comp_b}}}
\end{{axis}}
	\end{{tikzpicture}}}}

	"""
    # Terri graph
    return out.format(
        graphhspace=graphhspace,
        graphwidth=graphwidth,
        graphheight=graphheight,
        fund_name=fund_name,
        title=(
            r"""
		title={{Performance vs Benchmark}},"""
            if titleincl
            else ""
        ),
        zero_date=zero_date,
        max_return=max_return,
        graphx=graphx,
        graphbarwidth=graphbarwidth,
        return_plot=return_plot,
        comp_a_plot=comp_a_plot,
        comp_b_plot=comp_b_plot,
        series_abbrev=series_abbrev,
        comp_a=comp_a,
        comp_b=comp_b,
    )


# assumes each is in format "-4.342 \\%"


print("Fetching data...")

try:
    MASTER_FILEPATH = get_file_path("S:/Mandates/Funds/Fund Reporting/Master Data.xlsx")
    # MASTER_FILEPATH = "C:/Users/Lucid Trading/Desktop/tmp_trash/Master Data.xlsx"
    if platform.system() == "Darwin":  # macOS
        wb = op.load_workbook(Path(MASTER_FILEPATH))
    elif platform.system() == "Windows":
        wb = op.load_workbook(PureWindowsPath(Path(MASTER_FILEPATH)))

except:
    print("Error fetching data.")
    exit()

reports_generated = []
bad_reports = []

# find overview page
bigsheet = ""
for ws in wb.worksheets:
    if ws.title == "Platform Data":
        bigsheet = ws

usg_rets = []
usg_rets_mid = []
primem_rets = []

for ws in wb.worksheets:
    try:
        if (ws.title != "Mandate Template") and (ws["B2"].value == "Mandate Data"):
            report_name = ws.title
            if not (ws["C29"].value):
                continue

            print("*****" + report_name + "*****")
            crow = 7

            # find relevant period row
            while ws["F" + str(crow)].value:
                if ws["F" + str(crow)].value == ws["C23"].value:
                    break
                crow = crow + 1
            # crow = 88

            prev_pd_start = prev_start
            this_pd_start = curr_start
            print(
                "For period "
                + prev_pd_start.strftime("%m/%d")
                + " - "
                + this_pd_start.strftime("%m/%d")
            )

            overview_row = 7
            while bigsheet["B" + str(overview_row)].value:
                if bigsheet["B" + str(overview_row)].value == this_pd_start:
                    break
                overview_row = overview_row + 1
            if not bigsheet["B" + str(overview_row)].value:
                print("ERROR: Overview row not found for this period. Continuing...")
                continue

            lucid_aum = bigsheet["H" + str(overview_row)].value  # post sub/redemp
            program_size = 0
            for col in "CDEFG":
                if bigsheet[col + "6"].value:
                    if bigsheet[col + "6"].value.upper() == ws["C9"].value.upper():
                        program_size = bigsheet[
                            col + str(overview_row)
                        ].value  # post sub/redemp...
                        break

            # returns for each comparable, a/b/c, just taken in order from cols on sheet
            daycount = ws["C25"].value
            interval_tuple = (3, 12)
            if "QUARTERLY" in ws["C15"].value.upper():
                interval_tuple = (2, 4)  # because one row = 3 months
            r_a = (
                ws["Y" + str(crow)].value,
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[0])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Y", "E", "F", crow - interval_tuple[0], 7, daycount
                        ),
                        accs_since_start(ws, "Y", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[0] >= 7)
                    else "n/a"
                ),
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[1])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Y", "E", "F", crow - interval_tuple[1], 7, daycount
                        ),
                        accs_since_start(ws, "Y", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[1] >= 7)
                    else "n/a"
                ),
            )
            r_b = (
                ws["Z" + str(crow)].value,
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[0])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Z", "E", "F", crow - interval_tuple[0], 7, daycount
                        ),
                        accs_since_start(ws, "Z", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[0] >= 7)
                    else "n/a"
                ),
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[1])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Z", "E", "F", crow - interval_tuple[1], 7, daycount
                        ),
                        accs_since_start(ws, "Z", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[1] >= 7)
                    else "n/a"
                ),
            )
            r_c = (
                ws["AA" + str(crow)].value,
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[0])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "AA", "E", "F", crow - interval_tuple[0], 7, daycount
                        ),
                        accs_since_start(ws, "AA", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[0] >= 7)
                    else "n/a"
                ),
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[1])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "AA", "E", "F", crow - interval_tuple[1], 7, daycount
                        ),
                        accs_since_start(ws, "AA", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[1] >= 7)
                    else "n/a"
                ),
            )
            r_this_1 = (
                diff_period_rate(
                    ws["F" + str(crow - interval_tuple[0])].value,
                    this_pd_start,
                    daycount,
                    accs_since_start(
                        ws,
                        "N" if daycount == 360 else "O",
                        "E",
                        "F",
                        crow - interval_tuple[0],
                        7,
                        daycount,
                    ),
                    accs_since_start(
                        ws, "N" if daycount == 360 else "O", "E", "F", crow, 7, daycount
                    ),
                )
                if (crow - interval_tuple[0] >= 7)
                else "n/a"
            )

            r_this_2 = (
                diff_period_rate(
                    ws["F" + str(crow - interval_tuple[1])].value,
                    this_pd_start,
                    daycount,
                    accs_since_start(
                        ws,
                        "N" if daycount == 360 else "O",
                        "E",
                        "F",
                        crow - interval_tuple[1],
                        7,
                        daycount,
                    ),
                    accs_since_start(
                        ws, "N" if daycount == 360 else "O", "E", "F", crow, 7, daycount
                    ),
                )
                if (crow - interval_tuple[1] >= 7)
                else "n/a"
            )

            # plotting info
            ts_row_start = max(7, crow - 15)  #######################################

            ts_row_end = crow
            if "QUARTERLY" in ws["C15"].value.upper():
                interval_tuple = (6, 12)  # now revise to months count

            if ws["C6"].value.upper() == "USGFUND M":
                usg_rets = [r_this_2, r_a[2], r_b[2], r_c[2]]
                usg_rets_mid = [r_this_1, r_a[1], r_b[1], r_c[1]]
                print(usg_rets)
            if ws["C6"].value.upper() == "PRIMEFUND M":
                primem_rets = [r_this_2, r_a[2], r_b[2], r_c[2]]
                print("Storing...")
                print(primem_rets)

            if ws["C5"].value.upper() == "NOTE" and ws["C9"].value == "USG":
                r_this_2 = usg_rets[0]
                r_this_1 = usg_rets_mid[0]
                r_a = (r_a[0], usg_rets_mid[1], usg_rets[1])
                r_b = (r_b[0], usg_rets_mid[2], usg_rets[2])
                r_c = (r_c[0], usg_rets_mid[3], usg_rets[3])

            # populate template with parameters
            print("Populating report template...")

            report_data_fund = {
                "report_date": report_date,  # done
                "fundname": fund_name,  # done
                "toptableextraspace": "5.5em",
                "series_abbrev": series_abbrev,
                "port_limit": (
                    "Quarterly"
                    if "Q" in df_attributes["series_abbreviation"].iloc[0]
                    else "Monthly"
                ),
                "seriesname": df_attributes["series_name"].iloc[0],
                "fund_description": df_attributes["fund_description"].iloc[0],
                "series_description": df_attributes["series_description"].iloc[0],
                "benchmark": benchmark_name,  # done
                "tgt_outperform": target_outperform_range,  # done
                "exp_rat_footnote": expense_ratio_footnote_text,
                "prev_pd_start": pd.to_datetime(prev_start).strftime(
                    "%B %d, %Y"
                ),  # done
                "this_pd_start": pd.to_datetime(curr_start).strftime(
                    "%B %d, %Y"
                ),  # done
                "prev_pd_return": prev_return,  # done
                "prev_pd_benchmark": benchmark_short,  # done
                "prev_pd_outperform": prev_target_outperform,  # done
                "this_pd_end": pd.to_datetime(curr_end).strftime("%B %d, %Y"),  # done
                "this_pd_est_return": current_target_return,  # done
                "this_pd_est_outperform": target_outperform_net,  # done
                "benchmark_short": benchmark_short,  # done
                "interval1": month_wordify(interval_tuple[0]),  # TODO: review this
                "interval2": month_wordify(interval_tuple[1]),  # TODO: review this
                "descstretch": stretches(
                    df_attributes["fund_description"].iloc[0]
                    + df_attributes["series_description"].iloc[0]
                )[0],
                "pcompstretch": stretches(
                    df_attributes["fund_description"].iloc[0]
                    + df_attributes["series_description"].iloc[0]
                )[1],
                "addl_coll_breakdown": "",
                "oc_aaa": form_as_percent(oc_usg_aaa, 2),  # TODO: review
                "oc_tbills": "-",
                "oc_total": form_as_percent(oc_total, 2),  # TODO: review
                "usg_aaa_cat": (
                    "US Govt Repo" if fund_name == "USG" else "US Govt/AAA Repo"
                ),  # done
                "alloc_aaa": form_as_percent(aloc_usg_aaa, 2),  # TODO: review
                "alloc_tbills": form_as_percent(aloc_tbills, 2),  # TODO: review
                "alloc_total": form_as_percent(1, 1),  # TODO: review
                "tablevstretch": tablevstretch(fund_name),  # done
                # "return_table_plot": "\n\t\\textbf{Lucid USG - Series M}                    & \\textbf{5.55\\%}                              & \\textbf{-}                                  & \\textbf{5.55\\%}                               & \\textbf{-}                           & \\textbf{5.55\\%}                             & \\textbf{-}                          \\\\\n1m T-Bills                       & 5.55\\%                                       & \\textbf{+16 bps}                            & 5.55\\%                               & \\textbf{+17 bps}                     & 5.55\\%                              & \\textbf{+16 bps}                    \\\\\nCrane Govt MM Index                       & 5.55\\%                                       & \\textbf{+43 bps}                           & 5.55\\%                               & \\textbf{+43 bps}                     & 5.55\\%                              & \\textbf{+40 bps}                    \\\\ \\arrayrulecolor{light_grey}\\hline\n\t",
                "return_table_plot": return_table_plot(
                    fund_name=fund_name,  # done
                    prev_pd_return=prev_return,
                    series_abbrev=series_abbrev,
                    r_this_1=r_this_1,
                    r_this_2=r_this_2,
                    comp_a=ws["Y6"].value,
                    comp_b=ws["Z6"].value,
                    comp_c=ws["AA6"].value,
                    r_a=r_a,
                    r_b=r_b,
                    r_c=r_c,
                    s_a_0=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_a[0], 2),
                    ),  # TODO daycounts for spreads and round to 2 places
                    s_a_1=bps_spread(r_this_1, r_a[1]),
                    s_a_2=bps_spread(r_this_2, r_a[2]),
                    s_b_0=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_b[0], 2),
                    ),
                    s_b_1=bps_spread(r_this_1, r_b[1]),
                    s_b_2=bps_spread(r_this_2, r_b[2]),
                    s_c_0=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_c[0], 2),
                    ),
                    s_c_1=bps_spread(r_this_1, r_c[1]),
                    s_c_2=bps_spread(r_this_2, r_c[2]),
                ),
                "fund_size": get_fund_size(reporting_fund, report_date),
                "series_size": get_series_size(reporting_series, report_date),
                "lucid_aum": get_aum(report_date),
                "rating": df_attributes["rating"].iloc[0],  # done
                "rating_org": df_attributes["rating_org"].iloc[0],  # done
                "calc_frequency": "Monthly at par",  # done
                "next_withdrawal_date": curr_withdrawal,  # done
                "next_notice_date": curr_notice,  # done
                "min_invest": "$"
                + str(df_attributes["minimum_investment"].iloc[0]),  # done
                "wal": "28",
                "legal_fundname": df_attributes["legal_fund_name"].iloc[0],  # done
                "fund_inception": df_attributes["fund_inception"]
                .iloc[0]
                .strftime("%B %d, %Y"),  # done
                "series_inception": df_attributes["series_inception"]
                .iloc[0]
                .strftime("%B %d, %Y"),  # done
                "performance_graph": "\n\t\t\t  \\hspace*{-0.9cm}\\resizebox {!} {8cm} {\\begin{tikzpicture}\n\t\t\\begin{axis}[\n\t\t\ttitle style = {font = \\small},\n\t\t\taxis line style = {light_grey},\n\t\ttitle={{Performance vs Benchmark}},\n\t\t\tdate coordinates in=x, date ZERO=2023-02-09,\n\t\t\txticklabel=\\month/\\day/\\year,  \n\t\t\tymin=3, ymax=7, %MAXRETURN HERE\n\t\t\tlegend cell align = {left},\n\t\t\tlegend style={at={(0.3,1)},\n\t\t\t  anchor=north east, font=\\tiny, draw=none,fill=none},\n\t\t\t  x=0.15mm, %CHANGE THIS to tighten in graph, eg if quarterly\n\t\t\tbar width=2.5mm, ybar=2pt, %bar width is width, ybar is space between\n\t\t   % symbolic x coords={Firm 1, Firm 2, Firm 3, Firm 4, Firm 5},\n\t\t\txtick=data,\n\t\t\tx tick label style={rotate=90,anchor=east,font=\\tiny,/pgf/number format/assume math mode},\n\t\t\t\t yticklabel=\\pgfmathparse{\\tick}\\pgfmathprintnumber{\\pgfmathresult}\\,\\%,\n\t\t\ty tick label style = {/pgf/number format/.cd,\n\t\t\t\t\tfixed,\n\t\t\t\t\tfixed zerofill,\n\t\t\t\t\tprecision=2,\n\t\t\t\t\t/pgf/number format/assume math mode\n\t\t\t},\n\t\t\tnodes near coords align={vertical},\n\t\t\tytick distance=0.5,\n\t\t\txtick pos=bottom,ytick pos=left,\n\t\t\tevery node near coord/.append style={font=\\fontsize{6}{6}\\selectfont,/pgf/number format/.cd,\n\t\t\t\t\tfixed,\n\t\t\t\t\tfixed zerofill,\n\t\t\t\t\tprecision=2,/pgf/number format/assume math mode},\n\t\t\t]\n\t\t%\\addplot[ybar, nodes near coords, fill=blue] \n\t\t\\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={shadow yshift=-1pt, shadow xshift=1pt}] \n\t\t\tcoordinates {\n\t\t\t\t(2023-02-09,4.44) (2023-03-09,4.718) (2023-04-13,4.85) (2023-05-11,5.0) (2023-06-15,5.2) (2023-07-20,5.23) (2023-08-17,5.41) (2023-09-14,5.5) (2023-10-19,5.53) (2023-11-16,5.53) (2023-12-14,5.53) (2024-01-18,5.53) (2024-02-15,5.53) (2024-03-14,5.53) (2024-04-18,5.53) (2024-05-16,5.53) \n\t\t\t};\n\t\t\\addplot[draw=dark_red,ultra thick,smooth] \n\t\t\tcoordinates {\n\t\t\t\t(2023-02-09,4.22) (2023-03-09,4.53) (2023-04-13,4.63) (2023-05-11,3.96) (2023-06-15,5.17) (2023-07-20,5.04) (2023-08-17,5.25) (2023-09-14,5.35) (2023-10-19,5.37) (2023-11-16,5.38) (2023-12-14,5.36) (2024-01-18,5.33) (2024-02-15,5.36) (2024-03-14,5.36) (2024-04-18,5.37) (2024-05-16,5.37) \n\t\t\t};\n\t\t\\addplot[draw=dark_color,ultra thick,smooth] \n\t\t\tcoordinates {\n\t\t\t\t(2023-02-09,4.105) (2023-03-09,4.307) (2023-04-13,4.491) (2023-05-11,4.667) (2023-06-15,4.87) (2023-07-20,4.893) (2023-08-17,5.035) (2023-09-14,5.112) (2023-10-19,5.137) (2023-11-16,5.156) (2023-12-14,5.16) (2024-01-18,5.151) (2024-02-15,5.131) (2024-03-14,5.117) (2024-04-18,5.107) (2024-05-16,5.105) \n\t\t\t};\n\t\t\\legend{\\hphantom{A}USG Series M,\\hphantom{A}1m T-Bills,\\hphantom{A}Crane Govt MM Index}\n\t\t\\end{axis}\n\t\t\t\\end{tikzpicture}}\n\n\t\t\t",
            }

            script = ""
            if ws["C5"].value.upper() == "FUND":  # fund (series) report template
                # populate
                if (
                    ws["C9"].value == "USG"
                    or ws["C11"].value == "M"
                    or ws["C11"].value == "C1"
                ):
                    maxreturn = 7
                else:
                    maxreturn = 8

                if (
                    ws["C9"].value == "USG"
                    or ws["C11"].value == "M"
                    or ws["C11"].value == "C1"
                    or ws["C11"].value == "MIG"
                ):
                    minreturn = 3
                else:
                    minreturn = 0

                return_table_plot_backup = return_table_plot(
                    fund_name=ws["C9"].value,
                    prev_pd_return=form_as_percent(
                        ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                    ),
                    series_abbrev=ws["C11"].value,
                    r_this_1=r_this_1,
                    r_this_2=r_this_2,
                    comp_a=ws["Y6"].value,
                    comp_b=ws["Z6"].value,
                    comp_c=ws["AA6"].value,
                    r_a=r_a,
                    r_b=r_b,
                    r_c=r_c,
                    s_a_0=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_a[0], 2),
                    ),  # TODO daycounts for spreads and round to 2 places
                    s_a_1=bps_spread(r_this_1, r_a[1]),
                    s_a_2=bps_spread(r_this_2, r_a[2]),
                    s_b_0=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_b[0], 2),
                    ),
                    s_b_1=bps_spread(r_this_1, r_b[1]),
                    s_b_2=bps_spread(r_this_2, r_b[2]),
                    s_c_0=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_c[0], 2),
                    ),
                    s_c_1=bps_spread(r_this_1, r_c[1]),
                    s_c_2=bps_spread(r_this_2, r_c[2]),
                )

                performance_graph_backup = (
                    performance_graph(
                        True,
                        hspacemap(
                            fund_descriptions[ws["C9"].value.upper()]
                            + series_descriptions[ws["C6"].value.upper()],
                            ts_row_end - ts_row_start + 1,
                        ),
                        "!",
                        str(
                            heightmap(
                                fund_descriptions[ws["C9"].value.upper()]
                                + series_descriptions[ws["C6"].value.upper()]
                            )
                        )
                        + "cm",
                        ws["C9"].value,
                        ws["F" + str(ts_row_start)].value.strftime("%Y-%m-%d"),
                        minreturn,
                        maxreturn,
                        (
                            xmap(
                                fund_descriptions[ws["C9"].value.upper()]
                                + series_descriptions[ws["C6"].value.upper()],
                                ts_row_end - ts_row_start + 1,
                            )
                            if ws["C6"].value.upper() != "PRIMEFUND Q1"
                            else 0.065
                        ),
                        barwidthmap(
                            fund_descriptions[ws["C9"].value.upper()]
                            + series_descriptions[ws["C6"].value.upper()],
                            ts_row_end - ts_row_start + 1,
                        ),
                        plotify(
                            ws,
                            "F",
                            "N" if daycount == 360 else "O",
                            ts_row_start,
                            ts_row_end,
                        ),
                        plotify(ws, "F", "Y", ts_row_start, ts_row_end),
                        plotify(ws, "F", "Z", ts_row_start, ts_row_end),
                        ws["C11"].value,
                        ws["Y6"].value,
                        ws["Z6"].value,
                    )
                    if (ts_row_end - ts_row_start + 1 > 3)
                    else snapshot_graph(
                        -0.8,
                        "!",
                        "6.676cm",
                        0.25
                        + max(
                            round(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value
                                * 100,
                                2,
                            ),
                            round(r_a[0] * 100, 2),
                            round(r_b[0] * 100, 2),
                            round(r_c[0] * 100, 2),
                        )
                        + 0.28,
                        ws["C11"].value,
                        ws["Y6"].value,
                        ws["Z6"].value,
                        ws["AA6"].value,
                        round(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value
                            * 100,
                            2,
                        ),
                        round(r_a[0] * 100, 2),
                        round(r_b[0] * 100, 2),
                        round(r_c[0] * 100, 2),
                    )
                )

                script = fund_report_template.format(
                    report_date=report_data_fund["report_date"],
                    fundname=report_data_fund["fundname"],
                    toptableextraspace=report_data_fund["toptableextraspace"],
                    series_abbrev=report_data_fund["series_abbrev"],
                    port_limit=report_data_fund["port_limit"],
                    seriesname=report_data_fund["seriesname"],
                    fund_description=report_data_fund["fund_description"],
                    series_description=report_data_fund["series_description"],
                    benchmark=report_data_fund["benchmark"],  # TODO ENSURE THERE
                    tgt_outperform=report_data_fund[
                        "tgt_outperform"
                    ],  # TODO ENSURE THERE
                    exp_rat_footnote=report_data_fund["exp_rat_footnote"],
                    prev_pd_start=report_data_fund["prev_pd_start"],
                    this_pd_start=report_data_fund["this_pd_start"],
                    prev_pd_return=report_data_fund["prev_pd_return"],
                    prev_pd_benchmark=report_data_fund["prev_pd_benchmark"],
                    prev_pd_outperform=report_data_fund["prev_pd_outperform"],
                    this_pd_end=report_data_fund["this_pd_end"],  # TODO ENSURE matches
                    this_pd_est_return=report_data_fund["this_pd_est_return"],
                    # TODO ensure there
                    this_pd_est_outperform=report_data_fund[
                        "this_pd_est_outperform"
                    ],  # TODO ENSURE THERE
                    benchmark_short=report_data_fund["benchmark_short"],
                    interval1=report_data_fund["interval1"],
                    interval2=report_data_fund["interval2"],
                    descstretch=report_data_fund["descstretch"],
                    pcompstretch=report_data_fund["pcompstretch"],
                    addl_coll_breakdown=report_data_fund["addl_coll_breakdown"],
                    oc_aaa=report_data_fund["oc_aaa"],
                    oc_tbills=report_data_fund["oc_tbills"],
                    oc_total=report_data_fund["oc_total"],
                    usg_aaa_cat=report_data_fund["usg_aaa_cat"],
                    alloc_aaa=report_data_fund["alloc_aaa"],
                    alloc_tbills=report_data_fund["alloc_tbills"],
                    alloc_total=report_data_fund["alloc_total"],
                    tablevstretch=report_data_fund[
                        "tablevstretch"
                    ],  # only for fund report
                    return_table_plot=report_data_fund["return_table_plot"],
                    fund_size=report_data_fund["fund_size"],
                    series_size=report_data_fund[
                        "series_size"
                    ],  # post sub/redemp, TODO temporarily consold hardwired
                    lucid_aum=report_data_fund["lucid_aum"],
                    rating=report_data_fund["rating"],
                    rating_org=report_data_fund["rating_org"],
                    calc_frequency=report_data_fund["calc_frequency"],
                    next_withdrawal_date=report_data_fund["next_withdrawal_date"],
                    next_notice_date=report_data_fund["next_notice_date"],
                    min_invest=report_data_fund["min_invest"],
                    wal=report_data_fund["wal"],
                    legal_fundname=report_data_fund["legal_fundname"],
                    fund_inception=report_data_fund["fund_inception"],
                    series_inception=report_data_fund["series_inception"],
                    performance_graph=report_data_fund["performance_graph"],
                )

            elif ws["C5"].value.upper() == "NOTE":  # note report template

                report_data_note = {
                    "report_date": report_date_formal,
                    "fundname": "Test Fund Name",
                    "series_abbrev": "M",
                    "port_limit": "Monthly",
                    "issuer_name": "USG LLC",
                    "frequency": "Monthly",
                    "rating": "AAA",
                    "rating_org": "Edgar Jones LLC",
                    "benchmark": "1m T-Bills",
                    "tgt_outperform": "15-20",
                    "prev_pd_start": "April 15, 1990",
                    "this_pd_start": "May 15, 1990",
                    "prev_pd_return": "5.55\\%",
                    "prev_pd_benchmark": "1m TB",
                    "prev_pd_outperform": "99 bps",
                    "this_pd_end": "July 13",
                    "this_pd_est_return": "5.55\\%",
                    "this_pd_est_outperform": "99",
                    "benchmark_short": "1m TBoo",
                    "interval1": "12 Months",
                    "interval2": "12 Years",
                    "return_table_plot": "\n\t\\textbf{Lucid USG - Series M}                    & \\textbf{5.53\\%}                              & \\textbf{-}                                  & \\textbf{5.56\\%}                               & \\textbf{-}                           & \\textbf{5.60\\%}                             & \\textbf{-}                          \\\\\n1m T-Bills                       & 5.37\\%                                       & \\textbf{+16 bps}                            & 5.39\\%                               & \\textbf{+17 bps}                     & 5.44\\%                              & \\textbf{+16 bps}                    \\\\\nCrane Govt MM Index                       & 5.10\\%                                       & \\textbf{+43 bps}                           & 5.13\\%                               & \\textbf{+43 bps}                     & 5.20\\%                              & \\textbf{+40 bps}                    \\\\ \\arrayrulecolor{light_grey}\\hline\n\t",
                    "colltable": "\n\t\t\\renewcommand{\\arraystretch}{1.91}\\begin{tabular}{!{\\color{light_grey}\\vrule}\n\t\t>{\\columncolor[HTML]{EFEFEF}}p{3.5cm} \n\t\t>{\\columncolor[HTML]{EFEFEF}}c\n\t\t>{\\columncolor[HTML]{EFEFEF}}c!{\\color{light_grey}\\vrule}}\n\t\t\\arrayrulecolor{light_grey}\\hline\n\t\t\\multicolumn{3}{!{\\color{light_grey}\\vrule}l!{\\color{light_grey}\\vrule}}{\\rowcolor{lucid_blue}{\\color[HTML]{FFFFFF}\\textbf{Series Collateral Overview\\textsuperscript{4}}}} \\\\\n\t\t\\multicolumn{3}{!{\\color{light_grey}\\vrule}p{8.2cm}!{\\color{light_grey}\\vrule}}{\\rowcolor[HTML]{EFEFEF}{\\textbf{Series M}: Secured by \\textbf{US Government backed (USG) securities only}, with daily valuations \\& margining.}} \\\\\n\t\t& & \\\\\n\t\t & \\textbf{\\% Portfolio} & \\textbf{O/C Rate}\\\\\n\t\tUS Govt Repo & 98.8\\% & 107.0\\% \\\\\n\t\tT-Bills; Gov't MMF & 1.2\\% & - \\\\ \\cline{2-2} \\cline{3-3} \n\t\t\\textbf{Total} & 100.0\\% & \\textbf{107.0\\%} \\\\\\arrayrulecolor{light_grey}\\hline\n\t\t\\end{tabular}\n\t\t",
                    "zero_date": "1990-01-01",
                    "max_return": 3,
                    "return_plot": "(2024-01-18,5.53) (2024-02-15,5.53) (2024-03-14,5.53) (2024-04-18,5.53) (2024-05-16,5.53) ",
                    "comp_a_plot": "(2024-01-18,5.33) (2024-02-15,5.36) (2024-03-14,5.36) (2024-04-18,5.37) (2024-05-16,5.37) ",
                    "comp_b_plot": "(2024-01-18,5.151) (2024-02-15,5.131) (2024-03-14,5.117) (2024-04-18,5.107) (2024-05-16,5.105) ",
                    "performance_graph": "\n\t\t  \\hspace*{-0.86cm}\\resizebox {!} {6.676cm} {\\begin{tikzpicture}\n\t\\begin{axis}[\n\t\ttitle style = {font = \\small},\n\t\taxis line style = {light_grey},\n\t\t\ttitle={{Performance vs Benchmarks}},\n\t\tymin=2, ymax=6.53, %MAXRETURN HERE\n\t   symbolic x coords={Series M-8,1m T-Bills,Crane Govt},\n\t\txtick={Series M-8,1m T-Bills,Crane Govt},\n\t\tx tick label style={anchor=north,font=\\scriptsize,/pgf/number format/assume math mode},\n\t\tyticklabel=\\pgfmathparse{\\tick}\\pgfmathprintnumber{\\pgfmathresult}\\,\\%,\n\t\ty tick label style = {/pgf/number format/.cd,\n\t\t\t\tfixed,\n\t\t\t\tfixed zerofill,\n\t\t\t\tprecision=2,\n\t\t\t\t/pgf/number format/assume math mode\n\t\t},\n\t\tytick distance=0.5,\n\t\tbar width = 10mm, x = 3.7cm,\n\t\txtick pos=bottom,ytick pos=left,\n\t\tevery node near coord/.append style={font=\\fontsize{8}{8}\\selectfont,/pgf/number format/.cd,\n\t\t\t\tfixed,\n\t\t\t\tfixed zerofill,\n\t\t\t\tprecision=2,/pgf/number format/assume math mode},\n\t\t]\n\t\\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={shadow yshift=-1pt, shadow xshift=1pt}] \n\t\tcoordinates {\n\t\t\t(Series M-8,5.53) \n\t\t};\n\t\\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={shadow yshift=-1pt, shadow xshift=1pt}] \n\t\tcoordinates {\n\t\t\t(1m T-Bills,5.37) \n\t\t};\n\t\\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={shadow yshift=-1pt, shadow xshift=1pt}] \n\t\tcoordinates {\n\t\t\t(Crane Govt,5.1) \n\t\t};\n\t\\end{axis}\n\t\t\\end{tikzpicture}}\n\t\n\t\t",
                    "fund_size": "\\$123.0 million",
                    "series_size": "\\$123.0 million",
                    "lucid_aum": "\\$4.65 billion",
                    "fund_inception": "June 29 1990",
                    "cusip": "90366JAG2",
                    "note_abbrev": "M-8",
                    "principal_outstanding": "\\$20.7 million",
                    "issue_date": "December 14, 1990",
                    "maturity_date": "December 14, 2023",
                    "pd_end_date_long": "December 31, 2099",
                    "next_notice_date": "June 10, 9999",
                    "coupon_plot": "12/14/23 &\\textbf{{01/18/24}} &\\textbf{{5.53\\%}} &1m TB+20 &\\$20,700,000 &\\$109,766.71 &01/18/24 &\\$20,700,000 &108.2\\% \\\\01/18/24 &\\textbf{{02/15/24}} &\\textbf{{5.53\\%}} &1m TB+17 &\\hphantom{{\\$}}20,700,000 &\\hphantom{{\\$}}87,813.37 &02/15/24 &\\hphantom{{\\$}}20,700,000 &105.8\\% \\\\02/15/24 &\\textbf{{03/14/24}} &\\textbf{{5.53\\%}} &1m TB+17 &\\hphantom{{\\$}}20,700,000 &\\hphantom{{\\$}}87,813.37 &03/14/24 &\\hphantom{{\\$}}20,700,000 &107.0\\% \\\\03/14/24 &\\textbf{{04/18/24}} &\\textbf{{5.53\\%}} &1m TB+16 &\\hphantom{{\\$}}20,700,000 &\\hphantom{{\\$}}109,766.71 &04/18/24 &\\hphantom{{\\$}}20,700,000 &106.5\\% \\\\04/18/24 &\\textbf{{05/16/24}} &\\textbf{{5.53\\%}} &1m TB+16 &\\hphantom{{\\$}}20,700,000 &\\hphantom{{\\$}}87,813.37 &05/16/24 &\\hphantom{{\\$}}20,700,000 &107.0\\% \\\\05/16/24 &\\textbf{{06/13/24}} &\\textbf{{5.52\\%{{\\tiny (Est'd)}}}} &1m TB+18 &\\hphantom{{\\$}}20,700,000 &\\hphantom{{\\$}}n/a &06/13/24 &\\hphantom{{\\$}}20,700,000 &n/a \\\\\n\t\t& & & & & & & &      \\\\ \n\n\t\t\n\t\t& & & & & & & &      \\\\ \n\n\t\t\n\t\t& & & & & & & &      \\\\ \n\n\t\t\n\t\t& & & & & & & &      \\\\ \n\n\t\t\n\t\t& & & & & & & &      \\\\ \n\n\t\t\n\t\t\n\n\t\t",
                    "rets_disclaimer_if_m1": "",
                }

                script = note_report_template.format(
                    report_date=report_data_note["report_date"],
                    fundname=report_data_note["fundname"],
                    series_abbrev=report_data_note["series_abbrev"],
                    issuer_name=report_data_note["issuer_name"],
                    frequency=report_data_note["frequency"],
                    rating=report_data_note["rating"],
                    rating_org=report_data_note["rating_org"],
                    benchmark=report_data_note["benchmark"],
                    tgt_outperform=report_data_note["tgt_outperform"],
                    prev_pd_start=report_data_note["prev_pd_start"],
                    prev_pd_return=report_data_note["prev_pd_return"],
                    prev_pd_benchmark=report_data_note["prev_pd_benchmark"],
                    prev_pd_outperform=report_data_note["prev_pd_outperform"],
                    this_pd_end=report_data_note["this_pd_end"],
                    this_pd_est_return=report_data_note["this_pd_est_return"],
                    this_pd_est_outperform=report_data_note["this_pd_est_outperform"],
                    this_pd_start=report_data_note["this_pd_start"],
                    benchmark_short=report_data_note["benchmark_short"],
                    interval1=report_data_note["interval1"],
                    interval2=report_data_note["interval2"],
                    return_table_plot=report_data_note["return_table_plot"],
                    colltable=report_data_note["colltable"],
                    zero_date=report_data_note["zero_date"],
                    max_return=report_data_note["max_return"],
                    return_plot=report_data_note["return_plot"],
                    comp_a_plot=report_data_note["comp_a_plot"],
                    comp_b_plot=report_data_note["comp_b_plot"],
                    performance_graph=report_data_note["performance_graph"],
                    fund_size=report_data_note["fund_size"],
                    series_size=report_data_note["series_size"],
                    lucid_aum=report_data_note["lucid_aum"],
                    fund_inception=report_data_note["fund_inception"],
                    cusip=report_data_note["cusip"],
                    note_abbrev=report_data_note["note_abbrev"],
                    principal_outstanding=report_data_note["principal_outstanding"],
                    issue_date=report_data_note["issue_date"],
                    maturity_date=report_data_note["maturity_date"],
                    pd_end_date_long=report_data_note["pd_end_date_long"],
                    next_notice_date=report_data_note["next_notice_date"],
                    coupon_plot=report_data_note["coupon_plot"],
                    rets_disclaimer_if_m1=report_data_note["rets_disclaimer_if_m1"],
                )
            else:
                continue
            # write script to file
            print("Generating Latex file...")
            filepath = report_name.replace(" ", "_")
            script_file = filepath + ".tex"
            with open(script_file, "w") as out:
                out.write(script)
                out.close()
            # generate pdf
            print("Generating PDF...")
            pdf_file = filepath + ".pdf"
            cmd_str = "pdflatex -interaction nonstopmode {} {}".format(
                script_file, pdf_file
            )
            try:
                try:
                    x = subprocess.check_output(cmd_str)
                except:
                    print("File generated to {}.".format(pdf_file))
                    reports_generated.append(report_name)
            except:
                print("Error generating file {}.".format(pdf_file))
    except Exception as e:
        print("EXCEPTION")
        print(e)
        print("Error generating " + report_name)
        bad_reports.append(report_name)
        continue

# close workbook
wb.close()

if not bad_reports:
    print("All reports generated:")
else:
    print("Some reports generated:")

for x in reports_generated:
    print(x)
# TODO check page counts
if bad_reports:
    print("****ERROR generating following reports:")
    for x in bad_reports:
        print(x)
