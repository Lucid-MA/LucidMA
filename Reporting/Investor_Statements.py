# import sys
import os
import subprocess
import openpyxl as op
from pathlib import Path, PureWindowsPath
from datetime import datetime, timedelta

# intialize script templates
fund_report_template = r"""
\documentclass[9pt]{{article}}
\usepackage[T1]{{fontenc}}
\hfuzz=27pt 
\usepackage{{lmodern}}
\usepackage[dvipsnames,table,xcdraw]{{xcolor}}
\definecolor{{lucid_blue}}{{RGB}}{{0,18,82}}
\definecolor{{light_grey}}{{RGB}}{{198,198,198}}
\definecolor{{dark_red}}{{RGB}}{{144,8,8}}
\definecolor{{dark_color}}{{RGB}}{{9,143,68}}
\definecolor{{dark_grey}}{{RGB}}{{194,194,214}}
\usepackage[margin=0.7in,headsep=0.1in]{{geometry}}
\usepackage{{fancyhdr}}
\usepackage{{graphicx}}
\usepackage{{listings}}
\usepackage{{colortbl}}
\usepackage{{boldline}}
\graphicspath{{ {{images/}} }}
\pagestyle{{fancy}}
\usepackage{{adjustbox}}
\usepackage{{pgfplots}}
\renewcommand{{\familydefault}}{{\sfdefault}} % make font nice non bitmap
\renewcommand*{{\thepage}}{{\small\arabic{{page}}}}
\usepgfplotslibrary{{dateplot}}
\def\mywidth{{17.6cm}}
\def\halfwidthb{{12.4cm}}
\def\quarterwidthb{{6cm}}
\def\quarterwidth{{4.35cm}}
\def\eighthwidth{{2.175cm}}
\def\quarterwidtha{{3.41cm}}
\def\eighthwidtha{{1.2cm}}
\def\lesswidth{{10.9cm}}
\usetikzlibrary{{shadows,shadows.blur,shapes.geometric}}
\pgfplotsset{{compat=1.8}}
\pgfplotsset{{every axis/.append style={{
                    label style={{font=\tiny}},
                    tick label style={{font=\tiny}}  
                    }}}}
\lhead{{\includegraphics[width=9cm]{{lucid_logo.png}}}}
\rhead{{Report Date: {report_date}}}
\setlength{{\columnsep}}{{3em}}
\begin{{document}}
\twocolumn[{{
\begin{{center}}
\renewcommand{{\arraystretch}}{{1.5}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor{{lucid_blue}}}}m{{\mywidth}} !{{\color{{light_grey}}\vrule}}}}
{{\color[HTML]{{FFFFFF}} \large \textbf{{Monthly Report}}}}\\
\end{{tabular}}
\end{{center}}

\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor{{lucid_blue}}}}l !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{\quarterwidthb}} !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}l !{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
{{\color[HTML]{{FFFFFF}} \textbf{{Program Series}}}} & \multicolumn{{2}}{{p{{\halfwidthb}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{Lucid {fundname} - Series {series_abbrev}}}}} \\[1.5mm]
{{\color[HTML]{{FFFFFF}} \textbf{{Objective and Strategy}}}} & \multicolumn{{2}}{{p{{\halfwidthb}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}} {fund_description} 

{series_description}}} \\[{toptableextraspace}]
{{\color[HTML]{{FFFFFF}} \textbf{{Current Target Return}}\textsuperscript{{1}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{{benchmark} + {tgt_outperform} bps}}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Previous Period Return}}}} & {prev_pd_start} - {this_pd_start} & \textbf{{{prev_pd_return}}} ({prev_pd_benchmark} + {prev_pd_outperform}) \\ 
{{\color[HTML]{{FFFFFF}} \textbf{{Current Period Est'd Return}}\textsuperscript{{2}}}} & {this_pd_start} - {this_pd_end} & \textbf{{{this_pd_est_return}}}  ({benchmark_short} + {this_pd_est_outperform} bps) \\ \arrayrulecolor{{light_grey}}\hline
\end{{tabular}}

\begin{{center}}
\noindent \renewcommand{{\arraystretch}}{{{tablevstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}m{{5.2cm}}!{{\color{{light_grey}}\vrule}}m{{\eighthwidth}}c!{{\color{{light_grey}}\vrule}}m{{\eighthwidtha}}c!{{\color{{light_grey}}\vrule}}m{{\eighthwidtha}}c!{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
\rowcolor{{lucid_blue}} 
{{\color[HTML]{{FFFFFF}} \textbf{{Net Returns}}\textsuperscript{{3}}}}               & \multicolumn{{2}}{{m{{\quarterwidth}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}\centering {{\color[HTML]{{FFFFFF}} \textbf{{Previous Period}}}}}} & \multicolumn{{2}}{{m{{\quarterwidtha}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}\centering {{\color[HTML]{{FFFFFF}} \textbf{{{interval1}}}}}}} & \multicolumn{{2}}{{m{{\quarterwidtha}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}\centering {{\color[HTML]{{FFFFFF}} \textbf{{{interval2}}}}}}} \\
\rowcolor{{lucid_blue}}  
{{\color[HTML]{{FFFFFF}} Series / Comparables}} & {{\color[HTML]{{FFFFFF}} Return}}                & {{\color[HTML]{{FFFFFF}} Spread}}               & {{\color[HTML]{{FFFFFF}} Return\textsuperscript{{1}}}}        & {{\color[HTML]{{FFFFFF}} Spread}}        & {{\color[HTML]{{FFFFFF}} Return\textsuperscript{{1}}}}       & {{\color[HTML]{{FFFFFF}} Spread}}       \\ \arrayrulecolor{{light_grey}}\hline
{return_table_plot}
\end{{tabular}}
\end{{center}}
}}]
\hfill \break
\hfill \break

\noindent \renewcommand{{\arraystretch}}{{{descstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{3.7cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{4cm}}!{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
\multicolumn{{2}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{ \color[HTML]{{FFFFFF}}\textbf{{Fund and Series Details\textsuperscript{{4}}}}}}}} \\
Fund Size & {fund_size}\\
Series Size & {series_size}\\
Lucid AUM & {lucid_aum}\\
Series Rating & {rating} by {rating_org}\\
Series Withdrawal & {calc_frequency}\\
Next Withdrawal & {next_withdrawal_date}\\
Next Notice Date & {next_notice_date}\\
Min Investment & {min_invest}\\
Current WAL & {wal} days\\
\noindent\parbox[b]{{\hsize}}{{\vspace{{1mm}}Current Max Limit on all Series Assets}} & {next_withdrawal_date}\\[-1mm]
Fund Entity & {legal_fundname}\\
Fund Inception & {fund_inception}\\
Series Inception & {series_inception}\\ \arrayrulecolor{{light_grey}}\hline
\end{{tabular}}
\hspace*{{-0.2cm}}\begin{{tabular}}{{p{{8.45cm}}}}
\textit{{\scriptsize Please see fund Offering Memorandum and related documents for complete terms and Important Disclaimer attached.}}
\end{{tabular}}


\begin{{figure}}
\centering
\noindent\renewcommand{{\arraystretch}}{{{pcompstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{3.5cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}c
>{{\columncolor[HTML]{{EFEFEF}}}}c!{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
\multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{ \color[HTML]{{FFFFFF}}\textbf{{Portfolio Composition\textsuperscript{{5}}}}}}}} \\
\textbf{{Series Assets}} & \textbf{{\% Portfolio}} & \textbf{{O/C Rate}}\\
{usg_aaa_cat} & {alloc_aaa} & {oc_aaa} \\
{addl_coll_breakdown}
T-Bills; Gov't MMF & {alloc_tbills} & {oc_tbills} \\ \cline{{2-2}} \cline{{3-3}} 
\textbf{{Total}} & {alloc_total} & \textbf{{{oc_total}}} \\\arrayrulecolor{{light_grey}}\hline
\end{{tabular}}

% \tikzfading[name=fade down,
%     top color=transparent!30,
%     bottom color=transparent!0]

{performance_graph}
\end{{figure}}


\onecolumn



\pagebreak 

\footnotesize
\noindent\textbf{{\color{{lucid_blue}}Notes}}

\begin{{enumerate}}
\item Target returns based on the program manager's estimate of the projected returns for the respective series based on current market conditions. 

\item Current return (estimated) is based on the rates of the invested series portfolio as of the current period start date.  Actual period return based on the final net returns of portfolio.   

\item Annualized net returns of Fund Series and comparables are for the entirety of each period and are quoted on an Act/360 basis for Lucid Prime Series and Act/365 for Lucid USG series. Any interperiod subscriptions will have different returns based upon the respective interperiod portfolio investments and allocations. Net returns include the applicable series expense ratio and include any management fee waivers or maximum expense caps. Historical returns assume reinvestment at the applicable Fund Series, Libor, T-Bill or MMF Index rate at the end of each period.  Money Market index returns based on the average of the daily rates for the respective period. SOFR is the term reference rate for the applicable period (e.g. 1m or 3m) as published by the CME Group. {exp_rat_footnote}

\item All Fund details as of the last period end date. Fund accepts new subscriptions and redemptions on each Withdrawal Date.  Manager may accept subscriptions on any other day with approval, as fully described in the private offering memorandum.

\item Portfolio composition and Over-Collateralization Rate (``O/C Rate'') of the repo investments as of the business day prior to the last day of the most recent period. O/C Rate equals the market value of the collateral as a proportion of the respective repo investments. Eligible repo collateral details and classifications for the Series as fully described in the private offering memorandum.

\end{{enumerate}}

\noindent\textbf{{\color{{lucid_blue}}Important Disclaimer}}

\scriptsize

\noindent This material has been prepared by Lucid Management and Capital Partners LP or one of its affiliates, principals or advisors (``Lucid'') and may contain ``forward-looking statements'' which are based on Lucid's beliefs, as well as on a number of assumptions concerning future events, based on information currently available to Lucid. Readers are cautioned not to put undue reliance on such forward-looking statements, which are not a guarantee of future performance, and are subject to a number of uncertainties and other factors, many of which are outside Lucid's control, which could cause actual results to differ materially from such statements. This material is for distribution only under such circumstances as may be permitted by applicable law and it is solely intended for qualified institutions and individuals with existing relationships with Lucid, its affiliates or advisers.  It has no regard to the specific investment objectives, financial situation or particular needs of any recipient. It is published solely for informational and discussion purposes only and is not to be construed as a solicitation or an offer to buy or sell any securities, related financial instruments, actual fund or specific transaction. No representation or warranty, either express or implied, is provided in relation to the accuracy, completeness or reliability of the information contained herein, nor is it intended to be a complete statement or summary of the securities, markets or developments referred to in the materials.  It should not be regarded by recipients as a substitute for the exercise of their own judgment. Any opinions expressed in this material are subject to change without notice and may differ or be contrary to opinions expressed by other business areas or groups of Lucid as a result of using different assumptions and criteria. Lucid is under no obligation to update or keep current the information contained herein. Lucid, its partners, officers and employees' or clients may have or have had interests or long or short positions in the securities or other financial instruments referred to herein and may at any time make purchases and/or sales in them as principal or agent. Neither Lucid nor any of its affiliates, nor any of Lucid or any of its affiliates, partners, employees or agents accepts any liability for any loss or damage arising out of the use of all or any part of this material. Money market investments including repurchase agreements are not suitable for all investors. Past performance is not necessarily indicative of future results.  Prior to entering into a transaction you should consult with your own legal, regulatory, tax, financial and accounting advisers to the extent you deem necessary to make your own investment, hedging and trading decisions. Any transaction between you and Lucid will be subject to the detailed provisions of a private placement memorandum or a managed account agreement relating to that transaction. Additional information will be made available upon request. The indicative information in this document (the ``Information'') are provided to you for information purposes only and may not be complete. Any redistribution of this document without express written consent of Lucid is prohibited.  No part of this material may be reproduced in any form, or referred to in any publication, without express written consent of Lucid.  This presentation should only be considered current as of the date of the publication without regard to the date on which you may have accessed or received the information. Unless required by law or specifically agreed in writing, we have no obligation to continue to provide to you the Information, and we may cease doing so at any time in our sole discretion. \underline{{\textit{{Targeted Return Disclosure}}}} \textit{{The targeted returns included in this presentation are not intended as, and must not be regarded as, a representation, warranty or prediction that any Fund (or series thereof) will achieve any particular rate of return over any particular time period or that any Fund (or series thereof) will not incur losses. Although Lucid believes, based on these factors, that the referenced return targets are reasonable, return targets are subject to inherent limitations including, without limitation, the fact they cannot take into account the impact of future economic events on future trading and investment decisions. These events may include changes in interest rates and/or benchmarks greater than those occurring within the historical time period examined when developing the return targets, or future changes in laws or regulations. All targeted returns are net of Lucid's anticipated fees and expenses.}}

\vspace{{1mm}}
\noindent SEC ADV Part 2 firm brochure: \underline{{https://files.adviserinfo.sec.gov/IAPD/Content/Common/crd\_iapd\_Brochure.aspx?BRCHR\_VRSN\_ID=832231}}}}


\scriptsize
\begin{{center}}
\renewcommand{{\arraystretch}}{{1.2}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{8cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}l !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{9cm}}  !{{\color{{light_grey}}\vrule}}}}
\multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{ \color[HTML]{{FFFFFF}}\textbf{{Contact Information}}}}}} \\ \arrayrulecolor{{light_grey}}\hline
\textbf{{Investment Manager}} &  & \textbf{{For Investors (Subscriptions \& Withdrawals)}} \\\arrayrulecolor{{light_grey}}\hline
Lucid Management and Capital Partners LP &  & \underline{{Lucid.IR@sscinc.com}} with copy to: \\
295 Madison Avenue, 39th Floor &  & \underline{{operations@lucidma.com}} \\
New York, New York 10017 &  & \\
T: +1-212-551-1702 &  &  \\
Investor Relations: \underline{{carolina.siles@lucidma.com}} &  & \\
\textbf{{Fund Auditor:}} KPMG & & \textbf{{Fund Custodian:}} Bank of NY Mellon\\ \arrayrulecolor{{light_grey}}\hline
\end{{tabular}}
\end{{center}}
\end{{document}}
"""

note_report_template = r"""

\documentclass[9pt]{{article}}
\usepackage[T1]{{fontenc}}
\hfuzz=27pt 
\usepackage{{lmodern}}
\usepackage[dvipsnames,table,xcdraw]{{xcolor}}
\definecolor{{lucid_blue}}{{RGB}}{{0,18,82}}
\definecolor{{light_grey}}{{RGB}}{{198,198,198}}
\definecolor{{dark_red}}{{RGB}}{{144,8,8}}
\definecolor{{dark_color}}{{RGB}}{{9,143,68}}
\definecolor{{dark_grey}}{{RGB}}{{194,194,214}}
\definecolor{{darker_grey}}{{RGB}}{{143,143,143}}
\usepackage[margin=0.7in,headsep=0.1in]{{geometry}}
\usepackage{{fancyhdr}}
\usepackage{{graphicx}}
\usepackage{{subfig}}
\usepackage{{colortbl}}
\usepackage{{listings}}
\usepackage{{boldline}}
\usepackage{{ragged2e}}
\graphicspath{{ {{images/}} }}
\pagestyle{{fancy}}
\usepackage{{adjustbox}}
\usepackage{{pgfplots}}
\renewcommand{{\familydefault}}{{\sfdefault}} % make font nice non bitmap
\renewcommand*{{\thepage}}{{\small\arabic{{page}}}}
\usepgfplotslibrary{{dateplot}}
\def\mywidth{{17.6cm}}
\def\halfwidthb{{11.2cm}}
\def\quarterwidthb{{4cm}}
\def\quarterwidth{{4.35cm}}
\def\eighthwidth{{2.175cm}}
\def\quarterwidtha{{3.41cm}}
\def\eighthwidtha{{1.2cm}}
\usetikzlibrary{{shadows,shadows.blur,shapes.geometric}}
\pgfplotsset{{compat=1.8}}
\pgfplotsset{{every axis/.append style={{
                    label style={{font=\tiny}},
                    tick label style={{font=\tiny}}  
                    }}}}
\lhead{{\includegraphics[width=9cm]{{lucid_logo.png}}}}
\rhead{{Report Date: {report_date}}}
\setlength{{\columnsep}}{{3em}}
\begin{{document}}
\

\noindent\renewcommand{{\arraystretch}}{{1.5}}\begin{{tabular}}{{
>{{\columncolor[HTML]{{8F8F8F}}}}p{{8.3cm}} 
>{{\columncolor[HTML]{{8F8F8F}}}}p{{5.5cm}} 
>{{\columncolor[HTML]{{8F8F8F}}}}r}}
\textbf{{\large Lucid Medium Term Notes}} &  & \textbf{{Noteholder Report}} \\
\end{{tabular}}

\begin{{center}}
\renewcommand{{\arraystretch}}{{1.3}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor{{lucid_blue}}}}l !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{\quarterwidthb}} !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}l !{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
%\cline{{2-3}}


{{\color[HTML]{{FFFFFF}} \textbf{{Program Series}}}} & \multicolumn{{2}}{{p{{11.78cm}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}} 
\textbf{{Lucid {fundname} Series {series_abbrev} / Note Series {note_abbrev}}}
}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Issuer}}}} & \multicolumn{{2}}{{p{{11.78cm}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}} 
{issuer_name}
}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Redemption \& Coupon Frequency}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{{frequency}}}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Rating}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{{rating}}} by {rating_org}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Current Target Return of Notes}}\textsuperscript{{1}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{{benchmark} + {tgt_outperform} bps}}}}\\
{{\color[HTML]{{FFFFFF}} \textbf{{Previous Coupon Period}}}} & {prev_pd_start} - {this_pd_start} & \textbf{{{prev_pd_return}}} ({prev_pd_benchmark} + {prev_pd_outperform}) \\
{{\color[HTML]{{FFFFFF}} \textbf{{Current Period Est'd Coupon}}\textsuperscript{{2}}}} & {this_pd_start} - {this_pd_end} & \textbf{{{this_pd_est_return}}}  ({benchmark_short} + {this_pd_est_outperform} bps) \\ \arrayrulecolor{{light_grey}}\hline
\end{{tabular}}
\end{{center}}

\noindent{{\color{{lucid_blue}}\textbf{{Historical Performance of Program Series vs Benchmarks}} \textit{{(all-in net returns)}}\textsuperscript{{3}}}}

\noindent\renewcommand{{\arraystretch}}{{1.3}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}m{{5.2cm}}!{{\color{{light_grey}}\vrule}}m{{\eighthwidth}}c!{{\color{{light_grey}}\vrule}}m{{\eighthwidtha}}c!{{\color{{light_grey}}\vrule}}m{{\eighthwidtha}}c!{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
\rowcolor{{lucid_blue}} 
{{\color[HTML]{{FFFFFF}} \textbf{{ Series / Comparables}}}}               & \multicolumn{{2}}{{m{{\quarterwidth}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}\centering {{\color[HTML]{{FFFFFF}} \textbf{{Previous Period}}}}}} & \multicolumn{{2}}{{m{{\quarterwidtha}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}\centering {{\color[HTML]{{FFFFFF}} \textbf{{{interval1}}}}}}} & \multicolumn{{2}}{{m{{\quarterwidtha}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}\centering {{\color[HTML]{{FFFFFF}} \textbf{{{interval2}}}}}}} \\ 
\rowcolor{{lucid_blue}}  
{{\color[HTML]{{FFFFFF}}}} & {{\color[HTML]{{FFFFFF}} Return\textsuperscript{{1}}}}                & {{\color[HTML]{{FFFFFF}} Spread}}               & {{\color[HTML]{{FFFFFF}} Return\textsuperscript{{1}}}}        & {{\color[HTML]{{FFFFFF}} Spread}}        & {{\color[HTML]{{FFFFFF}} Return\textsuperscript{{1}}}}       & {{\color[HTML]{{FFFFFF}} Spread}}       \\ \arrayrulecolor{{light_grey}}\hline
{return_table_plot}
\end{{tabular}}

\vspace{{.8cm}}

\noindent\adjustbox{{valign=t}}{{\begin{{minipage}}{{9.5cm}}
\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor{{lucid_blue}}}}m{{8.2cm}} !{{\color{{light_grey}}\vrule}}}}
{{\color[HTML]{{FFFFFF}} \small \textbf{{Most Recent Period Returns vs Benchmarks\textsuperscript{{3}}}}}}\\
\end{{tabular}}
{performance_graph}
\end{{minipage}}}}%
\hfill
\adjustbox{{valign=t}}{{\begin{{minipage}}[t]{{15cm}}
    {colltable}
\end{{minipage}}}}

\begin{{center}}
\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{5.05cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{3.05cm}} !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{4.05cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{4.35cm}} !{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
\multicolumn{{4}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor{{lucid_blue}}{{\color[HTML]{{FFFFFF}}\textbf{{Program Overview\textsuperscript{{5}}}}}}}} \\ \arrayrulecolor{{light_grey}}\hline
\textbf{{Related Fund Series Size}} & \textbf{{{series_size}}} & Issuing \& Paying Agent & Bank of NY Mellon \\ 
\textbf{{Lucid {fundname} Program Size}} & \textbf{{{fund_size}}} & Collateral Agent & Bank of NY Mellon \\ 
\textbf{{Lucid Platform AUM}} & \textbf{{{lucid_aum}}} & Auditor & KPMG \\ 
{fundname} Program Inception & {fund_inception} & & \\ \arrayrulecolor{{light_grey}}\hline
\end{{tabular}}
\end{{center}}

\begin{{center}}
    \noindent{{\small Please see page 2 for Coupons by CUSIP, program documents for complete terms, and Important Disclaimer attached.}}
\end{{center}}


\pagebreak 

\noindent\renewcommand{{\arraystretch}}{{1.5}}\begin{{tabular}}{{
>{{\columncolor[HTML]{{8F8F8F}}}}m{{\mywidth}}}}
{{\large \textbf{{Coupons by CUSIP}}}}\\
\end{{tabular}}


\begin{{center}}
\noindent\renewcommand{{\arraystretch}}{{1.3}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
>{{\columncolor{{lucid_blue}}}}l !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1cm}} !{{\color{{light_grey}}\vrule}}
>{{\columncolor[HTML]{{EFEFEF}}}}l !{{\color{{light_grey}}\vrule}}}}
\arrayrulecolor{{light_grey}}\hline
%\cline{{2-3}}
{{\color[HTML]{{FFFFFF}} \textbf{{CUSIP}}}} & \multicolumn{{2}}{{p{{11.56cm}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}} 
\textbf{{{cusip}}}
}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Note Series}}}} & \multicolumn{{2}}{{p{{11.56cm}}!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}} 
\textbf{{{note_abbrev}}} (Secured by related fund series {series_abbrev})
}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Current Principal}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{{principal_outstanding}}}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Original Issue Date}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}{issue_date}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Final Maturity Date}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}{maturity_date}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Next Coupon Payment Date}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}{pd_end_date_long}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Next Redemption Date (Put Date)}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}\textbf{{{pd_end_date_long}}}}} \\
{{\color[HTML]{{FFFFFF}} \textbf{{Next Notice Date for Redemption}}}} & \multicolumn{{2}}{{l!{{\color{{light_grey}}\vrule}}}}{{\cellcolor[HTML]{{EFEFEF}}{next_notice_date}}} \\ \arrayrulecolor{{light_grey}}\hline
\end{{tabular}}
\end{{center}}

\begin{{center}}{{\footnotesize
\noindent\begin{{tabular}}{{
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1.45cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1.45cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1.70cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1.6cm}} 
>{{\columncolor[HTML]{{EFEFEF}}\RaggedLeft\arraybackslash}}p{{2cm}} 
>{{\columncolor[HTML]{{EFEFEF}}\RaggedLeft\arraybackslash}}p{{1.52cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1.52cm}} 
>{{\columncolor[HTML]{{EFEFEF}}\RaggedLeft\arraybackslash}}p{{1.64cm}} 
>{{\columncolor[HTML]{{EFEFEF}}}}p{{1.40cm}} }}
\textbf{{Interest Period Start}} & \textbf{{Interest Period End}} & \textbf{{Interest Rate}} & \textbf{{Spread\hphantom{{A}} to\hphantom{{A}}\hphantom{{A}} Benchmark}} & \RaggedRight\arraybackslash\textbf{{Note\hphantom{{A}} Series\hphantom{{A}}\hphantom{{A}} Principal}} & \RaggedRight\arraybackslash\textbf{{Interest Paid}} & \textbf{{Interest Payment Date}} & \RaggedRight\arraybackslash\textbf{{Related Fund Capital Account}} & \textbf{{Collateral O/C Rate}} \\ \arrayrulecolor{{light_grey}}\hline
{coupon_plot}
\end{{tabular}}
}}\end{{center}}


{{\small
\color{{gray}}
    \noindent\textbf{{Note:}}  The Series Portfolio (which includes the Collateral Securities) is pledged to BNYM as the Collateral Agent for the Noteholders and equals the Note Principal Amount. A copy of the most recent capital account statement is available by request at \underline{{operations@lucidma.com}}.{rets_disclaimer_if_m1}
}}

\pagebreak

\footnotesize
\noindent\textbf{{\color{{lucid_blue}}Report Notes}}

\begin{{enumerate}}
\item Target returns based on the program manager's estimate of the projected returns for the respective series based on current market conditions. 

\item Current coupon (estimated) is based on the rates of the invested portfolio of the Related Fund Interest as of the current period start date.  Actual rate set in arrears based on the final net returns of portfolio.  

\item Annualized net returns quoted on an Act/360 basis after all program costs. Historical returns assume reinvestment at the respective Series return, money market index, Libor rate or T-Bill Index rate at the end of each period.  Libor is the applicable USD London Interbank offered Rate for each calculation period, as published by the ICE Benchmark Administration Fixing. T-Bill is the offer rate for the T-Bills with a maturity matching the respective coupon period (or interpolated rate if the dates do not match). Crane Prime Institutional Money Market Index adjusted to an Actual/360 basis (for Prime notes) and Act/365 for USG Notes, based on the daily average for the periods. A1/P1 CP is dealer placed commercial paper corresponding to the period as published by Bloomberg. SOFR is the term reference rate for the applicable period (e.g. 1m or 3m) as published by the CME Group.

\item Over-Collateralization Rate (``O/C Rate'') of the repo investments securing the notes in the Related Fund Interest as of the business day prior to the end of the most recent period. O/C Rates will vary each period based on the specific risk characteristics of the collateral securities that meet the Lucid risk management standards. O/C Rate equals the market value of the collateral as a proportion of the respective repo investments. Eligible repo collateral details and classifications for the respective Series as fully described in the private offering memorandum. Historical Returns based on the investment program series performance since inception in the related fund series or notes; accordingly, specific Secured Notes that reference the series (e.g. M-1, M-2) purchased in the middle of coupon dates or issued after the inception date may have different returns based on inter-period issuances or different holding periods. Please review the specific interest rates for the respective secured note CUSIP on page 2 for the specific returns applicable to the note.

\item Program AUM based on the amounts invested in the series strategy through the Secured Notes or directly in the related fund entity (Lucid Prime Fund LLC).  For each program series (e.g. Series M), the returns of the Secured Notes and direct investments in the related fund are the same given the program structure. Lucid Platform AUM is the contracted assets under management as of the report date.

\end{{enumerate}}

\noindent Please refer to the Private Placement Memorandum and the Series Supplement for complete details. 

{{\color{{gray}} \noindent The SEC ADV Part 2 firm brochure on the administrator can be accessed via the the following link:

\noindent\underline{{https://files.adviserinfo.sec.gov/IAPD/Content/Common/crd\_iapd\_Brochure.aspx?BRCHR\_VRSN\_ID=832231}}}}

\begin{{center}}\noindent\begin{{tabular}}{{p{{\textwidth}}}}
\rowcolor{{lucid_blue}} 
{{\color[HTML]{{FFFFFF}} \textbf{{Contact Information}}}} \\
\rowcolor[HTML]{{EFEFEF}} 
 \\
\rowcolor[HTML]{{EFEFEF}} 
\textbf{{{issuer_name}, issuer}} \\
\rowcolor[HTML]{{EFEFEF}} 
 c/o Lucid Management and Capital Partners as administrator\\
\rowcolor[HTML]{{EFEFEF}} 
 295 Madison Avenue, 39th Floor\\
\rowcolor[HTML]{{EFEFEF}} 
 New York, NY 10017\\
\rowcolor[HTML]{{EFEFEF}} 
 T: +1-212-551-1704\\
\rowcolor[HTML]{{EFEFEF}} 
 Redemption Notices: \underline{{operations@lucidma.com}}\\
\rowcolor[HTML]{{EFEFEF}} 
 Investor Relations: \underline{{carolina.siles@lucidma.com}}\\
\rowcolor[HTML]{{EFEFEF}} 
 \\
\rowcolor[HTML]{{EFEFEF}} 
\textbf{{Bank of New York Mellon, Issuing \& Paying Agent / Collateral Agent}} \\
\rowcolor[HTML]{{EFEFEF}} 
 240 Greenwich Street, Suite 7E\\
\rowcolor[HTML]{{EFEFEF}} 
 New York, NY 10286\\
\rowcolor[HTML]{{EFEFEF}} 
 T: +1-212-815-5837\\
\rowcolor[HTML]{{EFEFEF}} 
 \underline{{audrey.williams@bnymellon.com}}\\
\rowcolor[HTML]{{EFEFEF}} 
 \\
\rowcolor[HTML]{{EFEFEF}} 
 \textbf{{KPMG, Auditor and Tax Reporting}}\\
\rowcolor[HTML]{{EFEFEF}} 
 345 Park Avenue\\
\rowcolor[HTML]{{EFEFEF}} 
 New York, NY 10154\\
\rowcolor[HTML]{{EFEFEF}} 
 \\
\rowcolor[HTML]{{EFEFEF}} 

\end{{tabular}}
\end{{center}}

\pagebreak

 \ \\
  \ \\
  \ \\
   \ \\
  \ \\
  \ \\

\noindent\textbf{{\color{{lucid_blue}}Important Disclaimer}}

\small

\noindent This material has been prepared by Lucid Management and Capital Partners LP or one of its affiliates, principals or advisors (``Lucid'') and may contain ``forward-looking statements'' which are based on Lucid's beliefs, as well as on a number of assumptions concerning future events, based on information currently available to Lucid. Readers are cautioned not to put undue reliance on such forward-looking statements, which are not a guarantee of future performance, and are subject to a number of uncertainties and other factors, many of which are outside Lucid's control, which could cause actual results to differ materially from such statements.

\noindent This material is for distribution only under such circumstances as may be permitted by applicable law and it is solely intended for qualified institutions and individuals with existing relationships with Lucid, its affiliates or advisers.  It has no regard to the specific investment objectives, financial situation or particular needs of any recipient. It is published solely for informational and discussion purposes only and is not to be construed as a solicitation or an offer to buy or sell any securities, related financial instruments, actual fund or specific transaction. No representation or warranty, either express or implied, is provided in relation to the accuracy, completeness or reliability of the information contained herein, nor is it intended to be a complete statement or summary of the securities, markets or developments referred to in the materials.  It should not be regarded by recipients as a substitute for the exercise of their own judgment. Any opinions expressed in this material are subject to change without notice and may differ or be contrary to opinions expressed by other business areas or groups of Lucid as a result of using different assumptions and criteria. Lucid is under no obligation to update or keep current the information contained herein. Lucid, its partners, officers and employees' or clients may have or have had interests or long or short positions in the securities or other financial instruments referred to herein and may at any time make purchases and/or sales in them as principal or agent. Neither Lucid nor any of its affiliates, nor any of Lucid or any of its affiliates, partners, employees or agents accepts any liability for any loss or damage arising out of the use of all or any part of this material.

\noindent Money market investments including repurchase agreements are not suitable for all investors. Past performance is not necessarily indicative of future results.  Prior to entering into a transaction you should consult with your own legal, regulatory, tax, financial and accounting advisers to the extent you deem necessary to make your own investment, hedging and trading decisions. Any transaction between you and Lucid will be subject to the detailed provisions of a private placement memorandum or a managed account agreement relating to that transaction. Additional information will be made available upon request.

\noindent The indicative information in this document (the ``Information'') are provided to you for information purposes only and may not be complete. Any redistribution of this document without express written consent of Lucid is prohibited.  No part of this material may be reproduced in any form, or referred to in any publication, without express written consent of Lucid.  This presentation should only be considered current as of the date of the publication without regard to the date on which you may have accessed or received the information. Unless required by law or specifically agreed in writing, we have no obligation to continue to provide to you the Information, and we may cease doing so at any time in our sole discretion.

\noindent\underline{{\textit{{Targeted Return Disclosure}}}}

\noindent\textit{{The targeted returns included in this presentation are not intended as, and must not be regarded as, a representation, warranty or prediction that any Fund (or series thereof) will achieve any particular rate of return over any particular time period or that any Fund (or series thereof) will not incur losses. Although Lucid believes, based on these factors, that the referenced return targets are reasonable, return targets are subject to inherent limitations including, without limitation, the fact they cannot take into account the impact of future economic events on future trading and investment decisions. These events may include changes in interest rates and/or benchmarks greater than those occurring within the historical time period examined when developing the return targets, or future changes in laws or regulations. All targeted returns are net of Lucid's anticipated fees and expenses.}}

\end{{document}}

"""

# only thing hardwired: descriptions for various entities. currently most match but might change in future
fund_descriptions = dict()
series_descriptions = dict()
fund_descriptions[
    "USG"
] = r"""The fund series seeks income generation with 100\% capital preservation and invests solely in US government backed securities (USG) and repurchase agreements secured by USG, subject to the proprietary Lucid Investment Process.
"""

fund_descriptions[
    "PRIME"
] = r"""
The fund series seeks income generation with 100\% capital preservation and primarily invests in repurchase agreements secured by US Government Backed securities (USG) and Eligible Collateral Securities as well as other Eligible Investments (T-Bills, A1/P1 Commercial Paper and money market funds).
"""

custom_qx_desc = r"""
The fund series seeks income generation with 100\% capital preservation and primarily invests in repurchase agreements limited to Investment Grade and BB rated Collateral Securities (subject to a 50\% limit on BB collateral), Single-A Bank Guaranteed Collateral or other high quality money market securities. 
"""

series_descriptions[
    "USGFUND M"
] = r"""
\textbf{{Series M}} portfolio assets are limited to 1 month maximum maturities.
"""

series_descriptions[
    "PRIMEFUND M"
] = r"""
\textbf{{Series M}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Highly Rated Investment Grade securities (at least 75\% rated between AAA and A- or USG securities).
"""

series_descriptions[
    "PRIMEFUND M1"
] = r"""
\textbf{{Series M1}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Highly Rated Investment Grade securities (at least 75\% rated between AAA and A- or USG securities).
"""

series_descriptions[
    "PRIMEFUND C1"
] = r"""
\textbf{{Series C1}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Highly Rated Investment Grade securities (at least 75\% rated between AAA and A- or USG securities).
"""

series_descriptions[
    "PRIMEFUND Q1"
] = r"""
\textbf{{Series Q1}} interests have (i) maximum maturities of 3 months on all series assets, (ii) quarterly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Investment Grade securities only or USG securities.
"""

# series_descriptions["PRIMEFUND QX"] = r"""
# \textbf{{Series QX}} interests have (i) maximum maturities of 3 months on all series assets and (ii) quarterly withdrawal dates.
# """

# Terri Change 02/01/2024

series_descriptions[
    "PRIMEFUND QX"
] = r"""
\textbf{{Series QX}} # interests have (i) maximum maturities of 3 months on all series assets, (ii) monthly withdrawal dates.                                                                           
"""

series_descriptions[
    "PRIMEFUND MIG"
] = r"""
\textbf{{Series MIG}} interests have (i) maximum maturities of 1 month on all series assets, (ii) monthly withdrawal dates and (iii) all Eligible Collateral for repurchase agreements must be Investment Grade securities only or USG securities.
"""

# intervals for various entities, default to 3 months/1 year unless specified
intervals = dict()
intervals["PRIMEFUND Q1"] = (6, 12)
intervals["PRIMEFUND QX"] = (6, 12)

# disclaimer for reduced 12/30-1/28 return
m1_1230_128_rets_disclaimer = r"""\\
    
    \noindent\textit{{* Returns for the period from 12/30/20 to 1/28/21 reflects the expected impact of sizable allocations of investor cash into the M1 Series after the notice date, with manager and series investor consent. For amounts invested within the series guidelines and notice dates, returns are expected to remain in line with overall M1 series targets.}}
"""


def secured_by_from(f, s):
    if f == "USG":
        return "US Government backed (USG) securities only"
    else:
        if s == "M":
            return "Highly Rated Investment Grade Collateral securities (at least 75\\% must be rated between AAA and A- or USG securities)"
        elif s == "Q1":
            return "Investment Grade Collateral securities only"
        elif s == "QX":
            return "Investment Grade and BB rated Collateral securities (subject to a 50\\% limit on BB collateral)"
        else:
            return "Highly Rated Investment Grade Collateral securities"


def accs_since_start(ws, r_col, tstart_col, tend_col, end, start, daycount):
    if end < start:
        return 1
    try:
        r = float(ws[r_col + str(end)].value)
        dt = (ws[tend_col + str(end)].value - ws[tstart_col + str(end)].value).days
        return accs_since_start(
            ws, r_col, tstart_col, tend_col, end - 1, start, daycount
        ) * (1 + r * dt / daycount)
    except:
        print(
            "Breaking accrual recursion: "
            + ws.title
            + " "
            + r_col
            + str(start)
            + ":"
            + r_col
            + str(end)
        )
        return 1


def diff_period_rate(t1, t2, daycount, r1, r2):
    try:
        dt = (t2 - t1).days
        outp = ((1.0 * r2 / r1) - 1) * daycount / dt
        return form_as_percent(outp, 2)
    except:
        return "n/a"


def form_as_percent(val, rnd):
    try:
        if float(val) == 0:
            return "-"
        return ("{:." + str(rnd) + "f}").format(100 * val) + "\\%"
    except:
        return "n/a"


def xl_average(ws, col, start, end):
    sum = 0
    count = 0
    for row in range(start, end + 1):
        if ws[col + str(row)].value:
            sum = sum + ws[col + str(row)].value
            count = count + 1
    if count == 0:
        return 1
    return 1.0 * sum / count


# graph dimension adjustments
maxchars = 385


def extraspacefromdesc(varistring):
    if len(varistring) > 500:
        return "9.4em"
    return "5.5em"


def heightmap(varistring):
    if len(varistring) > maxchars:
        return 6.676
    else:
        return 8


# fund & series details , portfolio comp table array stretches
def stretches(varistring):
    if len(varistring) > 500:
        return [1.6, 1.5]  # prime fund m
    return [1.72, 2.55]  # usg fund m


def hspacemap(varistring, nbars):
    if len(varistring) > maxchars:
        if nbars == 8:
            return -0.77
        if nbars == 7:
            return -0.88
        return -0.77  # default here, if 16
    else:
        return -0.9  # default here, if 16


# space between bars in bar chart (mm)
# Terri 02/05/2024 change spacing on M graph from 1.7 to
# return [1.7,1.7,1.7,1.36,0.942,0.674,0.558,0.471,0.395,0.35,0.314,0.103,0.255,0.234,0.215,0.198,0.188][nbars]
# return 0.156 # default here, if 16
def xmap(varistring, nbars):
    # max 16
    nbars = min(nbars, 16)
    if len(varistring) > maxchars:
        return [
            1.7,
            1.7,
            1.7,
            1.40,
            0.942,
            0.674,
            0.558,
            0.471,
            0.395,
            0.35,
            0.103,
            0.103,
            0.255,
            0.234,
            0.215,
            0.198,
            0.188,
        ][nbars]
    else:
        return 0.150  # default here, if 16


## width of each bar in bar chart (mm) Terri was 2.50 default
# def barwidthmap(varistring, nbars):
# 	# max 16
# 	nbars = min(nbars, 16)
# 	if (len(varistring) > maxchars):
# 		return [8,8,8,8,8,6,6,6,3,3,3,3,3,3,2.5,2.5,2.0][nbars]
# 	else:
# 		return 2.5 # default here, if 16


# width of each bar in bar chart (mm) Terri was 2.50 default
def barwidthmap(varistring, nbars):
    # max 16
    nbars = min(nbars, 16)
    if len(varistring) > maxchars:
        return [8, 8, 8, 8, 8, 6, 6, 6, 3, 3, 3, 3, 3, 3, 2.5, 2.5, 2.0][nbars]
    else:
        return 2.5  # default here, if 16


# for notes
def notehspacemap(nbars):
    return -0.79


def notexmap(nbars):
    if nbars == 10:
        return 0.284
    if nbars == 11:
        return 0.258
    if nbars == 12:
        return 0.231
    if nbars == 4:
        return 0.615
    return 0.17


def notebarwidthmap(nbars):
    if nbars == 10:
        return 4
    if nbars == 4 or nbars == 5:
        return 4
    return 2


def xl_max(fn, ws, col, start, end):
    if fn == "USG":
        return 3
    if fn == "Prime":
        return 3.3
    return 3


def tablevstretch(fund_name):
    if fund_name == "USG":
        return 1.2
    else:
        return 1


def return_table_plot(
    fund_name,
    prev_pd_return,
    series_abbrev,
    r_this_1,
    r_this_2,
    comp_a,
    comp_b,
    comp_c,
    r_a,
    r_b,
    r_c,
    s_a_0,
    s_a_1,
    s_a_2,
    s_b_0,
    s_b_1,
    s_b_2,
    s_c_0,
    s_c_1,
    s_c_2,
):
    if fund_name == "USG":
        out = r"""
    \textbf{{Lucid {fund_name} - Series {series_abbrev}}}                    & \textbf{{{prev_pd_return}}}                              & \textbf{{-}}                                  & \textbf{{{r_this_1}}}                               & \textbf{{-}}                           & \textbf{{{r_this_2}}}                             & \textbf{{-}}                          \\
{comp_a}                       & {r_a_0}                                       & \textbf{{{s_a_0}}}                            & {r_a_1}                               & \textbf{{{s_a_1}}}                     & {r_a_2}                              & \textbf{{{s_a_2}}}                    \\
{comp_b}                       & {r_b_0}                                       & \textbf{{{s_b_0}}}                           & {r_b_1}                               & \textbf{{{s_b_1}}}                     & {r_b_2}                              & \textbf{{{s_b_2}}}                    \\ \arrayrulecolor{{light_grey}}\hline
    """
        return out.format(
            fund_name=fund_name,
            prev_pd_return=prev_pd_return,
            series_abbrev=series_abbrev,
            r_this_1=r_this_1,
            r_this_2=r_this_2,
            comp_a=comp_a,
            comp_b=comp_b,
            r_a_0=form_as_percent(r_a[0], 2),
            r_a_1=r_a[1],  # already percent
            r_a_2=r_a[2],
            r_b_0=form_as_percent(r_b[0], 2),
            r_b_1=r_b[1],
            r_b_2=r_b[2],
            s_a_0=s_a_0,
            s_a_1=s_a_1,
            s_a_2=s_a_2,
            s_b_0=s_b_0,
            s_b_1=s_b_1,
            s_b_2=s_b_2,
        )
    else:
        out = r"""
    \textbf{{Lucid {fund_name} - Series {series_abbrev}}}                    & \textbf{{{prev_pd_return}}}                              & \textbf{{-}}                                  & \textbf{{{r_this_1}}}                               & \textbf{{-}}                           & \textbf{{{r_this_2}}}                             & \textbf{{-}}                          \\
{comp_a}                       & {r_a_0}                                       & \textbf{{{s_a_0}}}                            & {r_a_1}                               & \textbf{{{s_a_1}}}                     & {r_a_2}                              & \textbf{{{s_a_2}}}                    \\
{comp_b}                       & {r_b_0}                                       & \textbf{{{s_b_0}}}                           & {r_b_1}                               & \textbf{{{s_b_1}}}                     & {r_b_2}                              & \textbf{{{s_b_2}}}                    \\
{comp_c}                       & {r_c_0}                                       & \textbf{{{s_c_0}}}                            & {r_c_1}                               & \textbf{{{s_c_1}}}                     & {r_c_2}                              & \textbf{{{s_c_2}}}                    \\ \arrayrulecolor{{light_grey}}\hline
    """
        return out.format(
            fund_name=fund_name,
            prev_pd_return=prev_pd_return,
            series_abbrev=series_abbrev,
            r_this_1=r_this_1,
            r_this_2=r_this_2,
            comp_a=comp_a,
            comp_b=comp_b,
            comp_c=comp_c,
            r_a_0=form_as_percent(r_a[0], 2),
            r_a_1=r_a[1],  # already percent
            r_a_2=r_a[2],
            r_b_0=form_as_percent(r_b[0], 2),
            r_b_1=r_b[1],
            r_b_2=r_b[2],
            r_c_0=form_as_percent(r_c[0], 2),
            r_c_1=r_c[1],
            r_c_2=r_c[2],
            s_a_0=s_a_0,
            s_a_1=s_a_1,
            s_a_2=s_a_2,
            s_b_0=s_b_0,
            s_b_1=s_b_1,
            s_b_2=s_b_2,
            s_c_0=s_c_0,
            s_c_1=s_c_1,
            s_c_2=s_c_2,
        )


def addl_coll_breakdown(alloc_aa_a, oc_aa_a, alloc_bbb, oc_bbb, alloc_bb, oc_bb):
    if alloc_aa_a == "n/a":
        return ""
    else:
        outp = r"""IG Repo: AA to A & {alloc_aa_a} & {oc_aa_a} \\
        IG Repo: BBB & {alloc_bbb} & {oc_bbb} \\
        Repo: BB & {alloc_bb} & {oc_bb} \\"""
        return outp.format(
            alloc_aa_a=alloc_aa_a,
            oc_aa_a=oc_aa_a,
            alloc_bbb=alloc_bbb,
            oc_bbb=oc_bbb,
            alloc_bb=alloc_bb,
            oc_bb=oc_bb,
        )


def colltable(
    inclextra,
    secured_by,
    series_abbrev,
    include_aaa_in_usg_bucket,
    alloc_aaa,
    alloc_aa_a,
    alloc_bbb,
    alloc_bb,
    alloc_tbills,
    alloc_total,
    oc_total,
    oc_tbills,
    oc_bb,
    oc_bbb,
    oc_aa_a,
    oc_aaa,
):
    if inclextra:
        out = r"""
        \renewcommand{{\arraystretch}}{{{vstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
        >{{\columncolor[HTML]{{EFEFEF}}}}p{{3.5cm}} 
        >{{\columncolor[HTML]{{EFEFEF}}}}c
        >{{\columncolor[HTML]{{EFEFEF}}}}c!{{\color{{light_grey}}\vrule}}}}
        \arrayrulecolor{{light_grey}}\hline
        \multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{\color[HTML]{{FFFFFF}}\textbf{{Series Collateral Overview\textsuperscript{{4}}}}}}}} \\
        \multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}p{{8.2cm}}!{{\color{{light_grey}}\vrule}}}}{{\rowcolor[HTML]{{EFEFEF}}{{\textbf{{Series {series_abbrev}}}: Secured by \textbf{{{secured_by}}}, with daily valuations \& margining.}}}} \\
        & & \\
         & \textbf{{\% Portfolio}} & \textbf{{O/C Rate}}\\
        {usg_aaa_cat} & {alloc_aaa} & {oc_aaa} \\
        IG Repo: AA to A & {alloc_aa_a} & {oc_aa_a} \\
        IG Repo: BBB & {alloc_bbb} & {oc_bbb} \\
        Repo: BB & {alloc_bb} & {oc_bb} \\
        T-Bills; Gov't MMF & {alloc_tbills} & {oc_tbills} \\ \cline{{2-2}} \cline{{3-3}} 
        \textbf{{Total}} & {alloc_total} & \textbf{{{oc_total}}} \\\arrayrulecolor{{light_grey}}\hline
        \end{{tabular}}
        """
        return out.format(
            vstretch=1.48,
            secured_by=secured_by,
            series_abbrev=series_abbrev,
            usg_aaa_cat=(
                "US Govt/AAA Repo" if include_aaa_in_usg_bucket else "US Govt Repo"
            ),
            alloc_aaa=alloc_aaa,
            alloc_aa_a=alloc_aa_a,
            alloc_bbb=alloc_bbb,
            alloc_bb=alloc_bb,
            alloc_tbills=alloc_tbills,
            alloc_total=alloc_total,
            oc_total=oc_total,
            oc_tbills=oc_tbills,
            oc_bb=oc_bb,
            oc_bbb=oc_bbb,
            oc_aa_a=oc_aa_a,
            oc_aaa=oc_aaa,
        )
    else:
        out = r"""
        \renewcommand{{\arraystretch}}{{{vstretch}}}\begin{{tabular}}{{!{{\color{{light_grey}}\vrule}}
        >{{\columncolor[HTML]{{EFEFEF}}}}p{{3.5cm}} 
        >{{\columncolor[HTML]{{EFEFEF}}}}c
        >{{\columncolor[HTML]{{EFEFEF}}}}c!{{\color{{light_grey}}\vrule}}}}
        \arrayrulecolor{{light_grey}}\hline
        \multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}l!{{\color{{light_grey}}\vrule}}}}{{\rowcolor{{lucid_blue}}{{\color[HTML]{{FFFFFF}}\textbf{{Series Collateral Overview\textsuperscript{{4}}}}}}}} \\
        \multicolumn{{3}}{{!{{\color{{light_grey}}\vrule}}p{{8.2cm}}!{{\color{{light_grey}}\vrule}}}}{{\rowcolor[HTML]{{EFEFEF}}{{\textbf{{Series {series_abbrev}}}: Secured by \textbf{{{secured_by}}}, with daily valuations \& margining.}}}} \\
        & & \\
         & \textbf{{\% Portfolio}} & \textbf{{O/C Rate}}\\
        {usg_aaa_cat} & {alloc_aaa} & {oc_aaa} \\
        T-Bills; Gov't MMF & {alloc_tbills} & {oc_tbills} \\ \cline{{2-2}} \cline{{3-3}} 
        \textbf{{Total}} & {alloc_total} & \textbf{{{oc_total}}} \\\arrayrulecolor{{light_grey}}\hline
        \end{{tabular}}
        """
        return out.format(
            vstretch=1.91,
            secured_by=secured_by,
            series_abbrev=series_abbrev,
            usg_aaa_cat=(
                "US Govt/AAA Repo" if include_aaa_in_usg_bucket else "US Govt Repo"
            ),
            alloc_aaa=alloc_aaa,
            # alloc_cp=alloc_cp,
            alloc_tbills=alloc_tbills,
            alloc_total=alloc_total,
            oc_total=oc_total,
            oc_tbills=oc_tbills,
            # oc_cp=oc_cp,
            oc_aaa=oc_aaa,
        )


# max = float("-inf")
# try:
#   for row in range(start, end + 1):
#       if ws[col+str(row)].value:
#           if (100* ws[col+str(row)].value) > max:
#               max = 100 * ws[col+str(row)].value
# except:
#   maxval = 3
#   print("Defaulting to max of " + str(maxval) + " for plotting.")
#   return maxval
# return max


def plotify(ws, x_col, y_col, start, end):
    outp = ""
    for row in range(start, end + 1):
        if ws[x_col + str(row)].value and ws[y_col + str(row)].value:
            outp = (
                outp
                + "("
                + ws[x_col + str(row)].value.strftime("%Y-%m-%d")
                + ","
                + str(round(100 * ws[y_col + str(row)].value, 3))
                + ") "
            )
    return outp


def snapshot_graph(
    hspace,
    graphwidth,
    graphheight,
    maxreturn,
    series_abbrev,
    comp_a,
    comp_b,
    comp_c,
    this_r,
    ra,
    rb,
    rc,
):
    if comp_c is not None:
        out = r"""
          \hspace*{{{hspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
    \begin{{axis}}[
        title style = {{font = \small}},
        axis line style = {{light_grey}},
        title={{{{[TONY WAS HERE]Performance vs Benchmarks}}}},
        ymin=3, ymax={maxreturn}, %MAXRETURN HERE
       symbolic x coords={{Series {series_abbrev},{comp_a},{comp_b},{comp_c}}},
        xtick={{Series {series_abbrev},{comp_a},{comp_b},{comp_c}}},
        x tick label style={{anchor=north,font=\scriptsize,/pgf/number format/assume math mode}},
        yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
        y tick label style = {{/pgf/number format/.cd,
                fixed,
                fixed zerofill,
                precision=2,
                /pgf/number format/assume math mode
        }},
        ytick distance={tickdist},
        bar width = 8mm, x = 2.46cm,
        xtick pos=bottom,ytick pos=left,
        every node near coord/.append style={{font=\fontsize{{8}}{{8}}\selectfont,/pgf/number format/.cd,
                fixed,
                fixed zerofill,
                precision=2,/pgf/number format/assume math mode}},
        ]
    \addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
        coordinates {{
            (Series {series_abbrev},{this_r}) 
        }};
    \addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
        coordinates {{
            ({comp_a},{ra}) 
        }};
    \addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
        coordinates {{
            ({comp_b},{rb}) 
        }};
    \addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
        coordinates {{
            ({comp_c},{rc}) 
        }};
    \end{{axis}}
        \end{{tikzpicture}}}}

        """
        return out.format(
            hspace=hspace,
            tickdist=0.5 if max(this_r, ra, rb, rc) + 0.1 >= 1 else 0.25,
            graphwidth=graphwidth,
            graphheight=graphheight,
            maxreturn=max(this_r, ra, rb, rc) + 0.5,
            series_abbrev=series_abbrev,
            comp_a=benchmark_shorten(comp_a) if "CRANE" in comp_a.upper() else comp_a,
            comp_b=benchmark_shorten(comp_b) if "CRANE" in comp_b.upper() else comp_b,
            comp_c=benchmark_shorten(comp_c) if "CRANE" in comp_c.upper() else comp_c,
            this_r=this_r,
            ra=ra,
            rb=rb,
            rc=rc,
        )
    else:
        out = r"""
      \hspace*{{{hspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
\begin{{axis}}[
    title style = {{font = \small}},
    axis line style = {{light_grey}},
        title={{{{[WHO IS TONY]Performance vs Benchmarks}}}},
    ymin=2, ymax={maxreturn}, %MAXRETURN HERE
   symbolic x coords={{Series {series_abbrev},{comp_a},{comp_b}}},
    xtick={{Series {series_abbrev},{comp_a},{comp_b}}},
    x tick label style={{anchor=north,font=\scriptsize,/pgf/number format/assume math mode}},
    yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
    y tick label style = {{/pgf/number format/.cd,
            fixed,
            fixed zerofill,
            precision=2,
            /pgf/number format/assume math mode
    }},
    ytick distance={tickdist},
    bar width = 10mm, x = 3.7cm,
    xtick pos=bottom,ytick pos=left,
    every node near coord/.append style={{font=\fontsize{{8}}{{8}}\selectfont,/pgf/number format/.cd,
            fixed,
            fixed zerofill,
            precision=2,/pgf/number format/assume math mode}},
    ]
\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
    coordinates {{
        (Series {series_abbrev},{this_r}) 
    }};
\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
    coordinates {{
        ({comp_a},{ra}) 
    }};
\addplot[ybar, nodes near coords, fill=dark_grey, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
    coordinates {{
        ({comp_b},{rb}) 
    }};
\end{{axis}}
    \end{{tikzpicture}}}}

    """
    return out.format(
        hspace=hspace,
        tickdist=0.5 if maxreturn >= 1 else 0.25,
        graphwidth=graphwidth,
        graphheight=graphheight,
        maxreturn=max(this_r, ra, rb, rc) + 0.07,
        series_abbrev=series_abbrev,
        comp_a=benchmark_shorten(comp_a) if "CRANE" in comp_a.upper() else comp_a,
        comp_b=benchmark_shorten(comp_b) if "CRANE" in comp_b.upper() else comp_b,
        this_r=this_r,
        ra=ra,
        rb=rb,
    )


def performance_graph(
    titleincl,
    graphhspace,
    graphwidth,
    graphheight,
    fund_name,
    zero_date,
    max_return,
    graphx,
    graphbarwidth,
    return_plot,
    comp_a_plot,
    comp_b_plot,
    series_abbrev,
    comp_a,
    comp_b,
):
    out = r"""
      \hspace*{{{graphhspace}cm}}\resizebox {{{graphwidth}}} {{{graphheight}}} {{\begin{{tikzpicture}}
\begin{{axis}}[
    title style = {{font = \small}},
    axis line style = {{light_grey}},{title}
    date coordinates in=x, date ZERO={zero_date},
    xticklabel=\month/\day/\year,  
    ymin=0, ymax={max_return}, %MAXRETURN HERE
    legend cell align = {{left}},
    legend style={{at={{(0.05,1.0)}},
      anchor=north east, font=\small, draw=none,fill=none}},
      x={graphx}mm, %CHANGE THIS to tighten in graph, eg if quarterly
    bar width={graphbarwidth}mm, ybar=2pt, %bar width is width, ybar is space between
   % symbolic x coords={{Firm 1, Firm 2, Firm 3, Firm 4, Firm 5}},
    xtick=data,
    x tick label style={{rotate=90,anchor=east,font=\tiny,/pgf/number format/assume math mode}},
         yticklabel=\pgfmathparse{{\tick}}\pgfmathprintnumber{{\pgfmathresult}}\,\%,
    y tick label style = {{/pgf/number format/.cd,
            fixed,
            fixed zerofill,
            precision=2,
            /pgf/number format/assume math mode
    }},
    nodes near coords align={{vertical}},
    ytick distance=0.5,
    xtick pos=bottom,ytick pos=left,
    every node near coord/.append style={{font=\fontsize{{6}}{{6}}\selectfont,/pgf/number format/.cd,
            fixed,
            fixed zerofill,
            precision=2,/pgf/number format/assume math mode}},
    ]
%\addplot[ybar, nodes near coords, fill=blue] 
\addplot[ybar, nodes near coords, fill=lucid_blue, rounded corners=1pt,blur shadow={{shadow yshift=-1pt, shadow xshift=1pt}}] 
    coordinates {{
        {return_plot}
    }};
\addplot[draw=dark_red,ultra thick,smooth] 
    coordinates {{
        {comp_a_plot}
    }};
\addplot[draw=dark_color,ultra thick,smooth] 
    coordinates {{
        {comp_b_plot}
    }};
\legend{{\hphantom{{A}}{fund_name} Series {series_abbrev},\hphantom{{A}}{comp_a},\hphantom{{A}}{comp_b}}}
\end{{axis}}
    \end{{tikzpicture}}}}

    """
    # Terri graph
    return out.format(
        graphhspace=graphhspace,
        graphwidth=graphwidth,
        graphheight=graphheight,
        fund_name=fund_name,
        title=(
            r"""
        title={{[WHAT IS GOING]Performance vs Benchmark}},"""
            if titleincl
            else ""
        ),
        zero_date=zero_date,
        max_return=max_return,
        graphx=graphx,
        graphbarwidth=graphbarwidth,
        return_plot=return_plot,
        comp_a_plot=comp_a_plot,
        comp_b_plot=comp_b_plot,
        series_abbrev=series_abbrev,
        comp_a=comp_a,
        comp_b=comp_b,
    )


def coupon_plotify(ws, crow, daycount):
    outp = r""""""
    maxrows = 12
    startrow = max(7, crow - maxrows + 1)
    remaining_rows = max(0, maxrows - (crow + 2 - startrow))
    for i in range(startrow, crow + 2):
        addl = ws["E" + str(i)].value.strftime("%m/%d/%y") + " &"
        addl = (
            addl
            + r"""\textbf{{"""
            + ws["F" + str(i)].value.strftime("%m/%d/%y")
            + r"""}} &"""
        )
        addl = (
            addl
            + r"""\textbf{{"""
            + form_as_percent(ws[("N" if daycount == 360 else "O") + str(i)].value, 2)
            + (r"""{{\tiny (Est'd)}}""" if i == crow + 1 else "")
            + (
                "*"
                if ws["C12"].value == "Monthly1"
                and ws["E" + str(i)].value.strftime("%m/%d/%y") == "12/30/20"
                else ""
            )
            + r"""}} &"""
        )
        addl = (
            addl
            + benchmark_shorten(ws["U" + str(i)].value)
            + ("+" if int(10000 * ws["W" + str(i)].value) > 0 else "-")
            + str(int(abs(ws["X" + str(i)].value)))
            + " &"
        )
        addl = (
            addl
            + (r"\$" if i == startrow else r"\hphantom{{\$}}")
            + "{:,.0f}".format(ws["R" + str(i)].value)
            + " &"
        )
        addl = (
            addl
            + (
                r"\hphantom{{\$}}n/a"
                if i == crow + 1
                else (
                    (r"\$" if i == startrow else r"\hphantom{{\$}}")
                    + "{:,.2f}".format(ws["T" + str(i)].value)
                )
            )
            + " &"
        )
        addl = (
            addl
            + ws[("S" if i < crow + 1 else "F") + str(i)].value.strftime("%m/%d/%y")
            + " &"
        )
        addl = (
            addl
            + (r"\$" if i == startrow else r"\hphantom{{\$}}")
            + "{:,.0f}".format(ws["R" + str(i)].value)
            + " &"
        )
        addl = (
            addl
            + form_as_percent(ws["AP" + str(i)].value, 1)
            + (r" \\" if i <= crow + 1 or remaining_rows > 0 else " ")
        )
        outp = outp + addl

    for i in range(remaining_rows):
        addl = r"""
        """
        addl = addl + (
            r"""& & & & & & & &      \\ """ if i < remaining_rows - 1 else ""
        )
        outp = (
            outp
            + addl
            + r"""

        """
        )
    return outp


# DEPRECATED
def hardcoded_exp_ratio(f, s):
    if f.upper() == "USG":
        return 15.2
    if f.upper() == "PRIME":
        if s == "M":
            return 26.0
        if s == "C1":
            return 27.7
        if s == "MIG":
            return 34.1
        if s == "Q1":
            return 38.2
        if s == "M1":
            return 14.5


def hardcoded_exp_cap(f, s):
    if f.upper() == "USG":
        return 24.5
    if f.upper() == "PRIME":
        if s == "MIG":
            return 39.75
        if s == "Q1":
            return 49.75
        if s == "QX":
            return 49.75
    return 32.75  # default


def exp_rat_footnote(incl, cap, rat):
    if False:
        out = r"""Fund Series expense ratio currently capped at an all-in ratio of {cap} bps and can vary over time. The average expense ratio has been {rat} bps since inception."""
        return out.format(cap=cap, rat=rat)
    else:
        out = r"""Fund Series expense ratio currently capped at an all-in ratio of {cap} bps and can vary over time."""
        return out.format(cap=cap)


def wordify(val):
    try:
        prefix = 0
        if val >= 10**9:
            prefix = round(round(val, 9) / (10**9), 2)
            suffix = " billion"
        elif val >= 10**6:
            prefix = round(round(val, 6) / (10**6), 1)
            suffix = " million"
        elif val >= 10**3:
            prefix = int(round(val, 3) / (10**3))
            suffix = ",000"
        return r"\$" + str(prefix) + suffix
    except:
        return "n/a"


def month_wordify(months):
    if months < 12:
        return str(int(months)) + " Months" if int(months) != 1 else " Month"
    else:
        return (
            str(int(months / 12))
            + (" Years" if int(months / 12) != 1 else " Year")
            + ("" if months % 12 == 0 else (str(months % 12) + " Months"))
        )


def benchmark_shorten(s):
    if s == None:
        print("benchmark none")
        return ""
    if "1" in s.upper() and "LIBOR" in s.upper():
        return "1mL"
    if "3" in s.upper() and "LIBOR" in s.upper():
        return "3mL"
    if "1" in s.upper() and "BILL" in s.upper():
        return "1m TB"
    if "3" in s.upper() and "BILL" in s.upper():
        return "3m TB"
    if "CRANE GOV" in s.upper():
        return "Crane Govt"
    if "CRANE PRIME" in s.upper():
        return "Crane Prime"
    return s


# assumes each is in format "-4.342 \\%"
def bps_spread(t, b):
    try:
        val = round(float(t[0 : t.index("\\")]) - float(b[0 : b.index("\\")]), 2)
        return (
            "-"
            if int(abs(val) * 100) == 0
            else (
                ("+" if int(val * 100) > 0 else "-") + str(int(abs(val) * 100)) + " bps"
            )
        )
    except:
        return "n/a"


def issuer_from_fundname(f):
    if f == None:
        print("issuer none")
    if f.upper() == "USG":
        return "USG Assets LLC"
    if f.upper() == "PRIME":
        return "Prime Notes LLC"
    return f


def declare_ratings_org(r):
    if r == None:
        return ""
    if "EGAN JONES" in r.upper():
        return r + " (NAIC 1)"
    return r


def series_from_note(f, ent_name):
    if ent_name == "Monthly":
        return "M"
    elif ent_name == "Quarterly" or ent_name == "Quarterly1":
        return "Q1"
    elif ent_name == "Monthly1":
        return "M1"
    elif ent_name == "MonthlyIG":
        return "MIG"
    elif ent_name == "QuarterlyX":
        return "QX"
    elif ent_name == "Custom1":
        return "C1"
    else:
        return ""


def fund_inception_from_name(fund):
    if fund == "USG":
        return "June 29, 2017"
    if fund == "Prime":
        return "July 20, 2018"
        return "n/a"


print("Fetching data...")

try:
    MASTER_FILEPATH = "S:/Mandates/Funds/Fund Reporting/Master Data.xlsx"
    # MASTER_FILEPATH = "C:/Users/Lucid Trading/Desktop/tmp_trash/Master Data.xlsx"
    wb = op.load_workbook(PureWindowsPath(Path(MASTER_FILEPATH)))
except:
    print("Error fetching data.")
    exit()

reports_generated = []
bad_reports = []

# find overview page
bigsheet = ""
for ws in wb.worksheets:
    if ws.title == "Platform Data":
        bigsheet = ws

prime_graph_coup = ""  # TODO temporarily hardwired
usg_rets = []
usg_rets_mid = []
primem_rets = []
consolidated_monthly_series = [
    # "PRIMEFUND M",
    "PRIMEFUND C1",
    # "PRIMEFUND M1",
]  # TODO temporarily hardwired

for ws in wb.worksheets:
    try:
        if (ws.title != "Mandate Template") and (ws["B2"].value == "Mandate Data"):
            report_name = ws.title

            if not (ws["C29"].value):
                continue

            print("*****" + report_name + "*****")
            crow = 7

            # find relevant period row
            while ws["F" + str(crow)].value:
                if ws["F" + str(crow)].value == ws["C23"].value:
                    break
                crow = crow + 1

            prev_pd_start = ws["E" + str(crow)].value
            this_pd_start = ws["F" + str(crow)].value
            print(
                "For period "
                + prev_pd_start.strftime("%m/%d")
                + " - "
                + this_pd_start.strftime("%m/%d")
            )

            overview_row = 7
            while bigsheet["B" + str(overview_row)].value:
                if bigsheet["B" + str(overview_row)].value == this_pd_start:
                    break
                overview_row = overview_row + 1
            if not bigsheet["B" + str(overview_row)].value:
                print("ERROR: Overview row not found for this period. Continuing...")
                continue

            report_date = datetime.now()
            lucid_aum = bigsheet["H" + str(overview_row)].value  # post sub/redemp
            program_size = 0
            for col in "CDEFG":
                if bigsheet[col + "6"].value:
                    if bigsheet[col + "6"].value.upper() == ws["C9"].value.upper():
                        program_size = bigsheet[
                            col + str(overview_row)
                        ].value  # post sub/redemp...
                        break
            # todo hardcode consolidated series buckets here
            consolidated_monthly_bucket_size = 0
            for col in "IJKLMNOPQRSTUVWXYZ":
                if bigsheet[col + "6"].value:
                    if bigsheet[col + "6"].value.upper() in consolidated_monthly_series:
                        consolidated_monthly_bucket_size = (
                            consolidated_monthly_bucket_size
                            + bigsheet[col + str(overview_row)].value
                        )  # post sub/redemp...

            # consolidated_monthly_bucket_size = 909650767.36
            # returns for each comparable, a/b/c, just taken in order from cols on sheet
            daycount = ws["C25"].value
            # interval_tuple = (3, 12) if not (ws["C6"].value.upper() in intervals.keys()) else intervals[ws["C6"].value.upper()]
            interval_tuple = (3, 12)
            if "QUARTERLY" in ws["C15"].value.upper():
                interval_tuple = (2, 4)  # because one row = 3 months
            # print(ws["F" + str(crow - interval_tuple[0])].value)
            r_a = (
                ws["Y" + str(crow)].value,
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[0])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Y", "E", "F", crow - interval_tuple[0], 7, daycount
                        ),
                        accs_since_start(ws, "Y", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[0] >= 7)
                    else "n/a"
                ),
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[1])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Y", "E", "F", crow - interval_tuple[1], 7, daycount
                        ),
                        accs_since_start(ws, "Y", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[1] >= 7)
                    else "n/a"
                ),
            )
            r_b = (
                ws["Z" + str(crow)].value,
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[0])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Z", "E", "F", crow - interval_tuple[0], 7, daycount
                        ),
                        accs_since_start(ws, "Z", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[0] >= 7)
                    else "n/a"
                ),
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[1])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "Z", "E", "F", crow - interval_tuple[1], 7, daycount
                        ),
                        accs_since_start(ws, "Z", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[1] >= 7)
                    else "n/a"
                ),
            )
            r_c = (
                ws["AA" + str(crow)].value,
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[0])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "AA", "E", "F", crow - interval_tuple[0], 7, daycount
                        ),
                        accs_since_start(ws, "AA", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[0] >= 7)
                    else "n/a"
                ),
                (
                    diff_period_rate(
                        ws["F" + str(crow - interval_tuple[1])].value,
                        this_pd_start,
                        daycount,
                        accs_since_start(
                            ws, "AA", "E", "F", crow - interval_tuple[1], 7, daycount
                        ),
                        accs_since_start(ws, "AA", "E", "F", crow, 7, daycount),
                    )
                    if (crow - interval_tuple[1] >= 7)
                    else "n/a"
                ),
            )
            r_this_1 = (
                diff_period_rate(
                    ws["F" + str(crow - interval_tuple[0])].value,
                    this_pd_start,
                    daycount,
                    accs_since_start(
                        ws,
                        "N" if daycount == 360 else "O",
                        "E",
                        "F",
                        crow - interval_tuple[0],
                        7,
                        daycount,
                    ),
                    accs_since_start(
                        ws, "N" if daycount == 360 else "O", "E", "F", crow, 7, daycount
                    ),
                )
                if (crow - interval_tuple[0] >= 7)
                else "n/a"
            )

            r_this_2 = (
                diff_period_rate(
                    ws["F" + str(crow - interval_tuple[1])].value,
                    this_pd_start,
                    daycount,
                    accs_since_start(
                        ws,
                        "N" if daycount == 360 else "O",
                        "E",
                        "F",
                        crow - interval_tuple[1],
                        7,
                        daycount,
                    ),
                    accs_since_start(
                        ws, "N" if daycount == 360 else "O", "E", "F", crow, 7, daycount
                    ),
                )
                if (crow - interval_tuple[1] >= 7)
                else "n/a"
            )

            # plotting info
            ts_row_start = max(7, crow - 15)

            ts_row_end = crow
            if "QUARTERLY" in ws["C15"].value.upper():
                interval_tuple = (6, 12)  # now revise to months count

            # TODO temporary hardwired

            if ws["C6"].value.upper() == "USGFUND M":
                usg_rets = [r_this_2, r_a[2], r_b[2], r_c[2]]
                usg_rets_mid = [r_this_1, r_a[1], r_b[1], r_c[1]]
                print(usg_rets)
            if ws["C6"].value.upper() == "PRIMEFUND M":
                primem_rets = [r_this_2, r_a[2], r_b[2], r_c[2]]
                print("Storing...")
                print(primem_rets)

            if ws["C5"].value.upper() == "NOTE" and ws["C9"].value == "USG":
                r_this_2 = usg_rets[0]
                r_this_1 = usg_rets_mid[0]
                r_a = (r_a[0], usg_rets_mid[1], usg_rets[1])
                r_b = (r_b[0], usg_rets_mid[2], usg_rets[2])
                r_c = (r_c[0], usg_rets_mid[3], usg_rets[3])

            if ws["C6"].value.upper() == "PRIMENOTE M-1":
                print(primem_rets)
                r_this_2 = primem_rets[0]
                r_a = (r_a[0], r_a[1], primem_rets[1])
                r_b = (r_b[0], r_b[1], primem_rets[2])
                r_c = (r_c[0], r_c[1], primem_rets[3])

            # if (ws["C6"].value.upper() == "PRIMEFUND M"):
            #   prime_graph_coup = performance_graph(
            #           True,
            #           hspacemap(fund_descriptions[ws["C9"].value.upper()] + series_descriptions[ws["C6"].value.upper()], ts_row_end-ts_row_start + 1),
            #           "!",
            #           str(heightmap(fund_descriptions[ws["C9"].value.upper()] + series_descriptions[ws["C6"].value.upper()]))+"cm",
            #           ws["F" + str(ts_row_start)].value.strftime("%Y-%m-%d"),
            #           3.2, #TODO max return here, change
            #           xmap(fund_descriptions[ws["C9"].value.upper()] + series_descriptions[ws["C6"].value.upper()], ts_row_end-ts_row_start + 1),
            #           barwidthmap(fund_descriptions[ws["C9"].value.upper()] + series_descriptions[ws["C6"].value.upper()], ts_row_end-ts_row_start + 1),
            #           plotify(ws, 'F', "N" if daycount == 360 else "O", ts_row_start, ts_row_end),
            #           plotify(ws, 'F', 'Y', ts_row_start, ts_row_end),
            #           plotify(ws, 'F', 'Z', ts_row_start, ts_row_end),
            #           ws["C11"].value,
            #           ws["Y6"].value,
            #           ws["Z6"].value
            #         )

            # # end hardwired

            # populate template with parameters
            print("Populating report template...")
            script = ""
            if ws["C5"].value.upper() == "FUND":  # fund (series) report template
                # populate
                script = fund_report_template.format(
                    report_date=report_date.strftime("%B %d, %Y"),
                    fundname=ws["C9"].value,
                    toptableextraspace=extraspacefromdesc(
                        fund_descriptions[ws["C9"].value.upper()]
                        + series_descriptions[ws["C6"].value.upper()]
                    ),
                    series_abbrev=ws["C11"].value
                    + (" (NYP Custom)" if ws["C11"].value == "C1" else ""),
                    port_limit="Quarterly" if "Q" in ws["C11"].value else "Monthly",
                    seriesname=ws["C12"].value,
                    fund_description=(
                        custom_qx_desc
                        if ws["C6"].value.upper() == "PRIMEFUND QX"
                        else fund_descriptions[ws["C9"].value.upper()]
                    ),
                    series_description=series_descriptions[ws["C6"].value.upper()],
                    benchmark=ws["U" + str(crow + 1)].value,  # TODO ENSURE THERE
                    tgt_outperform=ws["C26"].value,  # TODO ENSURE THERE
                    exp_rat_footnote=exp_rat_footnote(
                        True,
                        hardcoded_exp_cap(ws["C9"].value, ws["C11"].value),
                        round(xl_average(ws, "AD", 7, crow), 1),
                    ),
                    prev_pd_start=prev_pd_start.strftime("%b %d"),
                    this_pd_start=this_pd_start.strftime("%b %d"),
                    prev_pd_return=form_as_percent(
                        ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                    ),
                    prev_pd_benchmark=benchmark_shorten(ws["U" + str(crow)].value),
                    prev_pd_outperform=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_a[0], 2),
                    )[1:],
                    this_pd_end=ws["C24"].value.strftime(
                        "%b %d"
                    ),  # TODO ENSURE matches
                    this_pd_est_return=form_as_percent(
                        ws[("N" if daycount == 360 else "O") + str(crow + 1)].value, 2
                    ),
                    # TODO ensure there
                    this_pd_est_outperform=int(
                        10000 * ws["W" + str(crow + 1)].value
                    ),  # TODO ENSURE THERE
                    benchmark_short=benchmark_shorten(ws["U" + str(crow + 1)].value),
                    interval1=month_wordify(interval_tuple[0]),
                    interval2=month_wordify(interval_tuple[1]),
                    descstretch=stretches(
                        fund_descriptions[ws["C9"].value.upper()]
                        + series_descriptions[ws["C6"].value.upper()]
                    )[0],
                    pcompstretch=stretches(
                        fund_descriptions[ws["C9"].value.upper()]
                        + series_descriptions[ws["C6"].value.upper()]
                    )[1],
                    addl_coll_breakdown=addl_coll_breakdown(
                        (
                            form_as_percent(ws["AK" + str(crow)].value, 1)
                            if ws["C9"].value != "USG"
                            else "n/a"
                        ),
                        form_as_percent(ws["AR" + str(crow)].value, 1),
                        (
                            form_as_percent(ws["AL" + str(crow)].value, 1)
                            if ws["C9"].value != "USG"
                            else "n/a"
                        ),
                        form_as_percent(ws["AS" + str(crow)].value, 1),
                        (
                            form_as_percent(ws["AM" + str(crow)].value, 1)
                            if ws["C9"].value != "USG"
                            else "n/a"
                        ),
                        form_as_percent(ws["AT" + str(crow)].value, 1),
                    ),
                    oc_aaa=form_as_percent(ws["AQ" + str(crow)].value, 1),
                    oc_tbills="-",
                    oc_total=form_as_percent(ws["AP" + str(crow)].value, 1),
                    usg_aaa_cat=(
                        "US Govt Repo"
                        if ws["C9"].value == "USG"
                        else "US Govt/AAA Repo"
                    ),
                    alloc_aaa=form_as_percent(ws["AJ" + str(crow)].value, 1),
                    alloc_tbills=form_as_percent(ws["AO" + str(crow)].value, 1),
                    alloc_total=form_as_percent(1, 1),
                    tablevstretch=tablevstretch(ws["C9"].value),  # only for fund report
                    return_table_plot=return_table_plot(
                        fund_name=ws["C9"].value,
                        prev_pd_return=form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        series_abbrev=ws["C11"].value,
                        r_this_1=r_this_1,
                        r_this_2=r_this_2,
                        comp_a=ws["Y6"].value,
                        comp_b=ws["Z6"].value,
                        comp_c=ws["AA6"].value,
                        r_a=r_a,
                        r_b=r_b,
                        r_c=r_c,
                        s_a_0=bps_spread(
                            form_as_percent(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            form_as_percent(r_a[0], 2),
                        ),  # TODO daycounts for spreads and round to 2 places
                        s_a_1=bps_spread(r_this_1, r_a[1]),
                        s_a_2=bps_spread(r_this_2, r_a[2]),
                        s_b_0=bps_spread(
                            form_as_percent(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            form_as_percent(r_b[0], 2),
                        ),
                        s_b_1=bps_spread(r_this_1, r_b[1]),
                        s_b_2=bps_spread(r_this_2, r_b[2]),
                        s_c_0=bps_spread(
                            form_as_percent(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            form_as_percent(r_c[0], 2),
                        ),
                        s_c_1=bps_spread(r_this_1, r_c[1]),
                        s_c_2=bps_spread(r_this_2, r_c[2]),
                    ),
                    fund_size=wordify(program_size),
                    series_size=wordify(
                        ws["G" + str(crow + 1)].value
                    ),  # post sub/redemp, TODO temporarily consold hardwired
                    lucid_aum=wordify(lucid_aum),
                    rating=ws["C13"].value,
                    rating_org=ws["C14"].value,
                    calc_frequency=ws["C17"].value,
                    next_withdrawal_date=ws["C24"].value.strftime("%B %d, %Y"),
                    next_notice_date=ws["C22"].value.strftime("%B %d, %Y"),
                    min_invest=wordify(ws["C21"].value),
                    wal=(ws["C24"].value - this_pd_start).days,
                    legal_fundname=ws["C10"].value,
                    fund_inception=ws["C18"].value.strftime("%B %d, %Y"),
                    series_inception=ws["E7"].value.strftime("%B %d, %Y"),
                    # max_return =
                    performance_graph=(
                        prime_graph_coup
                        if ws["C6"].value.upper() == "PRIMENOTE M-1"
                        else (
                            performance_graph(
                                True,
                                hspacemap(
                                    fund_descriptions[ws["C9"].value.upper()]
                                    + series_descriptions[ws["C6"].value.upper()],
                                    ts_row_end - ts_row_start + 1,
                                ),
                                "!",
                                str(
                                    heightmap(
                                        fund_descriptions[ws["C9"].value.upper()]
                                        + series_descriptions[ws["C6"].value.upper()]
                                    )
                                )
                                + "cm",
                                ws["C9"].value,
                                ws["F" + str(ts_row_start)].value.strftime("%Y-%m-%d"),
                                # max(xl_max(ws["C9"].value, ws, ("N" if daycount == 360 else "O"), ts_row_start, ts_row_end),xl_max(ws["C9"].value, ws, "Y", ts_row_start, ts_row_end),xl_max(ws["C9"].value, ws, "Z", ts_row_start, ts_row_end)),
                                (
                                    6
                                    if ws["C9"].value == "USG"
                                    else (
                                        7
                                        if ws["C11"].value == "MIG"
                                        else (
                                            7
                                            if ws["C11"].value == "Q1"
                                            else 7 if ws["C11"].value == "QX" else 6.75
                                        )
                                    )
                                ),  # TODO max return here hardwired graphmax maxreturn max_return
                                (
                                    xmap(
                                        fund_descriptions[ws["C9"].value.upper()]
                                        + series_descriptions[ws["C6"].value.upper()],
                                        ts_row_end - ts_row_start + 1,
                                    )
                                    if ws["C6"].value.upper() != "PRIMEFUND Q1"
                                    else 0.065
                                ),
                                barwidthmap(
                                    fund_descriptions[ws["C9"].value.upper()]
                                    + series_descriptions[ws["C6"].value.upper()],
                                    ts_row_end - ts_row_start + 1,
                                ),
                                plotify(
                                    ws,
                                    "F",
                                    "N" if daycount == 360 else "O",
                                    ts_row_start,
                                    ts_row_end,
                                ),
                                plotify(ws, "F", "Y", ts_row_start, ts_row_end),
                                plotify(ws, "F", "Z", ts_row_start, ts_row_end),
                                ws["C11"].value,
                                ws["Y6"].value,
                                ws["Z6"].value,
                            )
                            if (ts_row_end - ts_row_start + 1 > 3)
                            else snapshot_graph(
                                -0.8,
                                "!",
                                "6.676cm",
                                0.25
                                + max(
                                    round(
                                        ws[
                                            ("N" if daycount == 360 else "O")
                                            + str(crow)
                                        ].value
                                        * 100,
                                        2,
                                    ),
                                    round(r_a[0] * 100, 2),
                                    round(r_b[0] * 100, 2),
                                    round(r_c[0] * 100, 2),
                                )
                                + 0.28,
                                ws["C11"].value,
                                ws["Y6"].value,
                                ws["Z6"].value,
                                ws["AA6"].value,
                                round(
                                    ws[
                                        ("N" if daycount == 360 else "O") + str(crow)
                                    ].value
                                    * 100,
                                    2,
                                ),
                                round(r_a[0] * 100, 2),
                                round(r_b[0] * 100, 2),
                                round(r_c[0] * 100, 2),
                            )
                        )
                    ),
                )
            elif ws["C5"].value.upper() == "NOTE":  # note report template
                print("in note report template")
                script = note_report_template.format(
                    report_date=report_date.strftime("%B %d, %Y"),
                    fundname=ws["C9"].value,
                    series_abbrev=series_from_note(ws["C9"].value, ws["C12"].value),
                    issuer_name=issuer_from_fundname(ws["C9"].value),
                    frequency=ws["C15"].value,
                    rating=ws["C13"].value,
                    rating_org=declare_ratings_org(ws["C14"].value),
                    benchmark=ws["U" + str(crow + 1)].value,  # TODO ENSURE THERE
                    tgt_outperform=ws["C26"].value,
                    prev_pd_start=prev_pd_start.strftime("%b %d"),
                    prev_pd_return=form_as_percent(
                        ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                    ),
                    prev_pd_benchmark=benchmark_shorten(ws["U" + str(crow)].value),
                    prev_pd_outperform=bps_spread(
                        form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        form_as_percent(r_a[0], 2),
                    )[1:],
                    this_pd_end=ws["C24"].value.strftime(
                        "%b %d"
                    ),  # TODO ENSURE matches
                    this_pd_est_return=form_as_percent(
                        ws[("N" if daycount == 360 else "O") + str(crow + 1)].value, 2
                    ),
                    # TODO ensure there
                    this_pd_est_outperform=int(
                        10000 * ws["W" + str(crow + 1)].value
                    ),  # TODO ENSURE THERE
                    this_pd_start=this_pd_start.strftime("%b %d"),
                    benchmark_short=benchmark_shorten(ws["U" + str(crow + 1)].value),
                    interval1=month_wordify(interval_tuple[0]),
                    interval2=month_wordify(interval_tuple[1]),
                    return_table_plot=return_table_plot(
                        fund_name=ws["C9"].value,
                        prev_pd_return=form_as_percent(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value, 2
                        ),
                        series_abbrev=series_from_note(ws["C9"].value, ws["C12"].value),
                        r_this_1=r_this_1,
                        r_this_2=r_this_2,
                        comp_a=ws["Y6"].value,
                        comp_b=ws["Z6"].value,
                        comp_c=ws["AA6"].value,
                        r_a=r_a,
                        r_b=r_b,
                        r_c=r_c,
                        s_a_0=bps_spread(
                            form_as_percent(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            form_as_percent(r_a[0], 2),
                        ),  # TODO daycounts for spreads and round to 2 places
                        s_a_1=bps_spread(r_this_1, r_a[1]),
                        s_a_2=bps_spread(r_this_2, r_a[2]),
                        s_b_0=bps_spread(
                            form_as_percent(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            form_as_percent(r_b[0], 2),
                        ),
                        s_b_1=bps_spread(r_this_1, r_b[1]),
                        s_b_2=bps_spread(r_this_2, r_b[2]),
                        s_c_0=bps_spread(
                            form_as_percent(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            form_as_percent(r_c[0], 2),
                        ),
                        s_c_1=bps_spread(r_this_1, r_c[1]),
                        s_c_2=bps_spread(r_this_2, r_c[2]),
                    ),
                    colltable=colltable(
                        not (ws["C9"].value == "USG"),
                        secured_by_from(
                            ws["C9"].value,
                            series_from_note(ws["C9"].value, ws["C12"].value),
                        ),
                        series_from_note(ws["C9"].value, ws["C12"].value),
                        not (ws["C9"].value == "USG"),
                        form_as_percent(ws["AJ" + str(crow)].value, 1),
                        form_as_percent(ws["AK" + str(crow)].value, 1),
                        form_as_percent(ws["AL" + str(crow)].value, 1),
                        form_as_percent(ws["AM" + str(crow)].value, 1),
                        form_as_percent(ws["AO" + str(crow)].value, 1),
                        form_as_percent(1, 1),
                        form_as_percent(ws["AP" + str(crow)].value, 1),
                        "-",
                        form_as_percent(ws["AT" + str(crow)].value, 1),
                        form_as_percent(ws["AS" + str(crow)].value, 1),
                        form_as_percent(ws["AR" + str(crow)].value, 1),
                        form_as_percent(ws["AQ" + str(crow)].value, 1),
                    ),
                    zero_date=ws["F" + str(ts_row_start)].value.strftime("%Y-%m-%d"),
                    max_return=3,
                    return_plot=plotify(
                        ws,
                        "F",
                        "N" if daycount == 360 else "O",
                        ts_row_start,
                        ts_row_end,
                    ),
                    comp_a_plot=plotify(ws, "F", "Y", ts_row_start, ts_row_end),
                    comp_b_plot=plotify(ws, "F", "Z", ts_row_start, ts_row_end),
                    performance_graph=snapshot_graph(
                        -0.86,
                        "!",
                        "6.676cm",
                        0.5
                        + max(
                            round(
                                ws[("N" if daycount == 360 else "O") + str(crow)].value,
                                2,
                            ),
                            round(r_a[0] * 100, 2) if r_a[0] is not None else 0,
                            round(r_b[0] * 100, 2) if r_b[0] is not None else 0,
                            round(r_c[0] * 100, 2) if r_c[0] is not None else 0,
                        )
                        + 0.4,
                        ws["C11"].value,
                        ws["Y6"].value,
                        ws["Z6"].value,
                        ws["AA6"].value if ws["C9"].value != "USG" else None,
                        round(
                            ws[("N" if daycount == 360 else "O") + str(crow)].value
                            * 100,
                            2,
                        ),
                        round(r_a[0] * 100, 2) if r_a[0] is not None else 0,
                        round(r_b[0] * 100, 2) if r_b[0] is not None else 0,
                        round(r_c[0] * 100, 2) if r_c[0] is not None else 0,
                    ),
                    fund_size=wordify(program_size),
                    series_size=wordify(
                        ws["G" + str(crow + 1)].value
                    ),  # post sub/redemp, TODO temporarily consold hardwired
                    lucid_aum=wordify(lucid_aum),
                    fund_inception=(
                        ws["C18"].value.strftime("%B %d, %Y")
                        if fund_inception_from_name(ws["C9"].value) == "n/a"
                        else fund_inception_from_name(ws["C9"].value)
                    ),
                    cusip=ws["C16"].value,
                    note_abbrev=ws["C11"].value,
                    principal_outstanding=wordify(ws["R" + str(crow + 1)].value),
                    issue_date=ws["C18"].value.strftime("%B %d, %Y"),
                    maturity_date=ws["C20"].value.strftime("%B %d, %Y"),
                    pd_end_date_long=ws["C24"].value.strftime("%B %d, %Y"),
                    next_notice_date=ws["C22"].value.strftime("%B %d, %Y"),
                    coupon_plot=coupon_plotify(ws, crow, daycount),
                    rets_disclaimer_if_m1=(
                        m1_1230_128_rets_disclaimer
                        if ws["C12"].value == "Monthly1"
                        else ""
                    ),
                )
            else:
                continue
            # write script to file
            print("Generating Latex file...")
            filepath = report_name.replace(" ", "_")
            script_file = filepath + ".tex"
            with open(script_file, "w") as out:
                out.write(script)
                out.close()
            if "PrimeFund_C1" in filepath:
                # generate pdf
                print("Generating PDF...")
                # pdf_file = filepath + '.pdf'
                # Ensure the output directory exists
                output_directory = r"S:\Users\THoang\Data\MonthlyReports"  # Your desired output directory
                os.makedirs(output_directory, exist_ok=True)

                pdf_file = os.path.join(
                    output_directory, filepath + ".pdf"
                )  # Construct the full PDF path

                script_file = os.path.join(output_directory, script_file)

                # Check if the script file exists
                if not os.path.exists(script_file):
                    print(f"The script file {script_file} does not exist.")
                else:
                    pdf_file = os.path.join(
                        output_directory, filepath + ".pdf"
                    )  # Construct the full PDF path
                    cmd_str = f"pdflatex -interaction nonstopmode {script_file} -output-directory={output_directory}"
                    print(cmd_str)
                    try:
                        try:
                            x = subprocess.check_output(cmd_str)
                        except Exception as e:
                            print(
                                "Error generating file {} with error {}.".format(
                                    pdf_file, e
                                )
                            )
                            reports_generated.append(report_name)
                    except:
                        print("Error generating file {}.".format(pdf_file))
    except Exception as e:
        print("EXCEPTION")
        print(e)
        print("Error generating " + report_name)
        bad_reports.append(report_name)
        continue

# close workbook
wb.close()

if not bad_reports:
    print("All reports generated:")
else:
    print("Some reports generated:")

for x in reports_generated:
    print(x)
# TODO check page counts
if bad_reports:
    print("****ERROR generating following reports:")
    for x in bad_reports:
        print(x)
