import base64
import time
from datetime import datetime
from pathlib import PureWindowsPath, Path
import openpyxl as op
import pandas as pd
import requests
from sqlalchemy import text, Table, MetaData, Column, String, Date, DateTime
from sqlalchemy.exc import SQLAlchemyError

from Utils.Hash import hash_string
from Utils.database_utils import get_database_engine

# Assuming get_database_engine is already defined and returns a SQLAlchemy engine
engine = get_database_engine('sql_server_2')

def create_table_with_schema(tb_name):
    metadata = MetaData()
    metadata.bind = engine
    table = Table(tb_name, metadata,
                  Column("price_id", String(255), primary_key=True),
                  Column("price_date", Date),
                  Column("bond_id", String),
                  Column("price", String),
                  Column("timestamp", DateTime),
                  extend_existing=True)
    metadata.create_all(engine)
    print(f"Table {tb_name} created successfully or already exists.")


def upsert_data(tb_name, df):
    merge_sql = text("""
            MERGE INTO {} AS TARGET
            USING (
                SELECT :price_id AS price_id, :price_date AS price_date, :bond_id AS bond_id, :price AS price, :timestamp AS timestamp
                FROM (VALUES (1)) AS v(dummy)
            ) AS SOURCE
            ON TARGET.price_id = SOURCE.price_id
            WHEN MATCHED THEN
                UPDATE SET 
                    price_date = SOURCE.price_date, 
                    bond_id = SOURCE.bond_id, 
                    price = SOURCE.price, 
                    timestamp = SOURCE.timestamp
            WHEN NOT MATCHED THEN
                INSERT (price_id, price_date, bond_id, price, timestamp)
                VALUES (SOURCE.price_id, SOURCE.price_date, SOURCE.bond_id, SOURCE.price, SOURCE.timestamp);
        """.format(tb_name))

    # Executing the MERGE statement for each row in the DataFrame
    with engine.connect() as conn:
        try:
            with conn.begin():  # Start a transaction
                for idx, row in df.iterrows():
                    conn.execute(merge_sql, {
                        'price_id': row['price_id'],
                        'price_date': row['price_date'],
                        'bond_id': row['bond_id'],
                        'price': row['price'],
                        'timestamp': row['timestamp']
                    })
            print(
                f"Data for {df['price_date'][0]} upserted successfully into {tb_name}.")
        except SQLAlchemyError as e:
            print(f"An error occurred: {e}")
            raise

def fetch_idc(cusips, price_date):
    start_time = time.time()
    print("Fetching prices from IDC...")
    url = "https://rplus.intdata.com/cgi/nph-rplus"
    user = "d4lucid"
    password = "Spring17!"

    # Prepare the request
    auth = base64.b64encode(f"{user}:{password}".encode()).decode()
    headers = {"Authorization": f"Basic {auth}"}
    idc_req = f'GET,({" ,".join(cusips)}),(PRC),,,,TITLES=SHORT,DATEFORM=YMD'

    # Send the request
    response = requests.post(url, headers=headers, data={"Request": idc_req, "Done": "flag"})

    if response.status_code == 200:
        lines = response.text.strip().split("\n")
        data = []

        for line in lines[1:-1]:  # Skip the first line (header) and the last line (CRC)
            parts = line.split(",")
            cusip = parts[0].strip('"')  # Remove the quotes around the CUSIP
            price = parts[1]
            data.append([cusip, price])

        # Create a DataFrame with dates as columns and CUSIPs as rows
        df = pd.DataFrame(data, columns=["bond_id", "price"])
        df["price_date"] = price_date
        df["price_id"] = df.apply(
            lambda row: hash_string(f"{row['bond_id']}{'price_date'}"), axis=1)

        current_time = time.time()
        current_datetime = datetime.fromtimestamp(current_time)
        # Format datetime object to string in the desired format
        formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')
        # Assign formatted datetime to a new column in the DataFrame
        df["timestamp"] = formatted_datetime
        # Reorder the DataFrame columns to match the table column order
        df = df[["price_id", "price_date", "bond_id", "price", "timestamp"]]

        upsert_data(tb_name, df)
        print(f"Data uploaded exported to table")
        end_time = time.time()
        print(f"Time taken: {end_time - start_time:.2f} seconds")
    else:
        print(f"Error fetching data from IDC: {response.status_code}")
        return None

import logging
logging.basicConfig()
logging.getLogger('sqlalchemy.engine').setLevel(logging.INFO)


# Assuming df is your DataFrame after processing an unprocessed file
tb_name = "bronze_daily_price_idc"
create_table_with_schema(tb_name)

currdest = PureWindowsPath(Path("S:/Lucid/Data/Bond Data/Price Source/Price_Source.xlsx"))

try:
    wb = op.load_workbook(currdest)
except:
    print("file not found")
    exit()

if 'PM Prices' in wb.sheetnames:
    px_sheet = wb["PM Prices"]  # overwrite current PM page if exists
else:
    px_sheet = wb.copy_worksheet(wb["AM Prices"])  # make new (so copy AM) if doesn't

px_sheet.title = "PM Prices"
pxable_cusips = []

row = 2
curr = px_sheet.cell(row=row, column=1)
while (curr.value):
    pxable_cusips.append(curr.value)
    row = row + 1
    curr = px_sheet.cell(row=row, column=1)

# Get the current time in seconds since the epoch
current_time = time.time()
# Convert to a datetime object
date_time = datetime.fromtimestamp(current_time)
# Format the date as YYYYMMDD
query_date = date_time.strftime('%Y%m%d')

fetch_idc(pxable_cusips[:10], query_date)

print("Process completed.")
