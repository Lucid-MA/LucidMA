import os
import sys
from datetime import datetime, timedelta

import msal
import requests
import base64

# Get the absolute path of the current script
script_path = os.path.abspath(__file__)

# Get the directory of the script (Bronze_tables directory)
script_dir = os.path.dirname(script_path)

# Get the Reporting directory (parent of Bronze_tables)
reporting_dir = os.path.dirname(script_dir)

# Add the Reporting directory to sys.path so Python can find 'Utils' and 'Price'
sys.path.insert(0, reporting_dir)

import openpyxl
import pandas as pd
from sqlalchemy import text, Table, MetaData, Column, String, DateTime, Float, Date
from sqlalchemy.exc import SQLAlchemyError

# Import from Price.bloomberg_utils
from Price.bloomberg_utils import BloombergDataFetcher

# Import from Utils
from Utils.Common import get_file_path, get_current_timestamp, get_current_date
from Utils.Constants import (
    CP_1M,
    SOFR_1Y,
    SOFR_6M,
    SOFR_3M,
    CP_3M,
    CP_6M,
    CP_9M,
    SOFR_1M,
    LIBOR_1M,
    LIBOR_3M,
    TBILL_1M,
    TBILL_3M,
    EUR_FX,
    DGCXX,
)
from Utils.database_utils import get_database_engine

# Flag to enable publish to prod
PUBLISH_TO_PROD = True

# Flag to update database via excel file
MANUAL_REFRESH = True

# Assuming get_database_engine is already defined and returns a SQLAlchemy engine
if PUBLISH_TO_PROD:
    engine = get_database_engine("sql_server_2")
else:
    engine = get_database_engine("postgres")

tb_name = "bronze_daily_bloomberg_rates"

benchmark_file_path = get_file_path(r"S:/Lucid/Data/Historical Benchmarks.xlsx")

if MANUAL_REFRESH:
    import win32com.client as win32

    # Open the Excel file and refresh the data connection
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False  # Make Excel visible
    excel.DisplayAlerts = False  # Disable alerts

    workbook = excel.Workbooks.Open(benchmark_file_path)

    # Set calculation to automatic
    excel.Calculation = win32.constants.xlCalculationAutomatic

    # Refresh the specific sheet
    sheet = workbook.Sheets("bberg historical raw")
    sheet.Calculate()

    # Refresh the specific sheet
    sheet_2 = workbook.Sheets("dgcxx")
    sheet_2.Calculate()

    # Refresh all data connections
    workbook.RefreshAll()

    # Force a full recalculation
    excel.Calculate()

    # Save, close, and quit
    workbook.Save()
    workbook.Close()
    excel.Quit()

    # Release COM objects
    del sheet
    del sheet_2
    del workbook
    del excel


def authenticate_and_get_token():
    client_id = "10b66482-7a87-40ec-a409-4635277f3ed5"
    tenant_id = "86cd4a88-29b5-4f22-ab55-8d9b2c81f747"
    config = {
        "client_id": client_id,
        "authority": f"https://login.microsoftonline.com/{tenant_id}",
        "scope": ["https://graph.microsoft.com/Mail.Send"],
        "redirect_uri": "http://localhost:8080",
    }

    cache_file = "token_cache.bin"
    token_cache = msal.SerializableTokenCache()

    if os.path.exists(cache_file):
        with open(cache_file, "r") as f:
            token_cache.deserialize(f.read())

    client = msal.PublicClientApplication(
        config["client_id"], authority=config["authority"], token_cache=token_cache
    )

    accounts = client.get_accounts()
    if accounts:
        result = client.acquire_token_silent(config["scope"], account=accounts[0])
        if not result:
            print("No cached token found. Authenticating interactively...")
            result = client.acquire_token_interactive(scopes=config["scope"])
    else:
        print("No cached accounts found. Authenticating interactively...")
        result = client.acquire_token_interactive(scopes=config["scope"])

    if "error" in result:
        raise Exception(f"Error acquiring token: {result['error_description']}")

    with open(cache_file, "w") as f:
        f.write(token_cache.serialize())

    return result["access_token"]


def send_email(
    subject, body, recipients, cc_recipients, attachment_path=None, attachment_name=None
):
    token = authenticate_and_get_token()
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    email_data = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": body},
            "from": {"emailAddress": {"address": "operations@lucidma.com"}},
            "toRecipients": [
                {"emailAddress": {"address": recipient}} for recipient in recipients
            ],
            "ccRecipients": [
                {"emailAddress": {"address": cc_recipient}}
                for cc_recipient in cc_recipients
            ],
        }
    }

    if attachment_path and attachment_name:
        with open(attachment_path, "rb") as attachment:
            content_bytes = base64.b64encode(attachment.read()).decode("utf-8")

        email_data["message"]["attachments"] = [
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": attachment_name,
                "contentBytes": content_bytes,
            }
        ]

    response = requests.post(
        "https://graph.microsoft.com/v1.0/me/sendMail", headers=headers, json=email_data
    )
    if response.status_code != 202:
        raise Exception(f"Error sending email: {response.text}")
    else:
        print(f"Email '{subject}' sent successfully")


def create_table_with_schema(tb_name):
    metadata = MetaData()
    metadata.bind = engine
    table = Table(
        tb_name,
        metadata,
        Column("benchmark_date", String(255), primary_key=True),
        Column(CP_1M, Float, nullable=True),
        Column(CP_3M, Float, nullable=True),
        Column(CP_6M, Float, nullable=True),
        Column(CP_9M, Float, nullable=True),
        Column(SOFR_1M, Float, nullable=True),
        Column(SOFR_3M, Float, nullable=True),
        Column(SOFR_6M, Float, nullable=True),
        Column(SOFR_1Y, Float, nullable=True),
        Column(LIBOR_1M, Float, nullable=True),
        Column(LIBOR_3M, Float, nullable=True),
        Column(TBILL_1M, Float, nullable=True),
        Column(TBILL_1M + " Maturity", Date, nullable=True),
        Column(TBILL_3M, Float, nullable=True),
        Column(TBILL_3M + " Maturity", Date, nullable=True),
        Column(EUR_FX, Float, nullable=True),
        Column(DGCXX, Float, nullable=True),
        Column("timestamp", DateTime),
        extend_existing=True,
    )
    metadata.create_all(engine)
    print(f"Table {tb_name} created successfully or already exists.")


def upsert_data(tb_name, df):
    with engine.connect() as conn:
        try:
            with conn.begin():  # Start a transaction
                # Constructing the UPSERT SQL dynamically based on DataFrame columns
                column_names = ", ".join([f'"{col}"' for col in df.columns])

                value_placeholders = ", ".join(
                    [
                        f":{col.replace(' ', '_').replace('/', '_')}"
                        for col in df.columns
                    ]
                )
                # NOTE: THIS WORKS! For MS SQL, 'nan' data must be converted to None this way
                df = df.astype(object).where(pd.notnull(df), None)

                if PUBLISH_TO_PROD:
                    # Using MERGE statement for MS SQL Server
                    update_clause = ", ".join(
                        [
                            f'"{col}" = SOURCE."{col}"'
                            for col in df.columns
                            if col != "benchmark_date"
                        ]
                    )

                    upsert_sql = text(
                        f"""
                        MERGE INTO {tb_name} AS TARGET
                        USING (SELECT {','.join(f'SOURCE."{col}"' for col in df.columns)} FROM (VALUES ({value_placeholders})) AS SOURCE ({column_names})) AS SOURCE
                        ON TARGET."benchmark_date" = SOURCE."benchmark_date"
                        WHEN MATCHED THEN
                            UPDATE SET {update_clause}
                        WHEN NOT MATCHED THEN
                            INSERT ({column_names}) VALUES ({','.join(f'SOURCE."{col}"' for col in df.columns)});
                        """
                    )
                else:
                    update_clause = ", ".join(
                        [
                            f'"{col}"=EXCLUDED."{col}"'
                            for col in df.columns
                            if col
                            != "benchmark_date"  # Assuming "benchmark_date" is unique and used for conflict resolution
                        ]
                    )

                    upsert_sql = text(
                        f"""
                        INSERT INTO {tb_name} ({column_names})
                        VALUES ({value_placeholders})
                        ON CONFLICT ("benchmark_date")
                        DO UPDATE SET {update_clause};
                        """
                    )

                # Replace spaces and slashes with underscores in the DataFrame column names
                df.columns = [
                    col.replace(" ", "_").replace("/", "_") for col in df.columns
                ]

                # Execute upsert in a transaction
                conn.execute(upsert_sql, df.to_dict(orient="records"))
            print(f"Latest data upserted successfully into {tb_name}.")
        except SQLAlchemyError as e:
            print(f"An error occurred: {e}")
            raise


create_table_with_schema(tb_name)

if MANUAL_REFRESH:
    try:
        # Open the workbook and select the sheet
        wb = openpyxl.load_workbook(benchmark_file_path, read_only=True)
        sheet = wb["bberg historical raw"]

        # Find the last row with data
        last_row = sheet.max_row

        # Set the range of cells to read
        start_row = 12  # First row of data (after skipping rows)
        start_col = "D"  # Column letter for the start of the table
        end_col = "S"  # Column letter for the end of the table

        # Read the Excel file
        benchmark_df = pd.read_excel(
            benchmark_file_path,
            sheet_name="bberg historical raw",
            header=7,  # Header is on row 8 (index 7)
            usecols=f"{start_col}:{end_col}",
            skiprows=range(8, 11),  # Skip rows to start data from row 12
            nrows=last_row - start_row + 1,  # Explicitly specify number of rows to read
        )

        # Close the workbook
        wb.close()

        # Rename the columns
        new_column_names = [
            "benchmark_date",
            SOFR_1M,
            SOFR_3M,
            SOFR_6M,
            SOFR_1Y,
            LIBOR_1M,
            LIBOR_3M,
            CP_1M,
            CP_3M,
            CP_6M,
            CP_9M,
            TBILL_1M,
            TBILL_1M + " Maturity",
            TBILL_3M,
            TBILL_3M + " Maturity",
            EUR_FX,
        ]

        benchmark_df.columns = new_column_names

        # Convert the 'dates' column to 'YYYY-MM-DD' format
        benchmark_df["benchmark_date"] = (
            benchmark_df["benchmark_date"].dt.strftime("%Y-%m-%d").astype(str)
        )

        # Conver rates column to float
        for col in benchmark_df.columns[1:-1]:
            benchmark_df[col] = benchmark_df[col].apply(
                lambda x: x if pd.notna(x) and isinstance(x, (int, float)) else None
            )

        new_column_order = [
            "benchmark_date",
            CP_1M,
            CP_3M,
            CP_6M,
            CP_9M,
            SOFR_1M,
            SOFR_3M,
            SOFR_6M,
            SOFR_1Y,
            LIBOR_1M,
            LIBOR_3M,
            TBILL_1M,
            TBILL_1M + " Maturity",
            TBILL_3M,
            TBILL_3M + " Maturity",
            EUR_FX,
        ]

        benchmark_df = benchmark_df[new_column_order]

        # Convert specific columns to float, handling empty or 'nan' values
        columns_to_convert = [
            col
            for col in benchmark_df.columns
            if col
            not in [
                "benchmark_date",
                TBILL_1M + " Maturity",
                TBILL_3M + " Maturity",
            ]
        ]
        for col in columns_to_convert:
            benchmark_df[col] = pd.to_numeric(benchmark_df[col], errors="coerce")

        # DGCXX Index
        # Open the workbook and select the sheet
        wb = openpyxl.load_workbook(benchmark_file_path, read_only=True)
        sheet = wb["dgcxx"]

        # Find the last row with data
        last_row = sheet.max_row

        # Set the range of cells to read
        start_row = 11  # First row of data (after skipping rows)
        start_col = "D"  # Column letter for the start of the table
        end_col = "R"  # Column letter for the end of the table

        # Read the Excel file
        dgcxx_df = pd.read_excel(
            benchmark_file_path,
            sheet_name="dgcxx",
            usecols="G,I,J",
            skiprows=range(8, 10),  # Skip rows to start data from row 12
            nrows=last_row - start_row + 1,  # Explicitly specify number of rows to read
            names=["benchmark_date", "rate", "type"],  # Assign custom column headers
        )

        # Close the workbook
        wb.close()

        # Convert rate to numeric
        dgcxx_df["rate"] = pd.to_numeric(dgcxx_df["rate"], errors="coerce")
        dgcxx_df["benchmark_date"] = (
            dgcxx_df["benchmark_date"].dt.strftime("%Y-%m-%d").astype(str)
        )

        # Filter rows where benchmark_date is not null and type is 'Daily'
        filtered_df = dgcxx_df[
            (dgcxx_df["benchmark_date"].notna()) & (dgcxx_df["type"] == "Daily")
        ]

        # Sort the filtered DataFrame by benchmark_date in ascending order
        dgcxx_df = filtered_df.sort_values(by="benchmark_date", ascending=True)
        dgcxx_df = dgcxx_df.drop(columns=["type"]).rename(columns={"rate": DGCXX})
        benchmark_df = pd.merge(benchmark_df, dgcxx_df, on="benchmark_date", how="left")

        # Replace NaN values with None
        benchmark_df = benchmark_df.astype(object).where(pd.notnull(benchmark_df), None)
        benchmark_df["timestamp"] = get_current_timestamp()
        if benchmark_df is not None:
            upsert_data(tb_name, benchmark_df)
    except Exception as e:
        print("Failed to update the table manually. Error:", e)

else:
    try:
        securities = [
            "TSFR1M Index",
            "TSFR3M Index",
            "TSFR6M Index",
            "TSFR12M Index",
            "US0001M Index",
            "US0003M Index",
            "DCPA030Y Index",
            "DCPA090Y Index",
            "DCPA180Y Index",
            "DCPA270Y Index",
            "GBM Govt",
            "GB3 Govt",
            "EUR CURNCY",
            "DGCXX US Equity",
        ]

        fetcher = BloombergDataFetcher()
        security_attributes_df = fetcher.get_benchmark_security_attributes(
            securities, ["PX_LAST", "MATURITY", "PX_CLOSE_1D", "DVD_SH_LAST"]
        )
        security_attributes_df["benchmark_date"] = get_current_date()
        security_attributes_df["timestamp"] = get_current_timestamp()
        security_attributes_df["US0001M Index"] = None
        security_attributes_df["US0003M Index"] = None
        if security_attributes_df is not None:
            upsert_data(tb_name, security_attributes_df)

        # Send out successful email:
        current_date = datetime.now() - timedelta(days=0)
        valdate = current_date.strftime("%Y-%m-%d")
        subject = f"LRX - Daily Reference Data - {valdate}"

        recipients = [
            "tony.hoang@lucidma.com",
            # "amelia.thompson@lucidma.com",
            # "stephen.ng@lucidma.com",
            # "mattias.almers@lucidma.com",
            # "martin.stpierre@lucidma.com",
        ]
        cc_recipients = []

        # Format the styled DataFrame as an HTML table
        html_table = security_attributes_df.to_html(index=False, border=1, escape=False)

        html_content = f"""
                                <!DOCTYPE html>
                                <html lang="en">
                                <head>
                                    <meta charset="UTF-8">
                                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                    <style>
                                        table {{
                                            width: 100%;
                                            border-collapse: collapse;
                                        }}
                                        th, td {{
                                            border: 1px solid black;
                                            padding: 8px;
                                            text-align: center;
                                        }}
                                        th {{
                                            background-color: #f2f2f2;
                                        }}
                                        .header {{
                                            background-color: #d9edf7;
                                        }}
                                        .header span {{
                                            font-size: 24px;
                                            font-weight: bold;
                                        }}
                                        .subheader {{
                                            background-color: #dff0d8;
                                        }}
                                    </style>
                                </head>
                                <body>
                                    <table>
                                        <tr class="header">
                                            <td colspan="{len(security_attributes_df.columns)}"><span>Lucid Management and Capital Partners LP</span></td>
                                        </tr>
                                    </table>
                                    <table>
                                        {html_table}
                                    </table>
                                </body>
                                </html>
                                """

        send_email(
            subject,
            html_content,
            recipients,
            cc_recipients,
            # attachment_path,
            # attachment_name,
        )

    except Exception as e:
        subject = "Error obtaining daily reference data from Bloomberg"
        body = (
            f"Problem with refreshing latest reference data from the Bloomberg terminal. Might require login to establish active connection. \n"
            f"Error: {str(e)}"
        )
        recipients = [
            "tony.hoang@lucidma.com",
            # "amelia.thompson@lucidma.com",
            # "stephen.ng@lucidma.com",
        ]
        cc_recipients = []
        send_email(subject, body, recipients, cc_recipients)
