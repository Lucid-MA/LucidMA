import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from pandas._libs.tslibs.offsets import BDay
from sqlalchemy import create_engine, text
import pandas as pd

from Utils.database_utils import get_database_engine

def get_previous_business_day_filename(price_date):
    # Convert price_date to datetime if it's a string
    if isinstance(price_date, str):
        price_date = datetime.strptime(price_date, "%Y-%m-%d")

    # Get the previous business day
    previous_business_day = (price_date - BDay(1)).date()

    # Format filename
    filename = previous_business_day.strftime("Price_Source_%m_%d_%Y.xlsx")
    return filename


def replace_am_prices_with_pm_prices(current_excel, price_date):
    # Path to the archives folder
    archive_folder = "S:/Lucid/Data/Bond Data/Price Source/Archives"

    # Get filename for the previous business day
    prev_business_day_file = get_previous_business_day_filename(price_date)
    full_path = os.path.join(archive_folder, prev_business_day_file)

    if os.path.exists(full_path):
        # Load the workbook for the previous business day
        prev_day_book = load_workbook(full_path)
        if "PM Prices" in prev_day_book.sheetnames:
            pm_prices_df = pd.read_excel(full_path, sheet_name="PM Prices")

            # Insert a blank column between "Cusip/ISIN" and "IDC"
            pm_prices_df.insert(1, "", "")
            # Drop the "Unnamed: 1" column if it exists
            pm_prices_df = pm_prices_df.loc[:, ~pm_prices_df.columns.str.contains('^Unnamed')]

            # Check if the current Excel file exists
            if os.path.exists(current_excel):
                # Load the current workbook
                current_book = load_workbook(current_excel)
                if "AM Prices" in current_book.sheetnames:
                    # Remove the existing "AM Prices" sheet
                    std = current_book["AM Prices"]
                    current_book.remove(std)
                # Create a new "AM Prices" sheet to ensure it's fresh
                current_book.create_sheet("AM Prices")
                current_book.save(current_excel)
                current_book.close()
            else:
                # Create a new workbook
                current_book = Workbook()

                # Remove the default sheet
                default_sheet = current_book.active
                current_book.remove(default_sheet)

                # Create a new "AM Prices" sheet
                current_book.create_sheet("AM Prices")

                current_book.save(current_excel)
                current_book.close()

            # Write using pandas ExcelWriter with explicit handling
            with pd.ExcelWriter(current_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                pm_prices_df.to_excel(writer, sheet_name="AM Prices", index=False)

            print("Replaced AM Prices with PM Prices successfully.")
        else:
            print("No 'PM Prices' tab found in the file for the previous business day.")
    else:
        print("No file found for the previous business day.")

#
# def create_pm_prices_tab(price_date, excel_file):
#     engine = get_database_engine('sql_server_2')
#     query = text(f"""
#         SELECT
#             COALESCE(idc.bond_id, jppd.bond_id) AS "Cusip/ISIN",
#             COALESCE(idc.price, 'N/A') AS "IDC",
#             COALESCE(jppd.price, 'N/A') AS "Pricing Direct"
#         FROM
#             (SELECT bond_id, price FROM bronze_daily_price_idc WHERE price_date = :price_date) idc
#             FULL OUTER JOIN
#             (SELECT bond_id, price FROM bronze_daily_price_jppd WHERE price_date = :price_date) jppd
#             ON idc.bond_id = jppd.bond_id
#     """)
#
#     with engine.connect() as conn:
#         result = conn.execute(query, {'price_date': price_date})
#         data = result.fetchall()
#         df = pd.DataFrame(data, columns=result.keys())
#
#     if os.path.exists(excel_file):
#         book = load_workbook(excel_file)
#         mode = 'a'
#     else:
#         book = Workbook()
#         mode = 'w'
#
#     with pd.ExcelWriter(excel_file, engine='openpyxl', mode=mode) as writer:
#         if "PM Prices" in writer.book.sheetnames:
#             std = writer.book["PM Prices"]
#             writer.book.remove(std)
#         df.to_excel(writer, sheet_name="PM Prices", index=False)

def create_pm_prices_tab(price_date, excel_file):
    engine = get_database_engine('sql_server_2')
    query = text(f"""
        SELECT 
            COALESCE(idc.bond_id, jppd.bond_id) AS "Cusip/ISIN",
            '' AS "Blank", -- Add a blank column
            COALESCE(idc.price, 'N/A') AS "IDC",
            COALESCE(jppd.price, 'N/A') AS "Pricing Direct"
        FROM 
            (SELECT bond_id, price FROM bronze_daily_price_idc WHERE price_date = :price_date) idc
            FULL OUTER JOIN
            (SELECT bond_id, price FROM bronze_daily_price_jppd WHERE price_date = :price_date) jppd
            ON idc.bond_id = jppd.bond_id
    """)

    with engine.connect() as conn:
        result = conn.execute(query, {'price_date': price_date})
        data = result.fetchall()

        # Create a DataFrame with the specified column names
        df = pd.DataFrame(data, columns=["Cusip/ISIN", "", "IDC", "Pricing Direct"])

    if os.path.exists(excel_file):
        book = load_workbook(excel_file)
        mode = 'a'
    else:
        book = Workbook()
        mode = 'w'

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode=mode) as writer:
        if "PM Prices" in writer.book.sheetnames:
            std = writer.book["PM Prices"]
            writer.book.remove(std)
        df.to_excel(writer, sheet_name="PM Prices", index=False)


def create_helix_cusips_tab(excel_file):
    def generate_cusips():
        funds = ["Prime", "USG", "MMT", "LMCP Inv"]
        cusips_all = dict()
        for f in funds:
            cusips_all[f] = set()
            helix_pull_time = datetime.now()

        print("Fetching CUSIPs from Helix...")
        engine_1 = get_database_engine('sql_server_1')
        query = text(
            """
                select distinct
                case when tradepieces.company = 44 then 'USG Fund' when tradepieces.company = 45 then 'Prime Fund' when tradepieces.COMPANY = 46 then 'MMT IM Fund' when tradepieces.COMPANY = 48 then 'LMCP Inv Fund'  when tradepieces.COMPANY = 49 then 'LucidRepo' end Fund,
                ltrim(rtrim(Tradepieces.ISIN)) BondID
                from tradepieces 
                where (tradepieces.isvisible = 1 or tradepieces.company = 49)
                and tradepieces.company in (44,45,46,48,49)
            and ltrim(rtrim(Tradepieces.ISIN)) not in ('HEXZETA01','HEXZT----','HZLNT----','MCHY-----','MNTNCHRY1','OLIVEEUR-','OLIVEUSD-','OPPOR----','OPPORTUN1','PAAPLEUR-','PAAPLUSD-','PFIR-----','SSPRUCE--','STAPL----','STHAPPLE1','TREATY---','TREATYUS1','ALM2EUR--','ALM2USD--','ALMNDUSD1','ALMONDEUR','ALMONDUSD','ECYP-----','EELM-----','EWILLEUR-','EWILLUSD-')
                order by Fund ASC
            """
        )

        with engine_1.connect() as conn_1:
            result = conn_1.execute(query)
            data = result.fetchall()
            for row in data:
                if row[0] in ['USG Fund', 'Prime Fund', 'MMT IM Fund', 'LMCP Inv Fund', 'LucidRepo']:
                    k = row[0][:-5] if row[0] in ['USG Fund', 'Prime Fund'] else 'MMT'
                    cusips_all[k].add(row[1])

        for f in cusips_all.keys():
            to_add = set()
            for x in cusips_all[f]:
                if len(x) >= 3:
                    if x[:3] == "PNI":
                        to_add.add(x[3:len(x)])
            for extra_f in to_add:
                cusips_all[f].add(extra_f)

        return cusips_all

    cusips_data = generate_cusips()

    if os.path.exists(excel_file):
        book = load_workbook(excel_file)
        mode = 'a'
    else:
        book = Workbook()
        mode = 'w'

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode=mode) as writer:
        if "Helix Cusips" in writer.book.sheetnames:
            hcs = writer.book["Helix Cusips"]
            writer.book.remove(hcs)
        cusip_sheet = writer.book.create_sheet(title="Helix Cusips")
        col = 1
        for f in cusips_data.keys():
            row = 1
            cusip_sheet.cell(row=1, column=col, value=f)
            for c in cusips_data[f]:
                row += 1
                cusip_sheet.cell(row=row, column=col, value=c)
            col += 1


def main(excel_file):
    price_date = (datetime.now().date()).strftime('%Y-%m-%d')
    current_time = datetime.now().time()
    am_time = datetime.strptime("12:00:00", "%H:%M:%S").time()

    if current_time < am_time:
        print("Running AM tasks...")
        replace_am_prices_with_pm_prices(excel_file, price_date)
        create_helix_cusips_tab(excel_file)
    else:
        print("Running PM tasks...")
        create_pm_prices_tab(price_date, excel_file)
        create_helix_cusips_tab(excel_file)


if __name__ == "__main__":
    excel_file = "S:/Users/THoang/Data/Price Source/Price_Source.xlsx"
    main(excel_file)